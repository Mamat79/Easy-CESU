from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = "24515A"
TITLE_FILL = "173B40"
LIGHT_FILL = "E8EEF3"
LINE = "CBD5E1"


def safe_sheet_name(name: str) -> str:
    return re.sub(r"[\\/?*\[\]:]", " ", name).strip()[:31] or "Feuille"


def table_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    return cleaned[:240]


def auto_width(ws) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = 10
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 48))
        ws.column_dimensions[letter].width = width


def style_header(ws, row: int, col_count: int) -> None:
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for cell in ws[row][0:col_count]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 28


def style_sheet(ws) -> None:
    thin = Side(style="thin", color=LINE)
    border = Border(top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    auto_width(ws)


def write_table(ws, headers: list[str], rows: list[list[Any]], name: str) -> None:
    ws.append(headers)
    data_rows = rows or [["" for _ in headers]]
    for row in data_rows:
        ws.append(row)

    col_count = len(headers)
    row_count = len(data_rows) + 1
    style_header(ws, 1, col_count)
    style_sheet(ws)

    ref = f"A1:{get_column_letter(col_count)}{row_count}"
    table = Table(displayName=table_name(name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def as_rows(objects: list[dict[str, Any]], columns: list[tuple[str, str]]) -> tuple[list[str], list[list[Any]]]:
    headers = [header for _, header in columns]
    rows = [[row.get(key, "") for key, _ in columns] for row in objects]
    return headers, rows


def month_pivot_rows(client_month_rows: list[dict[str, Any]], year: int) -> list[dict[str, Any]]:
    by_client: dict[str, dict[str, Any]] = {}
    for row in client_month_rows:
        client = row.get("client", "")
        target = by_client.setdefault(client, {"client": client, "total_heures": 0.0, "total_net": 0.0})
        month = int(row.get("mois") or 0)
        if month:
            target[f"h_{month}"] = row.get("heures", "")
        target["total_heures"] += float(row.get("heures") or 0)
        target["total_net"] += float(row.get("montant_net") or 0)

    output = []
    for row in by_client.values():
        item = {"client": row["client"]}
        for month in range(1, 13):
            item[f"h_{month}"] = row.get(f"h_{month}", "")
        item["total_heures"] = round(float(row["total_heures"]), 2)
        item["total_net"] = round(float(row["total_net"]), 2)
        item["annee"] = year
        output.append(item)
    return sorted(output, key=lambda x: str(x["client"]).casefold())


def month_detail_rows(interventions: list[dict[str, Any]], month: int) -> list[dict[str, Any]]:
    rows = [row for row in interventions if row.get("mois") == month]
    return sorted(rows, key=lambda x: (str(x.get("client", "")).casefold(), str(x.get("date", ""))))


def set_number_format(ws, columns: str, number_format: str, start_row: int = 2) -> None:
    for column in columns.split(","):
        column = column.strip()
        if not column:
            continue
        for row in range(start_row, ws.max_row + 1):
            ws[f"{column}{row}"].number_format = number_format


def build_synthese(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Synthese"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Bilan activite {payload.get('year', '')}"
    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=18)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    totals = payload["summaries"]["totals"]
    rows = [
        ["Generation", payload.get("generated_at", "")],
        ["Clients", totals.get("nb_clients", 0)],
        ["Interventions", totals.get("nb_interventions", 0)],
        ["Heures", totals.get("heures", 0)],
        ["Total net", totals.get("montant_net", 0)],
        ["Total brut estime", totals.get("montant_brut", 0)],
    ]
    for row_index, row in enumerate(rows, start=3):
        ws.cell(row_index, 1, row[0])
        ws.cell(row_index, 2, row[1])
        ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=LIGHT_FILL)
        ws.cell(row_index, 1).font = Font(bold=True)
    ws["B6"].number_format = "0.00"
    ws["B7"].number_format = '#,##0.00 "€"'
    ws["B8"].number_format = '#,##0.00 "€"'

    headers = ["Mois", "Clients", "Interventions", "Heures", "Net"]
    for col, value in enumerate(headers, start=4):
        ws.cell(3, col, value)
    style_header(ws, 3, 8)

    monthly = payload["summaries"].get("monthly", [])
    for row_index, item in enumerate(monthly, start=4):
        values = [
            item.get("mois_libelle", ""),
            item.get("nb_clients", 0),
            item.get("nb_interventions", 0),
            item.get("heures", 0),
            item.get("montant_net", 0),
        ]
        for col, value in enumerate(values, start=4):
            ws.cell(row_index, col, value)
    if monthly:
        for row in range(4, 4 + len(monthly)):
            ws[f"G{row}"].number_format = "0.00"
            ws[f"H{row}"].number_format = '#,##0.00 "€"'
        chart = LineChart()
        chart.title = "Evolution mensuelle"
        chart.y_axis.title = "Montant net"
        chart.x_axis.title = "Mois"
        data = Reference(ws, min_col=8, min_row=3, max_row=3 + len(monthly))
        categories = Reference(ws, min_col=4, min_row=4, max_row=3 + len(monthly))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "A10")

    auto_width(ws)


def export_bilan_excel(payload: dict[str, Any], output: Path, input_path: Path | str) -> None:
    wb = Workbook()
    build_synthese(wb, payload)

    sheets = [
        (
            "Bilan mensuel",
            payload["summaries"].get("monthly", []),
            [
                ("annee", "Annee"),
                ("mois", "Mois"),
                ("mois_libelle", "Mois libelle"),
                ("nb_clients", "Clients"),
                ("nb_interventions", "Interventions"),
                ("heures", "Heures"),
                ("montant_net", "Montant net"),
                ("montant_brut", "Montant brut estime"),
            ],
            "BilanMensuel",
            {"F": "0.00", "G,H": '#,##0.00 "€"'},
        ),
        (
            "Bilan clients",
            payload["summaries"].get("clients", []),
            [
                ("client", "Client"),
                ("cesu", "Numero CESU"),
                ("adresse", "Adresse"),
                ("nb_mois_travailles", "Mois travailles"),
                ("nb_interventions", "Interventions"),
                ("heures", "Heures"),
                ("montant_net", "Montant net"),
                ("montant_brut", "Montant brut estime"),
                ("derniere_intervention", "Derniere intervention"),
            ],
            "BilanClients",
            {"F": "0.00", "G,H": '#,##0.00 "€"'},
        ),
        (
            "Interventions",
            payload.get("interventions", []),
            [
                ("annee", "Annee"),
                ("mois", "Mois"),
                ("mois_libelle", "Mois libelle"),
                ("client", "Client"),
                ("date", "Date"),
                ("duree_heures", "Duree heures"),
                ("salaire_net_horaire", "Salaire net horaire"),
                ("montant_net", "Montant net"),
                ("montant_brut", "Montant brut estime"),
                ("cesu", "Numero CESU"),
                ("adresse", "Adresse"),
                ("source_onglet", "Onglet source"),
            ],
            "Interventions",
            {"F": "0.00", "G,H,I": '#,##0.00 "€"'},
        ),
    ]

    for title, objects, columns, name, formats in sheets:
        ws = wb.create_sheet(title)
        headers, rows = as_rows(objects, columns)
        write_table(ws, headers, rows, name)
        for cols, fmt in formats.items():
            set_number_format(ws, cols, fmt)

    pivot_rows = month_pivot_rows(payload["summaries"].get("client_month", []), int(payload.get("year") or 0))
    ws = wb.create_sheet("Client par mois")
    headers, rows = as_rows(
        pivot_rows,
        [("client", "Client")]
        + [(f"h_{month}", label) for month, label in enumerate(["Janv h", "Fev h", "Mars h", "Avril h", "Mai h", "Juin h", "Juil h", "Aout h", "Sept h", "Oct h", "Nov h", "Dec h"], start=1)]
        + [("total_heures", "Total heures"), ("total_net", "Total net")],
    )
    write_table(ws, headers, rows, "ClientParMois")
    set_number_format(ws, ",".join(get_column_letter(col) for col in range(2, 15)), "0.00")
    set_number_format(ws, "O", '#,##0.00 "€"')

    for month in payload["summaries"].get("monthly", []):
        month_number = int(month.get("mois") or 0)
        detail_rows = month_detail_rows(payload.get("interventions", []), month_number)
        ws = wb.create_sheet(safe_sheet_name(f"{month_number:02d} {month.get('mois_libelle', '')}"))
        headers, rows = as_rows(
            detail_rows,
            [
                ("client", "Client"),
                ("date", "Date"),
                ("duree_heures", "Duree heures"),
                ("salaire_net_horaire", "Salaire net horaire"),
                ("montant_net", "Montant net"),
                ("montant_brut", "Montant brut estime"),
                ("cesu", "Numero CESU"),
                ("adresse", "Adresse"),
            ],
        )
        write_table(ws, headers, rows, f"Mois_{payload.get('year')}_{month_number:02d}")
        set_number_format(ws, "C", "0.00")
        set_number_format(ws, "D,E,F", '#,##0.00 "€"')

    ws = wb.create_sheet("Anomalies")
    anomalies = payload.get("anomalies", [])
    write_table(ws, ["Anomalie"], [[item] for item in anomalies] or [["Aucune anomalie detectee"]], "Anomalies")

    ws = wb.create_sheet("Parametres")
    rows = [
        ["Classeur suivi de paye", payload["sources"].get("suivi_paye", "")],
        ["Fichier clients", payload["sources"].get("fichier_clients", "")],
        ["Dossier notes", payload["sources"].get("notes_intervention_dir", "")],
        ["Salaire net horaire", payload["parameters"].get("salaire_net_horaire", "")],
        ["Coefficient brut", payload["parameters"].get("coefficient_brut", "")],
        ["Notes PDF activees", "Oui" if payload["parameters"].get("pdf_notes_enabled") else "Non"],
        ["Notes remplacees", "Oui" if payload["parameters"].get("replace_notes") else "Non"],
        ["Fichier donnees", str(input_path)],
        ["Generation", payload.get("generated_at", "")],
    ]
    write_table(ws, ["Parametre", "Valeur"], rows, "Parametres")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
