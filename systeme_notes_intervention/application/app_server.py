from __future__ import annotations

import json
import os
import re
import shutil
import socket
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import urllib.request
import webbrowser
from copy import deepcopy
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import load_workbook

try:
    from backup_restore import create_backup, extract_backup, extract_backup_from_transfer_kit, verify_backup
    from reminder_logic import (
        ANTICIPATION_UNITS,
        RECURRENCES,
        add_months,
        next_occurrence_after,
        occurrence_dates_until,
        parse_iso_date,
    )
except ImportError:  # pragma: no cover - used by packaged executable
    from application.backup_restore import create_backup, extract_backup, extract_backup_from_transfer_kit, verify_backup
    from application.reminder_logic import (
        ANTICIPATION_UNITS,
        RECURRENCES,
        add_months,
        next_occurrence_after,
        occurrence_dates_until,
        parse_iso_date,
    )

try:
    from excel_export import export_bilan_excel
except ImportError:  # pragma: no cover - used by packaged executable
    from application.excel_export import export_bilan_excel

try:
    from email_service import (
        DEFAULT_EMAIL_BODY,
        DEFAULT_EMAIL_SUBJECT,
        EMAIL_TEMPLATE_FIELDS,
        build_email_message,
        delete_smtp_password,
        get_smtp_password,
        normalize_smtp_settings,
        password_saved,
        render_email_template,
        save_smtp_password,
        smtp_connection,
        smtp_defaults,
    )
except ImportError:  # pragma: no cover - used by packaged executable
    from application.email_service import (
        DEFAULT_EMAIL_BODY,
        DEFAULT_EMAIL_SUBJECT,
        EMAIL_TEMPLATE_FIELDS,
        build_email_message,
        delete_smtp_password,
        get_smtp_password,
        normalize_smtp_settings,
        password_saved,
        render_email_template,
        save_smtp_password,
        smtp_connection,
        smtp_defaults,
    )

if getattr(sys, "frozen", False):
    ROOT_DIR = Path(sys.executable).resolve().parent
    RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", ROOT_DIR))
    APP_DIR = RESOURCE_DIR / "application"
else:
    APP_DIR = Path(__file__).resolve().parent
    ROOT_DIR = APP_DIR.parent
    RESOURCE_DIR = ROOT_DIR

STATIC_DIR = APP_DIR / "static"
APP_NAME = "Easy CESU"
APP_VERSION = "3.1.4"
V2_SCHEMA_VERSION = 2
V3_SCHEMA_VERSION = 3
V4_SCHEMA_VERSION = 4
V6_SCHEMA_VERSION = 6
DATABASE_SCHEMA_VERSION = V6_SCHEMA_VERSION
BROWSER_CLOSE_GRACE_SECONDS = 5.0
BROWSER_STALE_SECONDS = 90.0
BROWSER_SESSION_LOCK = threading.Lock()
BROWSER_SESSIONS: dict[str, float] = {}
BROWSER_STREAM_TOKENS: dict[str, object] = {}
BROWSER_CONNECTED_ONCE = False
BROWSER_EMPTY_SINCE: float | None = None
SUPPORT_REMINDER_INITIAL_DAYS = 30
SUPPORT_REMINDER_REPEAT_DAYS = 90
SUPPORT_LINKS = {
    "github_repository": "https://github.com/Mamat79/Easy-CESU",
    "github_star": "https://github.com/Mamat79/Easy-CESU",
    "github_issues": "https://github.com/Mamat79/Easy-CESU/issues/new",
    "paypal_me": "https://www.paypal.com/paypalme/MamatLeroy",
}


def user_data_root() -> Path:
    override = str(os.environ.get("EASY_CESU_DATA_ROOT") or "").strip()
    if override:
        return Path(override).expanduser()
    if sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support"
    elif os.name == "nt":
        base = Path(os.environ.get("LOCALAPPDATA") or Path.home() / "AppData" / "Local")
    else:
        base = Path(os.environ.get("XDG_DATA_HOME") or Path.home() / ".local" / "share")
    return Path(base) / "EasyCESU"


def legacy_user_data_root() -> Path:
    if sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support"
    elif os.name == "nt":
        base = Path(os.environ.get("LOCALAPPDATA") or Path.home() / "AppData" / "Local")
    else:
        base = Path(os.environ.get("XDG_DATA_HOME") or Path.home() / ".local" / "share")
    return Path(base) / APP_NAME


RUNTIME_DIR = user_data_root() if getattr(sys, "frozen", False) else ROOT_DIR
DATA_DIR = RUNTIME_DIR / "data" if getattr(sys, "frozen", False) else APP_DIR / "data"
ATTACHMENTS_DIR = RUNTIME_DIR / "attachments" if getattr(sys, "frozen", False) else ROOT_DIR / "attachments"
BACKUPS_DIR = RUNTIME_DIR / "backups" if getattr(sys, "frozen", False) else ROOT_DIR / "backups"
LOGS_DIR = RUNTIME_DIR / "logs" if getattr(sys, "frozen", False) else ROOT_DIR / "logs"
TEMP_DIR = RUNTIME_DIR / "temp" if getattr(sys, "frozen", False) else ROOT_DIR / "temp"
LEGACY_DB_FILE = "interventions.sqlite"
DB_PATH = DATA_DIR / LEGACY_DB_FILE
CONFIG_PATH = RUNTIME_DIR / "config" / "config.json" if getattr(sys, "frozen", False) else ROOT_DIR / "config.json"
DEFAULT_CONFIG_PATH = RESOURCE_DIR / "config.json"
CONFIG_CREATED_THIS_RUN = not CONFIG_PATH.exists()


def default_notes_dir() -> Path:
    return RUNTIME_DIR / "notes-intervention"


def default_export_dir() -> Path:
    return RUNTIME_DIR / "exports"

sys.path.insert(0, str(RESOURCE_DIR))
sys.path.insert(0, str(ROOT_DIR))
from generer_notes_et_donnees import (  # noqa: E402
    EMPLOYEE_LINES,
    Intervention,
    MONTH_LABELS,
    build_summaries,
    default_note_template_configuration,
    generate_note_pdf,
    generate_notes,
    french_money,
    hours_label,
    matching_existing_note,
    money,
    normalize_note_template_configuration,
    normalize_name,
    register_fonts,
    safe_filename,
    sorted_name_key,
)


DEFAULT_CONFIG = {
    "suivi_paye_dir": "",
    "notes_intervention_dir": str(default_notes_dir()),
    "export_dir": str(default_export_dir()),
    "fichier_clients": "",
    "suivi_paye_pattern": "Suivi de paye {year}.xlsx",
    "salaire_net_horaire_defaut": 22.0,
    "coefficient_brut_defaut": 1.2873125,
    "ecraser_notes_existantes": False,
    "initial_setup_completed": False,
    "support_reminder_enabled": True,
    "support_reminder_started_on": "",
    "support_reminder_last_seen_on": "",
}

PROFILE_METADATA_KEY = "active_profile"
PROFILE_METADATA_FIELDS = (
    "label",
    "name",
    "address",
    "phone",
    "email",
    "ss_number",
    "birth_info",
    "default_hourly_rate",
    "suivi_paye_dir",
    "fichier_clients",
    "suivi_paye_pattern",
    "first_name",
    "last_name",
    "commercial_name",
    "postal_code",
    "city",
    "siret",
    "primary_activity",
    "secondary_activities",
    "theme",
    "shortcut_icon",
    "backup_dir",
    "backup_retention_days",
    "daily_backup_enabled",
    "smtp_host",
    "smtp_port",
    "smtp_security",
    "smtp_username",
    "smtp_sender_name",
    "smtp_sender_email",
    "email_subject_template",
    "email_body_template",
)

ACTIVITIES = (
    "jardinage",
    "bricolage",
    "menage",
    "aide_a_domicile",
    "garde_d_enfants",
    "soutien_scolaire",
    "accompagnement",
    "assistance_administrative",
    "informatique",
    "autre",
)
SHORTCUT_ICONS = {
    "generique",
    "jardinage",
    "bricolage",
    "menage",
    "aide_a_domicile",
    "garde_d_enfants",
    "soutien_scolaire",
    "accompagnement",
    "assistance_administrative",
    "informatique",
}
NOTE_CATEGORIES = {"information", "compte_rendu", "fait", "a_faire", "client", "paiement", "materiel", "incident", "prive", "autre"}
NOTE_STATUSES = {"information", "a_faire", "en_attente", "termine", "annule"}
NOTE_PRIORITIES = {"basse", "normale", "haute"}
PAYMENT_STATUSES = {"a_recevoir", "partiellement_recu", "recu", "annule", "litige"}
INTERVENTION_STATUSES = {"prevue", "confirmee", "realisee", "annulee", "reportee", "a_declarer", "payee", "paiement_en_attente"}
ADMINISTRATIVE_STATUS_FIELDS = {"transmitted", "declared", "paid"}


def ensure_runtime_files() -> None:
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not CONFIG_PATH.exists():
        # La mise à jour ne déplace jamais les données historiques : elle les copie
        # vers le nouvel espace utilisateur stable, hors du dossier d'installation.
        legacy_root = legacy_user_data_root()
        legacy_config = legacy_root / "config.json"
        legacy_data = legacy_root / "application" / "data"
        legacy_import_disabled = os.environ.get("EASY_CESU_DISABLE_LEGACY_IMPORT") == "1"
        if getattr(sys, "frozen", False) and legacy_config.exists() and not legacy_import_disabled:
            shutil.copy2(legacy_config, CONFIG_PATH)
            if legacy_data.exists() and not DATA_DIR.exists():
                shutil.copytree(legacy_data, DATA_DIR)
        elif DEFAULT_CONFIG_PATH.exists():
            shutil.copy2(DEFAULT_CONFIG_PATH, CONFIG_PATH)
        else:
            CONFIG_PATH.write_text(json.dumps(DEFAULT_CONFIG, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    for folder in (DATA_DIR, ATTACHMENTS_DIR, BACKUPS_DIR, LOGS_DIR, TEMP_DIR):
        folder.mkdir(parents=True, exist_ok=True)


ensure_runtime_files()


def load_config() -> dict:
    with CONFIG_PATH.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def save_config() -> None:
    CONFIG_PATH.write_text(json.dumps(CONFIG, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


CONFIG = load_config()
DEFAULT_RATE = float(CONFIG.get("salaire_net_horaire_defaut", 22.0))
DEFAULT_BRUT_COEFF = float(CONFIG.get("coefficient_brut_defaut", 1.2873125))


def profile_id_from_label(label: str) -> str:
    base = "-".join(normalize_name(label).split()) or "compte"
    existing = {profile.get("id") for profile in CONFIG.get("profiles", [])}
    candidate = base
    index = 2
    while candidate in existing:
        candidate = f"{base}-{index}"
        index += 1
    return candidate


def default_profile() -> dict:
    profile = {
        "id": "mon-compte",
        "label": "Mon compte",
        "name": "",
        "address": "",
        "phone": "",
        "email": "",
        "ss_number": "",
        "birth_info": "",
        "default_hourly_rate": float(CONFIG.get("salaire_net_horaire_defaut", 22.0)),
        "notes_intervention_dir": str(default_notes_dir() / "mon-compte"),
        "export_dir": str(default_export_dir() / "mon-compte"),
        "suivi_paye_dir": "",
        "fichier_clients": "",
        "suivi_paye_pattern": "Suivi de paye {year}.xlsx",
        "data_file": "profiles/mon-compte/interventions.sqlite",
        "seed_from_sorties": False,
        "primary_activity": "autre",
        "secondary_activities": "",
        "backup_dir": str(BACKUPS_DIR / "mon-compte"),
        "backup_retention_days": 30,
        "daily_backup_enabled": True,
    }
    profile.update(smtp_defaults())
    return profile


def normalize_profile(profile: dict) -> dict:
    if not profile.get("id"):
        profile["id"] = profile_id_from_label(str(profile.get("label") or profile.get("name") or "compte"))
    profile.setdefault("label", profile.get("name") or profile["id"])
    profile.setdefault("name", "")
    profile.setdefault("address", "")
    profile.setdefault("phone", "")
    profile.setdefault("email", "")
    profile.setdefault("ss_number", "")
    profile.setdefault("birth_info", "")
    profile.setdefault("default_hourly_rate", float(CONFIG.get("salaire_net_horaire_defaut", 22.0)))
    profile.setdefault("notes_intervention_dir", CONFIG.get("notes_intervention_dir", str(default_notes_dir())))
    profile.setdefault("export_dir", CONFIG.get("export_dir", str(default_export_dir())))
    profile.setdefault("suivi_paye_dir", "")
    profile.setdefault("fichier_clients", "")
    profile.setdefault("suivi_paye_pattern", "Suivi de paye {year}.xlsx")
    profile.setdefault("data_file", f"profiles/{profile['id']}/interventions.sqlite")
    profile.setdefault("seed_from_sorties", False)
    profile.setdefault("first_name", "")
    profile.setdefault("last_name", "")
    profile.setdefault("commercial_name", "")
    profile.setdefault("postal_code", "")
    profile.setdefault("city", "")
    profile.setdefault("siret", "")
    profile.setdefault("primary_activity", "autre")
    profile.setdefault("secondary_activities", "")
    profile.setdefault("theme", "")
    profile.setdefault("shortcut_icon", "generique")
    profile.setdefault("backup_dir", str(BACKUPS_DIR / profile["id"]))
    profile.setdefault("backup_retention_days", 30)
    profile.setdefault("daily_backup_enabled", True)
    for field, value in smtp_defaults().items():
        profile.setdefault(field, value)
    return profile


def ensure_profiles_config() -> None:
    changed = False
    if "support_reminder_enabled" not in CONFIG:
        CONFIG["support_reminder_enabled"] = True
        changed = True
    if not str(CONFIG.get("support_reminder_started_on") or "").strip():
        CONFIG["support_reminder_started_on"] = date.today().isoformat()
        changed = True
    if "support_reminder_last_seen_on" not in CONFIG:
        CONFIG["support_reminder_last_seen_on"] = ""
        changed = True
    if "initial_setup_completed" not in CONFIG:
        # Une configuration existante ne doit pas relancer l'assistant apres une mise a jour.
        CONFIG["initial_setup_completed"] = not CONFIG_CREATED_THIS_RUN
        changed = True
    if not CONFIG.get("profiles"):
        CONFIG["profiles"] = [default_profile()]
        CONFIG["active_profile_id"] = CONFIG["profiles"][0]["id"]
        changed = True
    for profile in CONFIG["profiles"]:
        before = dict(profile)
        normalize_profile(profile)
        changed = changed or before != profile
    if not CONFIG.get("active_profile_id") or not any(
        profile.get("id") == CONFIG.get("active_profile_id") for profile in CONFIG["profiles"]
    ):
        CONFIG["active_profile_id"] = CONFIG["profiles"][0]["id"]
        changed = True
    if changed:
        save_config()


def active_profile() -> dict:
    ensure_profiles_config()
    active_id = CONFIG.get("active_profile_id")
    for profile in CONFIG["profiles"]:
        if profile.get("id") == active_id:
            return profile
    return CONFIG["profiles"][0]


def public_profile(profile: dict) -> dict:
    result = {
        "id": profile["id"],
        "label": profile.get("label", ""),
        "name": profile.get("name", ""),
        "address": profile.get("address", ""),
        "phone": profile.get("phone", ""),
        "email": profile.get("email", ""),
        "ss_number": profile.get("ss_number", ""),
        "birth_info": profile.get("birth_info", ""),
        "default_hourly_rate": float(profile.get("default_hourly_rate") or 22.0),
        "notes_intervention_dir": profile.get("notes_intervention_dir", ""),
        "export_dir": str(export_output_dir(profile)),
        "data_dir": str(profile_data_dir(profile)),
        "database_path": str(profile_db_path(profile)),
        "suivi_paye_dir": profile.get("suivi_paye_dir", ""),
        "fichier_clients": profile.get("fichier_clients", ""),
        "suivi_paye_pattern": profile.get("suivi_paye_pattern", "Suivi de paye {year}.xlsx"),
        "first_name": profile.get("first_name", ""),
        "last_name": profile.get("last_name", ""),
        "commercial_name": profile.get("commercial_name", ""),
        "postal_code": profile.get("postal_code", ""),
        "city": profile.get("city", ""),
        "siret": profile.get("siret", ""),
        "primary_activity": profile.get("primary_activity", "autre"),
        "secondary_activities": profile.get("secondary_activities", ""),
        "theme": profile.get("theme", ""),
        "shortcut_icon": profile.get("shortcut_icon", "generique"),
        "backup_dir": profile.get("backup_dir", str(BACKUPS_DIR / profile["id"])),
        "backup_retention_days": int(profile.get("backup_retention_days") or 30),
        "daily_backup_enabled": bool(profile.get("daily_backup_enabled", True)),
    }
    result.update(normalize_smtp_settings(profile))
    result["smtp_password_saved"] = password_saved(profile["id"])
    result["email_template_fields"] = sorted(EMAIL_TEMPLATE_FIELDS)
    return result


def sync_active_profile_runtime() -> None:
    global DEFAULT_RATE
    DEFAULT_RATE = float(active_profile().get("default_hourly_rate") or 22.0)


def profile_data_file_name(profile: dict) -> str:
    raw = str(profile.get("data_file") or LEGACY_DB_FILE).replace("\\", "/")
    name = Path(raw).name
    return name or LEGACY_DB_FILE


def profile_default_data_dir(profile: dict) -> Path:
    raw = Path(str(profile.get("data_file") or LEGACY_DB_FILE).replace("\\", "/"))
    if raw.is_absolute():
        return raw.parent
    if raw.parent == Path("."):
        return DATA_DIR
    return DATA_DIR / raw.parent


def profile_data_dir(profile: dict) -> Path:
    value = clean_text(profile.get("data_dir")) if "data_dir" in profile else ""
    if value:
        return Path(value).expanduser()
    return profile_default_data_dir(profile)


def profile_db_path(profile: dict) -> Path:
    raw = Path(str(profile.get("data_file") or LEGACY_DB_FILE).replace("\\", "/"))
    if raw.is_absolute() and not clean_text(profile.get("data_dir")):
        return raw.expanduser()
    return profile_data_dir(profile) / profile_data_file_name(profile)


def active_db_path() -> Path:
    return profile_db_path(active_profile())


def active_employee_lines() -> list[str]:
    profile = active_profile()
    lines = []
    if profile.get("name"):
        lines.append(str(profile["name"]).strip())
    lines.extend(line.strip() for line in str(profile.get("address", "")).splitlines() if line.strip())
    if profile.get("phone"):
        lines.append(f"Tel. : {str(profile['phone']).strip()}")
    if profile.get("email"):
        lines.append(f"Mel : {str(profile['email']).strip()}")
    identity_lines = []
    if profile.get("ss_number"):
        identity_lines.append(f"N° SS : {str(profile['ss_number']).strip()}")
    if profile.get("birth_info"):
        identity_lines.append(str(profile["birth_info"]).strip())
    if identity_lines:
        lines.append("")
        lines.extend(identity_lines)
    return lines or EMPLOYEE_LINES


ensure_profiles_config()
sync_active_profile_runtime()


def connect() -> sqlite3.Connection:
    db_path = active_db_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(db_path)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON")
    return db


@contextmanager
def db_connection() -> sqlite3.Connection:
    db = connect()
    try:
        yield db
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def table_columns(db: sqlite3.Connection, table: str) -> set[str]:
    return {str(row["name"]) for row in db.execute(f"PRAGMA table_info({table})").fetchall()}


def add_missing_columns(db: sqlite3.Connection, table: str, definitions: dict[str, str]) -> None:
    """Ajoute les champs V2 sans toucher aux données historiques."""

    existing = table_columns(db, table)
    for name, definition in definitions.items():
        if name not in existing:
            db.execute(f"ALTER TABLE {table} ADD COLUMN {name} {definition}")


def database_contains_user_data(db: sqlite3.Connection) -> bool:
    for table in ("clients", "interventions", "reminders"):
        try:
            if db.execute(f"SELECT 1 FROM {table} LIMIT 1").fetchone():
                return True
        except sqlite3.Error:
            continue
    return False


def apply_v2_schema_migrations() -> None:
    """Applique les migrations V2 sur une base existante après sauvegarde contrôlée."""

    with db_connection() as db:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS schema_migrations (
                version INTEGER PRIMARY KEY,
                applied_at TEXT NOT NULL,
                details TEXT NOT NULL DEFAULT ''
            )
            """
        )
        applied = {int(row["version"]) for row in db.execute("SELECT version FROM schema_migrations").fetchall()}
        needs_v2 = V2_SCHEMA_VERSION not in applied
        has_user_data = database_contains_user_data(db)

    if not needs_v2:
        return

    # Une base utilisée est sauvegardée avant toute évolution irréversible de structure.
    if has_user_data:
        backup_profile_to(BACKUPS_DIR, "avant-migration-v2")

    with db_connection() as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS service_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL COLLATE NOCASE UNIQUE,
                activity TEXT NOT NULL DEFAULT '',
                icon_key TEXT NOT NULL DEFAULT '',
                color TEXT NOT NULL DEFAULT '',
                default_hourly_rate REAL NOT NULL DEFAULT 0,
                default_duration_hours REAL NOT NULL DEFAULT 0,
                is_archived INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS intervention_services (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                intervention_id INTEGER NOT NULL,
                category_id INTEGER,
                label TEXT NOT NULL DEFAULT '',
                duration_hours REAL NOT NULL DEFAULT 0,
                hourly_rate REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(intervention_id) REFERENCES interventions(id) ON DELETE CASCADE,
                FOREIGN KEY(category_id) REFERENCES service_categories(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS intervention_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                intervention_id INTEGER,
                client_name TEXT NOT NULL DEFAULT '',
                body TEXT NOT NULL,
                category TEXT NOT NULL DEFAULT 'information',
                priority TEXT NOT NULL DEFAULT 'normal',
                status TEXT NOT NULL DEFAULT 'information',
                reminder_date TEXT NOT NULL DEFAULT '',
                carry_forward INTEGER NOT NULL DEFAULT 0,
                is_private INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(intervention_id) REFERENCES interventions(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS pending_payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                intervention_id INTEGER,
                client_name TEXT NOT NULL,
                expected_amount REAL NOT NULL DEFAULT 0,
                received_amount REAL NOT NULL DEFAULT 0,
                expected_date TEXT NOT NULL DEFAULT '',
                payment_method TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'to_receive',
                comment TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(intervention_id) REFERENCES interventions(id) ON DELETE SET NULL
            );

            CREATE INDEX IF NOT EXISTS idx_notes_client ON intervention_notes(client_name);
            CREATE INDEX IF NOT EXISTS idx_notes_intervention ON intervention_notes(intervention_id);
            CREATE INDEX IF NOT EXISTS idx_notes_status ON intervention_notes(status);
            CREATE INDEX IF NOT EXISTS idx_payments_client ON pending_payments(client_name);
            CREATE INDEX IF NOT EXISTS idx_payments_status ON pending_payments(status);
            """
        )
        add_missing_columns(
            db,
            "clients",
            {
                "activity": "TEXT NOT NULL DEFAULT ''",
                "instructions": "TEXT NOT NULL DEFAULT ''",
                "access_info": "TEXT NOT NULL DEFAULT ''",
                "payment_preferences": "TEXT NOT NULL DEFAULT ''",
                "usual_duration_hours": "REAL NOT NULL DEFAULT 0",
                "usual_frequency": "TEXT NOT NULL DEFAULT ''",
                "preferred_days": "TEXT NOT NULL DEFAULT ''",
                "is_archived": "INTEGER NOT NULL DEFAULT 0",
                "created_at": "TEXT NOT NULL DEFAULT ''",
            },
        )
        add_missing_columns(
            db,
            "interventions",
            {
                "planned_start": "TEXT NOT NULL DEFAULT ''",
                "planned_end": "TEXT NOT NULL DEFAULT ''",
                "actual_start": "TEXT NOT NULL DEFAULT ''",
                "actual_end": "TEXT NOT NULL DEFAULT ''",
                "break_minutes": "INTEGER NOT NULL DEFAULT 0",
                "travel_minutes": "INTEGER NOT NULL DEFAULT 0",
                "status": "TEXT NOT NULL DEFAULT 'realized'",
                "category_id": "INTEGER",
                "planned_amount": "REAL NOT NULL DEFAULT 0",
                "received_amount": "REAL NOT NULL DEFAULT 0",
                "payment_status": "TEXT NOT NULL DEFAULT ''",
            },
        )
        db.execute("UPDATE clients SET created_at = updated_at WHERE created_at = ''")
        db.execute("UPDATE interventions SET planned_amount = duration_hours * hourly_rate WHERE planned_amount = 0")
        db.execute("UPDATE interventions SET received_amount = duration_hours * hourly_rate WHERE paid = 1 AND received_amount = 0")
        db.execute("UPDATE interventions SET payment_status = CASE WHEN paid = 1 THEN 'received' ELSE 'to_receive' END WHERE payment_status = ''")
        db.execute(
            "INSERT OR IGNORE INTO schema_migrations (version, applied_at, details) VALUES (?, ?, ?)",
            (V2_SCHEMA_VERSION, now_stamp(), "Fondations universelles V2"),
        )


def apply_v3_schema_migrations() -> None:
    """Ajoute les modèles de documents sans modifier les données métier."""

    with db_connection() as db:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS schema_migrations (
                version INTEGER PRIMARY KEY,
                applied_at TEXT NOT NULL,
                details TEXT NOT NULL DEFAULT ''
            )
            """
        )
        applied = {int(row["version"]) for row in db.execute("SELECT version FROM schema_migrations").fetchall()}
        needs_v3 = V3_SCHEMA_VERSION not in applied
        has_user_data = database_contains_user_data(db)

    if not needs_v3:
        return
    if has_user_data:
        backup_profile_to(BACKUPS_DIR, "avant-migration-v3")

    with db_connection() as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS document_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL COLLATE NOCASE UNIQUE,
                document_type TEXT NOT NULL DEFAULT 'intervention_note',
                is_default INTEGER NOT NULL DEFAULT 0,
                configuration_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_document_templates_type
            ON document_templates(document_type, is_default);
            """
        )
        db.execute(
            "INSERT OR IGNORE INTO schema_migrations (version, applied_at, details) VALUES (?, ?, ?)",
            (V3_SCHEMA_VERSION, now_stamp(), "Fenêtre native et modèles de notes V3"),
        )


def apply_v4_schema_migrations() -> None:
    """Ajoute la préférence d'envoi des notes sans modifier les fiches existantes."""

    with db_connection() as db:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS schema_migrations (
                version INTEGER PRIMARY KEY,
                applied_at TEXT NOT NULL,
                details TEXT NOT NULL DEFAULT ''
            )
            """
        )
        applied = {int(row["version"]) for row in db.execute("SELECT version FROM schema_migrations").fetchall()}
        needs_v4 = V4_SCHEMA_VERSION not in applied
        has_user_data = database_contains_user_data(db)
    if not needs_v4:
        return
    if has_user_data:
        backup_profile_to(BACKUPS_DIR, "avant-migration-v4")
    with db_connection() as db:
        add_missing_columns(
            db,
            "clients",
            {
                "email_notes_enabled": "INTEGER NOT NULL DEFAULT 0",
                "email_review_before_send": "INTEGER NOT NULL DEFAULT 0",
            },
        )
        db.execute(
            "INSERT OR IGNORE INTO schema_migrations (version, applied_at, details) VALUES (?, ?, ?)",
            (V4_SCHEMA_VERSION, now_stamp(), "Envoi sélectionné des notes par email"),
        )


def apply_v6_schema_migrations() -> None:
    """Ajoute le suivi administratif sans transformer l'historique en retard."""

    with db_connection() as db:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS schema_migrations (
                version INTEGER PRIMARY KEY,
                applied_at TEXT NOT NULL,
                details TEXT NOT NULL DEFAULT ''
            )
            """
        )
        applied = {int(row["version"]) for row in db.execute("SELECT version FROM schema_migrations").fetchall()}
        has_user_data = database_contains_user_data(db)
        had_declared_column = "declared" in table_columns(db, "interventions")
        has_followup_table = bool(
            db.execute(
                "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'intervention_followup_ignores'"
            ).fetchone()
        )
        # Une ancienne tentative peut avoir inscrit le numéro de migration sans
        # avoir terminé le schéma. La structure réelle reste la référence.
        needs_v6 = V6_SCHEMA_VERSION not in applied or not had_declared_column or not has_followup_table

    if not needs_v6:
        return
    if has_user_data:
        backup_profile_to(BACKUPS_DIR, "avant-migration-v6")

    with db_connection() as db:
        add_missing_columns(db, "interventions", {"declared": "INTEGER NOT NULL DEFAULT 0"})
        if not had_declared_column:
            # L'état des anciennes déclarations est inconnu : elles ne doivent pas
            # toutes apparaître comme des démarches en retard après la mise à jour.
            db.execute("UPDATE interventions SET declared = 1")
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS intervention_followup_ignores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                intervention_id INTEGER NOT NULL,
                reminder_type TEXT NOT NULL CHECK(reminder_type IN ('transmitted', 'declared', 'paid')),
                created_at TEXT NOT NULL,
                UNIQUE(intervention_id, reminder_type),
                FOREIGN KEY(intervention_id) REFERENCES interventions(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_intervention_followup_ignores_intervention
            ON intervention_followup_ignores(intervention_id)
            """
        )
        db.execute(
            "INSERT OR IGNORE INTO schema_migrations (version, applied_at, details) VALUES (?, ?, ?)",
            (V6_SCHEMA_VERSION, now_stamp(), "Suivi Transmis, Déclaré et Payé"),
        )


def init_db() -> None:
    with db_connection() as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS clients (
                name TEXT PRIMARY KEY,
                cesu TEXT NOT NULL DEFAULT '',
                email TEXT NOT NULL DEFAULT '',
                email_notes_enabled INTEGER NOT NULL DEFAULT 0,
                email_review_before_send INTEGER NOT NULL DEFAULT 0,
                hourly_rate REAL NOT NULL DEFAULT 0,
                hourly_rate_custom INTEGER NOT NULL DEFAULT 0,
                address TEXT NOT NULL DEFAULT '',
                phone TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS interventions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                client TEXT NOT NULL,
                duration_hours REAL NOT NULL,
                hourly_rate REAL NOT NULL,
                task TEXT NOT NULL DEFAULT '',
                location TEXT NOT NULL DEFAULT '',
                transmitted INTEGER NOT NULL DEFAULT 0,
                declared INTEGER NOT NULL DEFAULT 0,
                paid INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_interventions_date ON interventions(date);
            CREATE INDEX IF NOT EXISTS idx_interventions_client ON interventions(client);

            CREATE TABLE IF NOT EXISTS easy_cesu_metadata (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS reminders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_name TEXT NOT NULL,
                title TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                reference_date TEXT NOT NULL,
                due_time TEXT NOT NULL DEFAULT '',
                recurrence_type TEXT NOT NULL DEFAULT 'once',
                recurrence_interval INTEGER NOT NULL DEFAULT 1,
                anticipation_value INTEGER NOT NULL DEFAULT 0,
                anticipation_unit TEXT NOT NULL DEFAULT 'days',
                is_active INTEGER NOT NULL DEFAULT 1,
                next_occurrence_date TEXT NOT NULL DEFAULT '',
                last_processed_date TEXT NOT NULL DEFAULT '',
                is_completed INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(client_name) REFERENCES clients(name) ON UPDATE CASCADE ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS reminder_occurrences (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reminder_id INTEGER NOT NULL,
                due_date TEXT NOT NULL,
                processed_at TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'pending',
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                UNIQUE(reminder_id, due_date),
                FOREIGN KEY(reminder_id) REFERENCES reminders(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_reminders_client ON reminders(client_name);
            CREATE INDEX IF NOT EXISTS idx_reminders_next ON reminders(next_occurrence_date);
            CREATE INDEX IF NOT EXISTS idx_reminder_occurrences_due ON reminder_occurrences(due_date);
            """
        )
        columns = {row["name"] for row in db.execute("PRAGMA table_info(clients)").fetchall()}
        if "email" not in columns:
            db.execute("ALTER TABLE clients ADD COLUMN email TEXT NOT NULL DEFAULT ''")
        if "hourly_rate" not in columns:
            db.execute("ALTER TABLE clients ADD COLUMN hourly_rate REAL NOT NULL DEFAULT 0")
        if "hourly_rate_custom" not in columns:
            db.execute("ALTER TABLE clients ADD COLUMN hourly_rate_custom INTEGER NOT NULL DEFAULT 0")
            db.execute(
                """
                UPDATE clients
                SET hourly_rate_custom = CASE
                    WHEN hourly_rate > 0 AND ABS(hourly_rate - ?) > 0.0001 THEN 1
                    ELSE 0
                END
                """,
                (DEFAULT_RATE,),
            )
            db.execute("UPDATE clients SET hourly_rate = 0 WHERE hourly_rate_custom = 0")
    apply_v2_schema_migrations()
    apply_v3_schema_migrations()
    apply_v4_schema_migrations()
    apply_v6_schema_migrations()
    ensure_default_document_template()
    refresh_clients()
    seed_interventions_if_empty()
    refresh_reminder_occurrences()


def now_stamp() -> str:
    return datetime.now().isoformat(timespec="seconds")


def profile_attachments_dir(profile: dict | None = None) -> Path:
    selected = profile or active_profile()
    return ATTACHMENTS_DIR / str(selected.get("id") or "compte")


def validate_reminder(data: dict) -> tuple[dict, str | None]:
    client_name = clean_text(data.get("client_name"))
    title = clean_text(data.get("title"))
    if not client_name:
        return {}, "Client obligatoire pour le rappel."
    if not title:
        return {}, "Titre du rappel obligatoire."
    try:
        reference_date = parse_iso_date(data.get("reference_date"), "Date de référence")
    except ValueError as exc:
        return {}, str(exc)
    due_time = clean_text(data.get("due_time"))
    if due_time:
        try:
            datetime.strptime(due_time, "%H:%M")
        except ValueError:
            return {}, "Heure invalide."
    recurrence_type = clean_text(data.get("recurrence_type")) or "once"
    if recurrence_type not in RECURRENCES:
        return {}, "Type de récurrence invalide."
    try:
        recurrence_interval = int(data.get("recurrence_interval", 1) or 1)
    except (TypeError, ValueError):
        return {}, "Intervalle de récurrence invalide."
    if recurrence_interval < 1 or recurrence_interval > 120:
        return {}, "L'intervalle de récurrence doit être compris entre 1 et 120."
    try:
        anticipation_value = int(data.get("anticipation_value", 0) or 0)
    except (TypeError, ValueError):
        return {}, "Délai d'anticipation invalide."
    if anticipation_value < 0 or anticipation_value > 365:
        return {}, "Le délai d'anticipation doit être compris entre 0 et 365."
    anticipation_unit = clean_text(data.get("anticipation_unit")) or "days"
    if anticipation_unit not in ANTICIPATION_UNITS:
        return {}, "Unité d'anticipation invalide."
    return {
        "client_name": client_name,
        "title": title,
        "description": clean_text(data.get("description")),
        "reference_date": reference_date.isoformat(),
        "due_time": due_time,
        "recurrence_type": recurrence_type,
        "recurrence_interval": recurrence_interval,
        "anticipation_value": anticipation_value,
        "anticipation_unit": anticipation_unit,
        "is_active": 1 if data.get("is_active", True) else 0,
    }, None


def insert_occurrence(db: sqlite3.Connection, reminder_id: int, due_date: date) -> None:
    db.execute(
        """
        INSERT INTO reminder_occurrences (reminder_id, due_date, created_at)
        VALUES (?, ?, ?)
        ON CONFLICT(reminder_id, due_date) DO NOTHING
        """,
        (reminder_id, due_date.isoformat(), now_stamp()),
    )


def refresh_reminder_occurrences(today: date | None = None) -> None:
    """Crée les occurrences échues et la prochaine occurrence de chaque rappel actif."""

    current_day = today or date.today()
    with db_connection() as db:
        reminders = db.execute(
            "SELECT * FROM reminders WHERE is_active = 1 AND is_completed = 0"
        ).fetchall()
        for row in reminders:
            reference = parse_iso_date(row["reference_date"], "Date de référence")
            recurrence_type = str(row["recurrence_type"])
            interval = int(row["recurrence_interval"])
            for occurrence in occurrence_dates_until(reference, recurrence_type, interval, current_day):
                insert_occurrence(db, int(row["id"]), occurrence)
            if recurrence_type == "once":
                insert_occurrence(db, int(row["id"]), reference)
                next_date = reference if reference >= current_day else ""
            else:
                next_date_value = next_occurrence_after(reference, recurrence_type, interval, current_day)
                if next_date_value:
                    insert_occurrence(db, int(row["id"]), next_date_value)
                next_date = next_date_value.isoformat() if next_date_value else ""
            db.execute(
                "UPDATE reminders SET next_occurrence_date = ?, updated_at = ? WHERE id = ?",
                (next_date, now_stamp(), row["id"]),
            )


def reminder_notification_date(item: dict) -> date:
    due = parse_iso_date(item["due_date"], "Échéance")
    value = int(item.get("anticipation_value") or 0)
    unit = item.get("anticipation_unit") or "days"
    if unit == "weeks":
        return date.fromordinal(due.toordinal() - value * 7)
    if unit == "months":
        return add_months(due, -value)
    return date.fromordinal(due.toordinal() - value)


def list_reminder_occurrences(
    client_name: str = "", start: date | None = None, end: date | None = None, include_processed: bool = False
) -> list[dict]:
    refresh_reminder_occurrences()
    query = """
        SELECT occurrence.*, reminder.client_name, reminder.title, reminder.description,
               reminder.due_time, reminder.recurrence_type, reminder.recurrence_interval,
               reminder.anticipation_value, reminder.anticipation_unit, reminder.is_active
        FROM reminder_occurrences AS occurrence
        JOIN reminders AS reminder ON reminder.id = occurrence.reminder_id
        WHERE 1 = 1
    """
    params: list[object] = []
    if client_name:
        query += " AND reminder.client_name = ?"
        params.append(client_name)
    if start:
        query += " AND occurrence.due_date >= ?"
        params.append(start.isoformat())
    if end:
        query += " AND occurrence.due_date <= ?"
        params.append(end.isoformat())
    if not include_processed:
        query += " AND occurrence.status = 'pending' AND reminder.is_active = 1"
    query += " ORDER BY occurrence.due_date, reminder.due_time, reminder.title COLLATE NOCASE"
    with db_connection() as db:
        items = [dict(row) for row in db.execute(query, params).fetchall()]
    for item in items:
        item["is_active"] = bool(item["is_active"])
        item["notification_date"] = reminder_notification_date(item).isoformat()
    return items


def reminders_for_client(client_name: str) -> list[dict]:
    with db_connection() as db:
        reminders = [dict(row) for row in db.execute(
            "SELECT * FROM reminders WHERE client_name = ? ORDER BY is_active DESC, reference_date, title COLLATE NOCASE",
            (client_name,),
        ).fetchall()]
    for reminder in reminders:
        reminder["is_active"] = bool(reminder["is_active"])
        reminder["is_completed"] = bool(reminder["is_completed"])
        reminder["occurrences"] = list_reminder_occurrences(client_name, include_processed=True)
        reminder["occurrences"] = [item for item in reminder["occurrences"] if item["reminder_id"] == reminder["id"]]
    return reminders


def create_reminder(data: dict) -> dict:
    item, error = validate_reminder(data)
    if error:
        raise ValueError(error)
    with db_connection() as db:
        if not db.execute("SELECT 1 FROM clients WHERE name = ?", (item["client_name"],)).fetchone():
            raise ValueError("Le client associé au rappel est introuvable.")
        stamp = now_stamp()
        cursor = db.execute(
            """
            INSERT INTO reminders (
                client_name, title, description, reference_date, due_time, recurrence_type,
                recurrence_interval, anticipation_value, anticipation_unit, is_active,
                next_occurrence_date, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item["client_name"], item["title"], item["description"], item["reference_date"], item["due_time"],
                item["recurrence_type"], item["recurrence_interval"], item["anticipation_value"],
                item["anticipation_unit"], item["is_active"], item["reference_date"], stamp, stamp,
            ),
        )
        reminder_id = int(cursor.lastrowid)
    refresh_reminder_occurrences()
    return next(item for item in reminders_for_client(str(data["client_name"])) if item["id"] == reminder_id)


def update_reminder(reminder_id: int, data: dict) -> dict:
    item, error = validate_reminder(data)
    if error:
        raise ValueError(error)
    with db_connection() as db:
        existing = db.execute("SELECT * FROM reminders WHERE id = ?", (reminder_id,)).fetchone()
        if not existing:
            raise KeyError("Rappel introuvable.")
        if not db.execute("SELECT 1 FROM clients WHERE name = ?", (item["client_name"],)).fetchone():
            raise ValueError("Le client associé au rappel est introuvable.")
        db.execute(
            """
            UPDATE reminders SET client_name = ?, title = ?, description = ?, reference_date = ?, due_time = ?,
                recurrence_type = ?, recurrence_interval = ?, anticipation_value = ?, anticipation_unit = ?,
                is_active = ?, is_completed = CASE WHEN ? = 1 THEN is_completed ELSE 0 END,
                next_occurrence_date = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                item["client_name"], item["title"], item["description"], item["reference_date"], item["due_time"],
                item["recurrence_type"], item["recurrence_interval"], item["anticipation_value"],
                item["anticipation_unit"], item["is_active"], item["is_active"], item["reference_date"], now_stamp(), reminder_id,
            ),
        )
        # Les occurrences traitées restent l'historique ; les échéances à venir sont recalculées.
        db.execute("DELETE FROM reminder_occurrences WHERE reminder_id = ? AND status = 'pending'", (reminder_id,))
    refresh_reminder_occurrences()
    return next(item for item in reminders_for_client(item["client_name"]) if item["id"] == reminder_id)


def set_reminder_occurrence_status(reminder_id: int, occurrence_id: int, status: str) -> dict:
    if status not in {"completed", "skipped", "pending"}:
        raise ValueError("Statut de rappel invalide.")
    with db_connection() as db:
        occurrence = db.execute(
            "SELECT * FROM reminder_occurrences WHERE id = ? AND reminder_id = ?", (occurrence_id, reminder_id)
        ).fetchone()
        if not occurrence:
            raise KeyError("Occurrence de rappel introuvable.")
        reminder = db.execute("SELECT * FROM reminders WHERE id = ?", (reminder_id,)).fetchone()
        processed_at = now_stamp() if status in {"completed", "skipped"} else ""
        db.execute(
            "UPDATE reminder_occurrences SET status = ?, processed_at = ? WHERE id = ?",
            (status, processed_at, occurrence_id),
        )
        if reminder and reminder["recurrence_type"] == "once" and status == "completed":
            db.execute(
                "UPDATE reminders SET is_completed = 1, next_occurrence_date = '', last_processed_date = ?, updated_at = ? WHERE id = ?",
                (occurrence["due_date"], now_stamp(), reminder_id),
            )
        elif status in {"completed", "skipped"}:
            db.execute(
                "UPDATE reminders SET last_processed_date = ?, updated_at = ? WHERE id = ?",
                (occurrence["due_date"], now_stamp(), reminder_id),
            )
    refresh_reminder_occurrences()
    return {"reminder_id": reminder_id, "occurrence_id": occurrence_id, "status": status}


def delete_reminder(reminder_id: int) -> None:
    with db_connection() as db:
        cursor = db.execute("DELETE FROM reminders WHERE id = ?", (reminder_id,))
        if cursor.rowcount == 0:
            raise KeyError("Rappel introuvable.")


def reminders_overview(year: int | None = None, month: int | None = None) -> dict:
    current_day = date.today()
    items = list_reminder_occurrences()
    for item in items:
        due = parse_iso_date(item["due_date"])
        notification = parse_iso_date(item["notification_date"])
        item["state"] = "late" if due < current_day else "today" if due == current_day else "upcoming"
        item["is_notifiable"] = notification <= current_day
    if year and month:
        items = [item for item in items if item["due_date"].startswith(f"{year:04d}-{month:02d}-")]
    return {
        "today": current_day.isoformat(),
        "late": [item for item in items if item["state"] == "late"],
        "today_items": [item for item in items if item["state"] == "today"],
        "upcoming": [item for item in items if item["state"] == "upcoming"],
        "notifications": [item for item in items if item["is_notifiable"]],
        "items": items,
    }


def notes_output_dir() -> Path:
    profile = active_profile()
    value = profile.get("notes_intervention_dir") or CONFIG.get("notes_intervention_dir") or str(
        default_notes_dir()
    )
    return Path(str(value))


def ensure_folder(folder: Path) -> Path:
    if folder.exists() and not folder.is_dir():
        raise ValueError("Le chemin choisi n'est pas un dossier.")
    try:
        folder.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise ValueError(f"Impossible d'accéder au dossier : {folder}") from exc
    return folder


def ensure_notes_output_dir(folder: Path) -> Path:
    return ensure_folder(folder)


def export_output_dir(profile: dict | None = None) -> Path:
    selected_profile = profile or active_profile()
    value = selected_profile.get("export_dir") or CONFIG.get("export_dir") or str(default_export_dir())
    return Path(str(value)).expanduser()


def set_notes_output_dir(value: object, create: bool = False) -> Path:
    raw = str(value or "").strip().strip('"')
    if not raw:
        raise ValueError("Dossier de génération obligatoire.")
    folder = Path(raw).expanduser()
    if folder.exists() and not folder.is_dir():
        raise ValueError("Le chemin choisi n'est pas un dossier.")
    if create:
        ensure_notes_output_dir(folder)
    profile = active_profile()
    profile["notes_intervention_dir"] = str(folder)
    if profile.get("data_file") == LEGACY_DB_FILE:
        CONFIG["notes_intervention_dir"] = str(folder)
    save_config()
    return folder


def set_export_output_dir(value: object, create: bool = False) -> Path:
    raw = str(value or "").strip().strip('"')
    if not raw:
        raise ValueError("Dossier d'exports obligatoire.")
    folder = Path(raw).expanduser()
    if folder.exists() and not folder.is_dir():
        raise ValueError("Le chemin choisi pour les exports n'est pas un dossier.")
    if create:
        ensure_folder(folder)
    profile = active_profile()
    profile["export_dir"] = str(folder)
    if profile.get("data_file") == LEGACY_DB_FILE:
        CONFIG["export_dir"] = str(folder)
    save_config()
    return folder


def set_profile_data_dir(profile: dict, value: object, create: bool = True) -> Path:
    raw = str(value or "").strip().strip('"')
    if not raw:
        raise ValueError("Dossier de données obligatoire.")
    folder = Path(raw).expanduser()
    if folder.exists() and not folder.is_dir():
        raise ValueError("Le chemin choisi pour les données n'est pas un dossier.")

    old_db = profile_db_path(profile)
    if create:
        ensure_folder(folder)
    new_db = folder / profile_data_file_name(profile)
    if old_db.resolve() != new_db.resolve() and old_db.exists() and not new_db.exists():
        new_db.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(old_db, new_db)
    profile["data_dir"] = str(folder)
    if profile.get("data_file") == LEGACY_DB_FILE:
        CONFIG["data_dir"] = str(folder)
    save_config()
    return folder


def set_data_output_dir(value: object) -> Path:
    folder = set_profile_data_dir(active_profile(), value, create=True)
    init_db()
    return folder


def set_profile_database_file(profile: dict, value: object, create: bool = True) -> Path:
    raw = str(value or "").strip().strip('"')
    if not raw:
        raise ValueError("Fichier de base de données obligatoire.")
    database_path = Path(raw).expanduser()
    if database_path.exists() and not database_path.is_file():
        raise ValueError("Le chemin choisi pour la base de données n'est pas un fichier.")
    if create:
        database_path.parent.mkdir(parents=True, exist_ok=True)
    profile["data_file"] = str(database_path)
    profile.pop("data_dir", None)
    if database_path.exists():
        apply_profile_metadata(profile, read_profile_metadata_from_db(database_path))
    sync_legacy_source_config(profile)
    if profile.get("id") == CONFIG.get("active_profile_id"):
        sync_active_profile_runtime()
        init_db()
        write_profile_metadata_to_db()
    save_config()
    return database_path


def set_database_file(value: object) -> Path:
    return set_profile_database_file(active_profile(), value, create=True)


def validate_database_source(source_file: object) -> Path:
    raw = str(source_file or "").strip().strip('"')
    if not raw:
        raise ValueError("Fichier de base à importer obligatoire.")
    source_path = Path(raw).expanduser()
    if not source_path.exists() or not source_path.is_file():
        raise ValueError("Le fichier de base à importer est introuvable.")
    db = None
    try:
        db = sqlite3.connect(f"file:{source_path}?mode=ro", uri=True)
        names = {row[0] for row in db.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()}
    except sqlite3.Error as exc:
        raise ValueError("Le fichier choisi n'est pas une base SQLite valide.") from exc
    finally:
        if db is not None:
            db.close()
    if not ({"clients", "interventions"}.issubset(names) or "easy_cesu_metadata" in names):
        raise ValueError("Le fichier choisi ne ressemble pas à une base Easy CESU.")
    return source_path


def profile_import_destination(profile: dict) -> Path:
    raw = Path(str(profile.get("data_file") or LEGACY_DB_FILE).replace("\\", "/"))
    filename = raw.name or LEGACY_DB_FILE
    data_dir = clean_text(profile.get("data_dir"))
    if data_dir:
        return Path(data_dir).expanduser() / filename
    if raw.is_absolute() or raw.parent == Path("."):
        return DATA_DIR / filename
    return DATA_DIR / raw


def copy_sqlite_database(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temp_destination = destination.with_name(f"{destination.name}.import_tmp")
    if temp_destination.exists():
        temp_destination.unlink()
    source_db = sqlite3.connect(f"file:{source}?mode=ro", uri=True)
    target_db = sqlite3.connect(temp_destination)
    try:
        source_db.backup(target_db)
    finally:
        target_db.close()
        source_db.close()
    os.replace(temp_destination, destination)


def backup_existing_database_before_import(destination: Path, source: Path, profile: dict) -> Path | None:
    if not destination.exists():
        return None
    if destination.resolve(strict=False) == source.resolve(strict=False):
        return None
    backup_dir = destination.parent / "sauvegardes"
    ensure_folder(backup_dir)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    profile_slug = profile.get("id") or profile_id_from_label(profile.get("label", "compte"))
    backup_path = backup_dir / f"easy-cesu-{profile_slug}-avant-import-{stamp}.sqlite"
    shutil.copy2(destination, backup_path)
    return backup_path


def import_database_from(source_file: object) -> dict:
    source_path = validate_database_source(source_file)
    profile = active_profile()
    destination = profile_import_destination(profile)
    backup_path = backup_existing_database_before_import(destination, source_path, profile)
    if destination.resolve(strict=False) != source_path.resolve(strict=False):
        copy_sqlite_database(source_path, destination)

    profile["data_dir"] = str(destination.parent)
    profile["data_file"] = destination.name
    metadata_imported = apply_profile_metadata(profile, read_profile_metadata_from_db(destination))
    sync_active_profile_runtime()
    save_config()
    init_db()
    write_profile_metadata_to_db()
    return {
        "source": str(source_path),
        "database": str(destination),
        "backup": str(backup_path) if backup_path else "",
        "metadata_imported": metadata_imported,
        "settings": app_settings(),
        "clients": clients_list(),
    }


def sync_legacy_source_config(profile: dict) -> None:
    if profile.get("data_file") != LEGACY_DB_FILE:
        return
    CONFIG["suivi_paye_dir"] = profile.get("suivi_paye_dir", "")
    CONFIG["fichier_clients"] = profile.get("fichier_clients", "")
    CONFIG["suivi_paye_pattern"] = profile.get("suivi_paye_pattern", "Suivi de paye {year}.xlsx")


def set_source_data_dir(value: object) -> Path:
    raw = str(value or "").strip().strip('"')
    if not raw:
        raise ValueError("Dossier source obligatoire.")
    folder = Path(raw).expanduser()
    if folder.exists() and not folder.is_dir():
        raise ValueError("Le chemin choisi pour les sources n'est pas un dossier.")
    profile = active_profile()
    profile["suivi_paye_dir"] = str(folder)
    sync_legacy_source_config(profile)
    save_config()
    return folder


def set_clients_file(value: object) -> Path:
    raw = str(value or "").strip().strip('"')
    if not raw:
        raise ValueError("Fichier clients obligatoire.")
    file_path = Path(raw).expanduser()
    if file_path.exists() and not file_path.is_file():
        raise ValueError("Le chemin choisi pour les clients n'est pas un fichier.")
    profile = active_profile()
    profile["fichier_clients"] = str(file_path)
    sync_legacy_source_config(profile)
    save_config()
    return file_path


def backup_database_to(destination_dir: object | None = None) -> dict:
    profile = active_profile()
    db_path = active_db_path()
    if not db_path.exists():
        init_db()
    write_profile_metadata_to_db()
    raw = str(destination_dir or "").strip().strip('"')
    folder = Path(raw).expanduser() if raw else profile_data_dir(profile) / "sauvegardes"
    ensure_folder(folder)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    profile_slug = profile.get("id") or profile_id_from_label(profile.get("label", "compte"))
    backup_path = folder / f"easy-cesu-{profile_slug}-{stamp}.sqlite"
    source_db = sqlite3.connect(db_path)
    backup_db = sqlite3.connect(backup_path)
    try:
        source_db.backup(backup_db)
    finally:
        backup_db.close()
        source_db.close()
    return {
        "database": str(db_path),
        "backup": str(backup_path),
        "settings": app_settings(),
    }


def backup_profile_to(destination_dir: object | None = None, reason: str = "") -> dict:
    """Crée une sauvegarde ZIP transférable du compte actif.

    La base est prise par l'API SQLite ``backup`` : aucun simple copier-coller
    d'un fichier en cours d'écriture ne peut produire une archive incohérente.
    """

    profile = active_profile()
    write_profile_metadata_to_db()
    raw = clean_text(destination_dir)
    destination = Path(raw).expanduser() if raw else BACKUPS_DIR
    ensure_folder(destination)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"-{reason}" if reason else ""
    filename = f"EasyCESU-{profile['id']}-{stamp}{suffix}.zip"
    config = {
        "app_name": APP_NAME,
        "app_version": APP_VERSION,
        "profile": deepcopy(profile),
        "created_at": now_stamp(),
    }
    result = create_backup(
        destination / filename,
        active_db_path(),
        config,
        str(profile["id"]),
        profile_attachments_dir(profile),
    )
    return {
        "backup": result["path"],
        "manifest": result["manifest"],
        "settings": app_settings(),
    }


def validate_backup_source(source_file: object, bundled_backup_path: Path | None = None) -> tuple[Path, bool]:
    raw = clean_text(source_file)
    if not raw:
        raise ValueError("Archive de sauvegarde obligatoire.")
    source = Path(raw).expanduser()
    try:
        verify_backup(source)
        return source, False
    except ValueError as direct_error:
        if bundled_backup_path is None:
            raise direct_error
        try:
            extracted = extract_backup_from_transfer_kit(source, bundled_backup_path)
        except ValueError:
            raise direct_error
        return extracted, True


def restored_profile_from_backup(profile_data: dict) -> dict:
    raw_profile = profile_data.get("profile")
    if not isinstance(raw_profile, dict):
        raise ValueError("Le profil contenu dans la sauvegarde est invalide.")
    source_profile = deepcopy(raw_profile)
    requested_id = clean_text(source_profile.get("id")) or profile_id_from_label(
        clean_text(source_profile.get("label")) or "compte"
    )
    existing = profile_by_id(requested_id)
    generic_first_profile = is_unused_first_profile()
    if existing and not generic_first_profile:
        requested_id = profile_id_from_label(clean_text(source_profile.get("label")) or requested_id)
    source_profile["id"] = requested_id
    source_profile["label"] = clean_text(source_profile.get("label")) or requested_id
    source_profile["data_file"] = f"profiles/{requested_id}/interventions.sqlite"
    source_profile.pop("data_dir", None)
    source_profile.pop("workspace_root", None)
    # Les chemins de l'ancien ordinateur ne sont jamais réutilisés sans validation.
    source_profile["notes_intervention_dir"] = str(default_notes_dir() / requested_id)
    source_profile["export_dir"] = str(default_export_dir() / requested_id)
    source_profile["suivi_paye_dir"] = ""
    source_profile["fichier_clients"] = ""
    source_profile["suivi_paye_pattern"] = "Suivi de paye {year}.xlsx"
    return normalize_profile(source_profile)


def is_unused_first_profile() -> bool:
    if len(CONFIG.get("profiles", [])) != 1 or CONFIG["profiles"][0].get("id") != "mon-compte":
        return False
    with db_connection() as db:
        clients = db.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
        interventions = db.execute("SELECT COUNT(*) FROM interventions").fetchone()[0]
        reminders = db.execute("SELECT COUNT(*) FROM reminders").fetchone()[0]
    return not clients and not interventions and not reminders


def restore_profile_from_backup(source_file: object) -> dict:
    with tempfile.TemporaryDirectory(prefix="easycesu-restore-") as raw_temp:
        staging = Path(raw_temp)
        source, from_transfer_kit = validate_backup_source(source_file, staging / "sauvegarde-incluse.zip")
        details = verify_backup(source)
        replace_generic_profile = is_unused_first_profile()
        restored_profile = restored_profile_from_backup(details["profile"])
        previous_backup = ""
        current_db = active_db_path()
        if current_db.exists() and not replace_generic_profile:
            previous_backup = str(backup_profile_to(BACKUPS_DIR, "avant-restauration")["backup"])
        extract_backup(source, staging)
        staged_database = staging / "database.sqlite"
        target_database = DATA_DIR / "profiles" / restored_profile["id"] / "interventions.sqlite"
        copy_sqlite_database(staged_database, target_database)
        staged_attachments = staging / "attachments"
        target_attachments = profile_attachments_dir(restored_profile)
        if staged_attachments.exists():
            if target_attachments.exists():
                shutil.rmtree(target_attachments)
            shutil.copytree(staged_attachments, target_attachments)

    if replace_generic_profile:
        CONFIG["profiles"] = [restored_profile]
    else:
        CONFIG.setdefault("profiles", []).append(restored_profile)
    CONFIG["active_profile_id"] = restored_profile["id"]
    CONFIG["initial_setup_completed"] = False
    sync_active_profile_runtime()
    save_config()
    init_db()
    write_profile_metadata_to_db()
    return {
        "source": clean_text(source_file),
        "from_transfer_kit": from_transfer_kit,
        "backup_before_restore": previous_backup,
        "restored_profile": public_profile(restored_profile),
        "settings": app_settings(),
        "clients": clients_list(),
    }


def app_settings() -> dict:
    profile = active_profile()
    return {
        "notes_intervention_dir": str(notes_output_dir()),
        "export_dir": str(export_output_dir(profile)),
        "data_dir": str(profile_data_dir(profile)),
        "default_hourly_rate": DEFAULT_RATE,
        "active_profile_id": profile["id"],
        "initial_setup_required": not bool(CONFIG.get("initial_setup_completed")),
        "workspace_root": str(profile.get("workspace_root") or ""),
        "user_data_root": str(RUNTIME_DIR),
        "backups_dir": str(BACKUPS_DIR),
        "profile": public_profile(profile),
        "profiles": [public_profile(item) for item in CONFIG.get("profiles", [])],
    }


def optional_config_date(value: object) -> date | None:
    raw = clean_text(value)
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def support_reminder_status(current_date: date | None = None) -> dict:
    today = current_date or date.today()
    enabled = bool(CONFIG.get("support_reminder_enabled", True))
    started_on = optional_config_date(CONFIG.get("support_reminder_started_on")) or today
    last_seen_on = optional_config_date(CONFIG.get("support_reminder_last_seen_on"))
    first_due_on = started_on + timedelta(days=SUPPORT_REMINDER_INITIAL_DAYS)
    next_due_on = (
        last_seen_on + timedelta(days=SUPPORT_REMINDER_REPEAT_DAYS)
        if last_seen_on is not None
        else first_due_on
    )
    return {
        "enabled": enabled,
        "due": enabled and today >= next_due_on,
        "next_due_on": next_due_on.isoformat(),
    }


def community_info(current_date: date | None = None) -> dict:
    return {
        "developer_name": "Mamat Leroy",
        "repository_url": SUPPORT_LINKS["github_repository"],
        "issues_url": SUPPORT_LINKS["github_issues"],
        "support_url": SUPPORT_LINKS["paypal_me"],
        "support_reminder": support_reminder_status(current_date),
    }


def update_support_reminder(action: object, current_date: date | None = None) -> dict:
    today = current_date or date.today()
    normalized_action = clean_text(action).lower()
    if normalized_action == "dismiss":
        CONFIG["support_reminder_last_seen_on"] = today.isoformat()
    elif normalized_action == "disable":
        CONFIG["support_reminder_enabled"] = False
        CONFIG["support_reminder_last_seen_on"] = today.isoformat()
    elif normalized_action == "enable":
        if not bool(CONFIG.get("support_reminder_enabled", True)):
            CONFIG["support_reminder_started_on"] = today.isoformat()
            CONFIG["support_reminder_last_seen_on"] = ""
        CONFIG["support_reminder_enabled"] = True
    else:
        raise ValueError("Action de rappel de soutien invalide.")
    save_config()
    return support_reminder_status(today)


def open_external_link(link_id: object) -> dict:
    normalized_id = clean_text(link_id)
    url = SUPPORT_LINKS.get(normalized_id)
    if url is None:
        raise ValueError("Lien externe non autorisé.")
    return {"opened": bool(webbrowser.open(url, new=2)), "link_id": normalized_id}


def set_default_hourly_rate(value: object) -> float:
    global DEFAULT_RATE
    try:
        rate = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("Tarif horaire par défaut invalide.") from exc
    if rate <= 0:
        raise ValueError("Le tarif horaire par défaut doit être supérieur à zéro.")
    DEFAULT_RATE = round(rate, 4)
    profile = active_profile()
    profile["default_hourly_rate"] = DEFAULT_RATE
    if profile.get("data_file") == LEGACY_DB_FILE:
        CONFIG["salaire_net_horaire_defaut"] = DEFAULT_RATE
    save_config()
    write_profile_metadata_to_db()
    return DEFAULT_RATE


def update_settings(data: dict) -> dict:
    if "notes_intervention_dir" in data:
        set_notes_output_dir(data.get("notes_intervention_dir"))
    if "export_dir" in data:
        set_export_output_dir(data.get("export_dir"))
    if "data_dir" in data:
        set_data_output_dir(data.get("data_dir"))
    if "database_path" in data:
        set_database_file(data.get("database_path"))
    if "suivi_paye_dir" in data:
        set_source_data_dir(data.get("suivi_paye_dir"))
    if "fichier_clients" in data:
        set_clients_file(data.get("fichier_clients"))
    if "default_hourly_rate" in data:
        set_default_hourly_rate(data.get("default_hourly_rate"))
    elif "salaire_net_horaire_defaut" in data:
        set_default_hourly_rate(data.get("salaire_net_horaire_defaut"))
    return app_settings()


def clean_text(value: object) -> str:
    return str(value or "").strip()


def profile_metadata_payload(profile: dict) -> dict:
    payload = {field: profile.get(field, "") for field in PROFILE_METADATA_FIELDS}
    payload["default_hourly_rate"] = float(profile.get("default_hourly_rate") or DEFAULT_RATE)
    return {
        "version": 1,
        "saved_at": now_stamp(),
        "profile": payload,
    }


def write_profile_metadata_to_db(profile: dict | None = None) -> None:
    selected_profile = profile or active_profile()
    with db_connection() as db:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS easy_cesu_metadata (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            INSERT INTO easy_cesu_metadata (key, value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET
                value = excluded.value,
                updated_at = excluded.updated_at
            """,
            (
                PROFILE_METADATA_KEY,
                json.dumps(profile_metadata_payload(selected_profile), ensure_ascii=False),
                now_stamp(),
            ),
        )


def read_profile_metadata_from_db(database_path: Path) -> dict:
    if not database_path.exists():
        return {}
    db = None
    try:
        db = sqlite3.connect(database_path)
        row = db.execute(
            "SELECT value FROM easy_cesu_metadata WHERE key = ?",
            (PROFILE_METADATA_KEY,),
        ).fetchone()
    except sqlite3.Error:
        return {}
    finally:
        if db is not None:
            db.close()
    if not row:
        return {}
    try:
        payload = json.loads(row[0])
    except (TypeError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def apply_profile_metadata(profile: dict, metadata: dict) -> bool:
    imported_profile = metadata.get("profile") if isinstance(metadata, dict) else None
    if not isinstance(imported_profile, dict):
        return False
    imported = False
    for field in PROFILE_METADATA_FIELDS:
        if field == "default_hourly_rate":
            continue
        if field in imported_profile and clean_text(imported_profile.get(field)):
            profile[field] = clean_text(imported_profile.get(field))
            imported = True
    if "default_hourly_rate" in imported_profile:
        try:
            rate = float(imported_profile.get("default_hourly_rate"))
        except (TypeError, ValueError):
            rate = 0
        if rate > 0:
            profile["default_hourly_rate"] = round(rate, 4)
            imported = True
    if imported:
        sync_legacy_source_config(profile)
    return imported


def apply_profile_fields(profile: dict, data: dict, *, allow_label_empty: bool = False) -> None:
    label = clean_text(data.get("label", profile.get("label", "")))
    if not label and not allow_label_empty:
        raise ValueError("Nom du compte obligatoire.")
    if label:
        profile["label"] = label
    for field in ("name", "address", "phone", "email", "ss_number", "birth_info"):
        if field in data:
            profile[field] = clean_text(data.get(field))
    for field in (
        "first_name",
        "last_name",
        "commercial_name",
        "postal_code",
        "city",
        "siret",
        "secondary_activities",
        "theme",
    ):
        if field in data:
            profile[field] = clean_text(data.get(field))
    if "primary_activity" in data:
        activity = clean_text(data.get("primary_activity")) or "autre"
        if activity not in ACTIVITIES:
            raise ValueError("Activité principale inconnue.")
        profile["primary_activity"] = activity
    if "shortcut_icon" in data:
        shortcut_icon = clean_text(data.get("shortcut_icon")) or "generique"
        if shortcut_icon not in SHORTCUT_ICONS:
            raise ValueError("Icône de raccourci inconnue.")
        profile["shortcut_icon"] = shortcut_icon
    if "backup_dir" in data:
        raw_backup_dir = clean_text(data.get("backup_dir"))
        if raw_backup_dir:
            folder = Path(raw_backup_dir).expanduser()
            if folder.exists() and not folder.is_dir():
                raise ValueError("Le chemin choisi pour les sauvegardes n'est pas un dossier.")
            folder.mkdir(parents=True, exist_ok=True)
            probe = folder / ".easy-cesu-write-test"
            try:
                probe.write_text("ok", encoding="utf-8")
                probe.unlink()
            except OSError as exc:
                raise ValueError("Le dossier de sauvegarde n'est pas accessible en écriture.") from exc
            profile["backup_dir"] = str(folder)
    if "backup_retention_days" in data:
        try:
            retention = int(data.get("backup_retention_days"))
        except (TypeError, ValueError) as exc:
            raise ValueError("Nombre de sauvegardes invalide.") from exc
        if not 1 <= retention <= 365:
            raise ValueError("Conserver entre 1 et 365 sauvegardes.")
        profile["backup_retention_days"] = retention
    if "daily_backup_enabled" in data:
        profile["daily_backup_enabled"] = bool(data.get("daily_backup_enabled"))
    smtp_fields = set(smtp_defaults())
    if smtp_fields.intersection(data):
        combined = {field: profile.get(field) for field in smtp_fields}
        combined.update({field: data.get(field) for field in smtp_fields if field in data})
        profile.update(normalize_smtp_settings(combined))
    password = str(data.get("smtp_password") or "")
    if password:
        save_smtp_password(profile["id"], password)
    if data.get("smtp_clear_password"):
        delete_smtp_password(profile["id"])
    if "notes_intervention_dir" in data:
        raw_notes_dir = clean_text(data.get("notes_intervention_dir"))
        if raw_notes_dir:
            folder = Path(raw_notes_dir).expanduser()
            if folder.exists() and not folder.is_dir():
                raise ValueError("Le chemin choisi pour les notes n'est pas un dossier.")
            profile["notes_intervention_dir"] = str(folder)
    if "export_dir" in data:
        raw_export_dir = clean_text(data.get("export_dir"))
        if raw_export_dir:
            folder = Path(raw_export_dir).expanduser()
            if folder.exists() and not folder.is_dir():
                raise ValueError("Le chemin choisi pour les exports n'est pas un dossier.")
            profile["export_dir"] = str(folder)
    if "data_dir" in data:
        raw_data_dir = clean_text(data.get("data_dir"))
        if raw_data_dir:
            set_profile_data_dir(profile, raw_data_dir, create=True)
    if "database_path" in data:
        raw_database_path = clean_text(data.get("database_path"))
        if raw_database_path:
            set_profile_database_file(profile, raw_database_path, create=True)
    if "default_hourly_rate" in data or "salaire_net_horaire_defaut" in data:
        rate_value = data.get("default_hourly_rate", data.get("salaire_net_horaire_defaut"))
        try:
            rate = float(rate_value)
        except (TypeError, ValueError) as exc:
            raise ValueError("Tarif horaire par défaut invalide.") from exc
        if rate <= 0:
            raise ValueError("Le tarif horaire par défaut doit être supérieur à zéro.")
        profile["default_hourly_rate"] = round(rate, 4)
    if "suivi_paye_dir" in data:
        profile["suivi_paye_dir"] = clean_text(data.get("suivi_paye_dir"))
    if "fichier_clients" in data:
        profile["fichier_clients"] = clean_text(data.get("fichier_clients"))
    if "suivi_paye_pattern" in data:
        pattern = clean_text(data.get("suivi_paye_pattern")) or "Suivi de paye {year}.xlsx"
        profile["suivi_paye_pattern"] = pattern
    sync_legacy_source_config(profile)


def profile_by_id(profile_id: str) -> dict | None:
    for profile in CONFIG.get("profiles", []):
        if profile.get("id") == profile_id:
            return profile
    return None


def bootstrap_payload() -> dict:
    today = date.today()
    return {
        "today": today.isoformat(),
        "year": today.year,
        "month": today.month,
        "default_rate": DEFAULT_RATE,
        "settings": app_settings(),
        "clients": clients_list(),
        "reminders": reminders_overview(),
    }


def create_profile(data: dict) -> dict:
    label = clean_text(data.get("label"))
    if not label:
        raise ValueError("Nom du compte obligatoire.")
    profile_id = profile_id_from_label(label)
    profile = normalize_profile(
        {
            "id": profile_id,
            "label": label,
            "name": clean_text(data.get("name")),
            "address": clean_text(data.get("address")),
            "phone": clean_text(data.get("phone")),
            "email": clean_text(data.get("email")),
            "ss_number": clean_text(data.get("ss_number")),
            "birth_info": clean_text(data.get("birth_info")),
            "default_hourly_rate": DEFAULT_RATE,
            "notes_intervention_dir": clean_text(data.get("notes_intervention_dir"))
            or str(default_notes_dir() / label),
            "export_dir": clean_text(data.get("export_dir")) or str(default_export_dir() / label),
            "suivi_paye_dir": clean_text(data.get("suivi_paye_dir")),
            "fichier_clients": clean_text(data.get("fichier_clients")),
            "data_file": f"profiles/{profile_id}/interventions.sqlite",
            "seed_from_sorties": False,
        }
    )
    apply_profile_fields(profile, data)
    CONFIG["profiles"].append(profile)
    CONFIG["active_profile_id"] = profile["id"]
    sync_active_profile_runtime()
    save_config()
    init_db()
    write_profile_metadata_to_db()
    return bootstrap_payload()


def update_profile(profile_id: str, data: dict) -> dict:
    profile = profile_by_id(profile_id)
    if profile is None:
        raise KeyError("Compte introuvable.")
    apply_profile_fields(profile, data)
    if profile.get("data_file") == LEGACY_DB_FILE:
        CONFIG["notes_intervention_dir"] = profile.get("notes_intervention_dir", CONFIG["notes_intervention_dir"])
        CONFIG["export_dir"] = profile.get("export_dir", CONFIG.get("export_dir", str(default_export_dir())))
        if profile.get("data_dir"):
            CONFIG["data_dir"] = profile["data_dir"]
        CONFIG["salaire_net_horaire_defaut"] = float(profile.get("default_hourly_rate") or DEFAULT_RATE)
    if profile.get("id") == CONFIG.get("active_profile_id"):
        sync_active_profile_runtime()
        init_db()
        write_profile_metadata_to_db()
    save_config()
    return bootstrap_payload() if profile.get("id") == CONFIG.get("active_profile_id") else {"settings": app_settings()}


def switch_profile(profile_id: str) -> dict:
    if profile_by_id(profile_id) is None:
        raise KeyError("Compte introuvable.")
    CONFIG["active_profile_id"] = profile_id
    sync_active_profile_runtime()
    save_config()
    init_db()
    return bootstrap_payload()


def delete_profile(profile_id: str) -> dict:
    profiles = CONFIG.get("profiles", [])
    if len(profiles) <= 1:
        raise ValueError("Impossible de supprimer le dernier compte.")
    profile = profile_by_id(profile_id)
    if profile is None:
        raise KeyError("Compte introuvable.")
    delete_smtp_password(profile_id)
    CONFIG["profiles"] = [item for item in profiles if item.get("id") != profile_id]
    if CONFIG.get("active_profile_id") == profile_id:
        CONFIG["active_profile_id"] = CONFIG["profiles"][0]["id"]
    sync_active_profile_runtime()
    save_config()
    init_db()
    return bootstrap_payload()


def powershell_executable() -> str:
    windows_root = os.environ.get("SystemRoot") or os.environ.get("WINDIR") or r"C:\Windows"
    candidate = Path(windows_root) / "System32" / "WindowsPowerShell" / "v1.0" / "powershell.exe"
    return str(candidate) if candidate.exists() else "powershell.exe"


def run_macos_dialog(script: str, *arguments: str) -> tuple[Path | None, bool]:
    """Exécute un sélecteur Cocoa via AppleScript et distingue l'annulation."""

    try:
        completed = subprocess.run(
            ["/usr/bin/osascript", "-e", script, *arguments],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=300,
            check=False,
        )
    except FileNotFoundError as exc:
        raise ValueError("Le sélecteur de fichiers macOS est introuvable.") from exc
    except subprocess.TimeoutExpired as exc:
        raise ValueError("Le choix de fichier a pris trop de temps.") from exc
    selected = completed.stdout.strip().splitlines()[-1] if completed.stdout.strip() else ""
    if selected == "__EASY_CESU_CANCELLED__":
        return None, True
    if completed.returncode != 0:
        message = completed.stderr.strip() or "Impossible d'ouvrir le sélecteur macOS."
        raise ValueError(message)
    if not selected:
        return None, True
    return Path(selected), False


def choose_folder_macos(description: str, current_dir: Path) -> tuple[Path | None, bool]:
    script = """
on run argv
    set dialogPrompt to item 1 of argv
    set initialPath to item 2 of argv
    try
        if initialPath is not "" then
            set selectedFolder to choose folder with prompt dialogPrompt default location POSIX file initialPath
        else
            set selectedFolder to choose folder with prompt dialogPrompt
        end if
        return POSIX path of selectedFolder
    on error number -128
        return "__EASY_CESU_CANCELLED__"
    end try
end run
"""
    initial = current_dir if current_dir.exists() else Path.home()
    return run_macos_dialog(script, description, str(initial))


def choose_file_macos(description: str, current_file: Path) -> tuple[Path | None, bool]:
    script = """
on run argv
    set dialogPrompt to item 1 of argv
    set initialPath to item 2 of argv
    try
        if initialPath is not "" then
            set selectedFile to choose file with prompt dialogPrompt default location POSIX file initialPath
        else
            set selectedFile to choose file with prompt dialogPrompt
        end if
        return POSIX path of selectedFile
    on error number -128
        return "__EASY_CESU_CANCELLED__"
    end try
end run
"""
    if current_file.is_file():
        initial = current_file.parent
    elif current_file.is_dir():
        initial = current_file
    else:
        initial = Path.home()
    return run_macos_dialog(script, description, str(initial))


def choose_folder(description: str, current_dir: Path) -> tuple[Path | None, bool]:
    if sys.platform == "darwin":
        return choose_folder_macos(description, current_dir)
    if os.name != "nt":
        raise ValueError("Le choix de dossier natif n'est pas disponible sur ce système.")
    script = r"""
Add-Type -AssemblyName System.Windows.Forms
[Console]::OutputEncoding = [System.Text.Encoding]::UTF8
$dialog = New-Object System.Windows.Forms.FolderBrowserDialog
$dialog.Description = $env:FOLDER_DESCRIPTION
$dialog.ShowNewFolderButton = $true
if ($env:CURRENT_DIR -and (Test-Path -LiteralPath $env:CURRENT_DIR)) {
    $dialog.SelectedPath = $env:CURRENT_DIR
}
$result = $dialog.ShowDialog()
if ($result -eq [System.Windows.Forms.DialogResult]::OK) {
    Write-Output $dialog.SelectedPath
    exit 0
}
exit 2
"""
    env = os.environ.copy()
    env["FOLDER_DESCRIPTION"] = description
    env["CURRENT_DIR"] = str(current_dir)
    startupinfo = None
    creationflags = 0
    if os.name == "nt":
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = 1
        creationflags = getattr(subprocess, "CREATE_NEW_CONSOLE", 0)
    try:
        completed = subprocess.run(
            [powershell_executable(), "-NoProfile", "-STA", "-ExecutionPolicy", "Bypass", "-Command", script],
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=300,
            check=False,
            startupinfo=startupinfo,
            creationflags=creationflags,
        )
    except FileNotFoundError as exc:
        raise ValueError("PowerShell est introuvable pour ouvrir le choix de dossier.") from exc
    except subprocess.TimeoutExpired as exc:
        raise ValueError("Le choix de dossier a pris trop de temps.") from exc
    if completed.returncode == 2:
        return None, True
    if completed.returncode != 0:
        message = completed.stderr.strip() or "Impossible d'ouvrir le choix de dossier."
        raise ValueError(message)
    selected = completed.stdout.strip().splitlines()[-1] if completed.stdout.strip() else ""
    if not selected:
        return None, True
    return Path(selected), False


def choose_file(description: str, current_file: Path, file_filter: str) -> tuple[Path | None, bool]:
    if sys.platform == "darwin":
        return choose_file_macos(description, current_file)
    if os.name != "nt":
        raise ValueError("Le choix de fichier natif n'est pas disponible sur ce système.")
    script = r"""
Add-Type -AssemblyName System.Windows.Forms
[Console]::OutputEncoding = [System.Text.Encoding]::UTF8
$dialog = New-Object System.Windows.Forms.OpenFileDialog
$dialog.Title = $env:FILE_DESCRIPTION
$dialog.Filter = $env:FILE_FILTER
$dialog.CheckFileExists = $true
$dialog.CheckPathExists = $true
$dialog.ValidateNames = $true
$dialog.Multiselect = $false
if ($env:CURRENT_FILE -and (Test-Path -LiteralPath $env:CURRENT_FILE -PathType Leaf)) {
    $dialog.InitialDirectory = Split-Path -Parent $env:CURRENT_FILE
    $dialog.FileName = Split-Path -Leaf $env:CURRENT_FILE
} elseif ($env:CURRENT_FILE -and (Test-Path -LiteralPath $env:CURRENT_FILE -PathType Container)) {
    $dialog.InitialDirectory = $env:CURRENT_FILE
}
$result = $dialog.ShowDialog()
if ($result -eq [System.Windows.Forms.DialogResult]::OK) {
    Write-Output ([System.IO.Path]::GetFullPath($dialog.FileName))
    exit 0
}
exit 2
"""
    env = os.environ.copy()
    env["FILE_DESCRIPTION"] = description
    env["FILE_FILTER"] = file_filter
    env["CURRENT_FILE"] = str(current_file)
    startupinfo = None
    creationflags = 0
    if os.name == "nt":
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = 1
        creationflags = getattr(subprocess, "CREATE_NEW_CONSOLE", 0)
    try:
        completed = subprocess.run(
            [powershell_executable(), "-NoProfile", "-STA", "-ExecutionPolicy", "Bypass", "-Command", script],
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=300,
            check=False,
            startupinfo=startupinfo,
            creationflags=creationflags,
        )
    except FileNotFoundError as exc:
        raise ValueError("PowerShell est introuvable pour ouvrir le choix de fichier.") from exc
    except subprocess.TimeoutExpired as exc:
        raise ValueError("Le choix de fichier a pris trop de temps.") from exc
    if completed.returncode == 2:
        return None, True
    if completed.returncode != 0:
        message = completed.stderr.strip() or "Impossible d'ouvrir le choix de fichier."
        raise ValueError(message)
    selected = completed.stdout.strip().splitlines()[-1] if completed.stdout.strip() else ""
    if not selected:
        return None, True
    return Path(selected), False


def choose_notes_output_dir() -> tuple[Path | None, bool]:
    selected, cancelled = choose_folder("Choisir le dossier des notes d'intervention", notes_output_dir())
    if cancelled or selected is None:
        return selected, cancelled
    return set_notes_output_dir(selected, create=False), False


def choose_export_output_dir() -> tuple[Path | None, bool]:
    selected, cancelled = choose_folder("Choisir le dossier des exports Excel et JSON", export_output_dir())
    if cancelled or selected is None:
        return selected, cancelled
    return set_export_output_dir(selected, create=False), False


def choose_data_output_dir() -> tuple[Path | None, bool]:
    selected, cancelled = choose_folder("Choisir le dossier des donnees Easy CESU", profile_data_dir(active_profile()))
    if cancelled or selected is None:
        return selected, cancelled
    return set_data_output_dir(selected), False


def configure_workspace_root(root: Path) -> dict:
    """Range les fichiers d'un compte sous un dossier principal unique."""
    workspace_root = ensure_folder(root.expanduser())
    profile = active_profile()
    profile_folder = str(profile.get("id") or "mon-compte")
    data_dir = workspace_root / "Donnees" / profile_folder
    notes_dir = workspace_root / "Notes d'intervention" / profile_folder
    export_dir = workspace_root / "Exports" / profile_folder

    set_profile_data_dir(profile, data_dir, create=True)
    set_notes_output_dir(notes_dir, create=True)
    set_export_output_dir(export_dir, create=True)
    profile["workspace_root"] = str(workspace_root)
    CONFIG["initial_setup_completed"] = True
    save_config()
    init_db()
    write_profile_metadata_to_db()
    return {
        "workspace_root": str(workspace_root),
        "data_dir": str(data_dir),
        "notes_intervention_dir": str(notes_dir),
        "export_dir": str(export_dir),
        "settings": app_settings(),
    }


def choose_workspace_root() -> tuple[dict | None, bool]:
    profile = active_profile()
    configured_root = clean_text(profile.get("workspace_root"))
    documents = Path.home() / "Documents"
    current = Path(configured_root) if configured_root else (documents if documents.exists() else Path.home())
    selected, cancelled = choose_folder("Choisir ou creer le dossier principal Easy CESU", current)
    if cancelled or selected is None:
        return None, True
    workspace_root = selected if selected.name.casefold() == APP_NAME.casefold() else selected / APP_NAME
    return configure_workspace_root(workspace_root), False


def choose_database_file() -> tuple[Path | None, bool]:
    selected, cancelled = choose_file(
        "Choisir ou creer la base de donnees Easy CESU",
        active_db_path(),
        "Bases Easy CESU (*.sqlite;*.sqlite3;*.db)|*.sqlite;*.sqlite3;*.db|Tous les fichiers (*.*)|*.*",
    )
    if cancelled or selected is None:
        return selected, cancelled
    return set_database_file(selected), False


def choose_import_database_file() -> tuple[Path | None, bool]:
    selected, cancelled = choose_file(
        "Importer une base de donnees Easy CESU",
        active_db_path(),
        "Bases Easy CESU (*.sqlite;*.sqlite3;*.db)|*.sqlite;*.sqlite3;*.db|Tous les fichiers (*.*)|*.*",
    )
    if cancelled or selected is None:
        return selected, cancelled
    return selected, False


def choose_import_backup_file() -> tuple[Path | None, bool]:
    selected, cancelled = choose_file(
        "Restaurer une sauvegarde Easy CESU",
        BACKUPS_DIR,
        "Sauvegardes Easy CESU (*.zip)|*.zip|Tous les fichiers (*.*)|*.*",
    )
    if cancelled or selected is None:
        return selected, cancelled
    return selected, False


def choose_source_data_dir() -> tuple[Path | None, bool]:
    current = Path(active_profile().get("suivi_paye_dir") or str(Path.home()))
    selected, cancelled = choose_folder("Choisir le dossier du suivi de paye", current)
    if cancelled or selected is None:
        return selected, cancelled
    return set_source_data_dir(selected), False


def choose_clients_file() -> tuple[Path | None, bool]:
    profile = active_profile()
    current = Path(profile.get("fichier_clients") or profile.get("suivi_paye_dir") or str(Path.home()))
    selected, cancelled = choose_file("Choisir le fichier clients", current, "Fichiers Excel (*.xlsx)|*.xlsx|Tous les fichiers (*.*)|*.*")
    if cancelled or selected is None:
        return selected, cancelled
    return set_clients_file(selected), False


def choose_backup_database_dir() -> tuple[Path | None, bool]:
    selected, cancelled = choose_folder("Choisir le dossier de sauvegarde Easy CESU", BACKUPS_DIR)
    return selected, cancelled


def client_variants(qualite: object, nom: object, prenom: object) -> list[str]:
    nom_s = str(nom).strip() if nom else ""
    prenom_s = str(prenom).strip() if prenom else ""
    qualite_s = str(qualite).strip() if qualite else ""
    variants = [
        " ".join(x for x in [nom_s, prenom_s] if x),
        " ".join(x for x in [prenom_s, nom_s] if x),
        " ".join(x for x in [qualite_s, nom_s, prenom_s] if x),
    ]
    return [v for v in variants if v]


def looks_like_client_name(name: str) -> bool:
    normalized = normalize_name(name)
    if normalized in {"client", "clients", "nom", "prenom", "qualite", "numero cesu"}:
        return False
    if re.match(r"^\d{4}-\d{2}-\d{2}", name) or "00:00:00" in name:
        return False
    return len(normalized) >= 2


def open_clients_source(path: Path) -> object | None:
    """Ouvre une source Excel optionnelle sans empêcher l'application de démarrer."""

    try:
        if not path.exists():
            return None
        return load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - les fichiers externes ne doivent jamais bloquer Easy CESU
        print(f"Source clients ignorée : {path} ({exc})", flush=True)
        return None


def read_clients_from_sources() -> dict[str, dict[str, str]]:
    clients: dict[str, dict[str, str]] = {}
    by_norm: dict[str, str] = {}

    profile = active_profile()
    suivi_dir_raw = str(profile.get("suivi_paye_dir") or "").strip()
    suivi_path = None
    if suivi_dir_raw:
        try:
            suivi_dir = Path(suivi_dir_raw)
            pattern = profile.get("suivi_paye_pattern", "Suivi de paye {year}.xlsx")
            suivi_path = suivi_dir / str(pattern).format(year=date.today().year)
            if not suivi_path.exists() and suivi_dir.exists():
                fallback_files = sorted(suivi_dir.glob("Suivi de paye *.xlsx"), reverse=True)
                suivi_path = fallback_files[0] if fallback_files else suivi_path
        except OSError as exc:
            print(f"Dossier source clients ignoré : {suivi_dir_raw} ({exc})", flush=True)
            suivi_path = None

    wb = open_clients_source(suivi_path) if suivi_path else None
    if wb:
        if "Variables" in wb.sheetnames:
            ws = wb["Variables"]
            for row in ws.iter_rows(min_row=1, values_only=True):
                name = str(row[2]).strip() if len(row) >= 3 and row[2] else ""
                if not name or not looks_like_client_name(name):
                    continue
                cesu = str(row[3]).strip() if len(row) >= 4 and row[3] else ""
                clients[name] = {
                    "name": name,
                    "cesu": cesu,
                    "email": "",
                    "hourly_rate": 0,
                    "hourly_rate_custom": 0,
                    "address": "",
                    "phone": "",
                }
                by_norm[normalize_name(name)] = name
                by_norm[sorted_name_key(name)] = name
        wb.close()

    fichier_clients_raw = str(profile.get("fichier_clients") or "").strip()
    fichier_clients = Path(fichier_clients_raw) if fichier_clients_raw else None
    wb = open_clients_source(fichier_clients) if fichier_clients else None
    if wb:
        if "Liste clients" in wb.sheetnames:
            ws = wb["Liste clients"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                cesu, qualite, nom, prenom, address, phone = row[:6]
                variants = client_variants(qualite, nom, prenom)
                variants = [variant for variant in variants if looks_like_client_name(variant)]
                if not variants:
                    continue
                matched_name = None
                for variant in variants:
                    matched_name = by_norm.get(normalize_name(variant)) or by_norm.get(sorted_name_key(variant))
                    if matched_name:
                        break
                name = matched_name or variants[0]
                existing = clients.setdefault(
                    name,
                    {
                        "name": name,
                        "cesu": "",
                        "email": "",
                        "hourly_rate": 0,
                        "hourly_rate_custom": 0,
                        "address": "",
                        "phone": "",
                    },
                )
                if cesu and not existing["cesu"]:
                    existing["cesu"] = str(cesu).strip()
                if address and not existing["address"]:
                    existing["address"] = str(address).strip()
                if phone and not existing["phone"]:
                    existing["phone"] = str(phone).strip()
                by_norm[normalize_name(name)] = name
                by_norm[sorted_name_key(name)] = name
        wb.close()

    return clients


def refresh_clients() -> None:
    clients = read_clients_from_sources()
    stamp = now_stamp()
    with db_connection() as db:
        for client in clients.values():
            db.execute(
                """
                INSERT INTO clients (name, cesu, email, hourly_rate, hourly_rate_custom, address, phone, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(name) DO UPDATE SET
                    cesu = CASE WHEN clients.cesu = '' THEN excluded.cesu ELSE clients.cesu END,
                    email = CASE WHEN clients.email = '' THEN excluded.email ELSE clients.email END,
                    address = CASE WHEN clients.address = '' THEN excluded.address ELSE clients.address END,
                    phone = CASE WHEN clients.phone = '' THEN excluded.phone ELSE clients.phone END,
                    updated_at = excluded.updated_at
                """,
                (
                    client["name"],
                    client["cesu"],
                    client["email"],
                    client["hourly_rate"],
                    client["hourly_rate_custom"],
                    client["address"],
                    client["phone"],
                    stamp,
                ),
            )


def seed_interventions_if_empty() -> None:
    if not active_profile().get("seed_from_sorties"):
        return
    with db_connection() as db:
        count = db.execute("SELECT COUNT(*) FROM interventions").fetchone()[0]
        if count:
            return
    seed_files = sorted((ROOT_DIR / "sorties").glob("donnees_interventions_*.json"), reverse=True)
    if not seed_files:
        return
    payload = json.loads(seed_files[0].read_text(encoding="utf-8"))
    rows = payload.get("interventions", [])
    stamp = now_stamp()
    with db_connection() as db:
        for row in rows:
            db.execute(
                """
                INSERT INTO interventions
                    (date, client, duration_hours, hourly_rate, task, location, transmitted, paid, created_at, updated_at)
                VALUES (?, ?, ?, ?, '', '', 0, 0, ?, ?)
                """,
                (
                    row["date"],
                    row["client"],
                    float(row["duree_heures"]),
                    float(row.get("salaire_net_horaire", DEFAULT_RATE)),
                    stamp,
                    stamp,
                ),
            )


def row_to_dict(row: sqlite3.Row) -> dict:
    item = dict(row)
    item["duration_hours"] = round(float(item["duration_hours"]), 4)
    item["hourly_rate"] = round(float(item["hourly_rate"]), 4)
    item["amount_net"] = money(item["duration_hours"] * item["hourly_rate"])
    item["planned_amount"] = round(float(item.get("planned_amount") or item["duration_hours"] * item["hourly_rate"]), 2)
    item["received_amount"] = round(float(item.get("received_amount") or 0), 2)
    item["break_minutes"] = int(item.get("break_minutes") or 0)
    item["travel_minutes"] = int(item.get("travel_minutes") or 0)
    item["paid"] = bool(item["paid"])
    item["transmitted"] = bool(item["transmitted"])
    item["declared"] = bool(item["declared"])
    return item


def list_interventions(year: int | None = None, month: int | None = None, client: str = "") -> list[dict]:
    query = "SELECT * FROM interventions WHERE 1=1"
    params: list[object] = []
    if year:
        query += " AND substr(date, 1, 4) = ?"
        params.append(f"{year:04d}")
    if month:
        query += " AND substr(date, 6, 2) = ?"
        params.append(f"{month:02d}")
    if client:
        query += " AND client = ?"
        params.append(client)
    query += " ORDER BY date DESC, client COLLATE NOCASE"
    with db_connection() as db:
        return [row_to_dict(row) for row in db.execute(query, params).fetchall()]


def clients_list() -> list[dict]:
    with db_connection() as db:
        rows = db.execute("SELECT * FROM clients ORDER BY name COLLATE NOCASE").fetchall()
    clients = []
    for row in rows:
        item = dict(row)
        item["hourly_rate_custom"] = bool(item.get("hourly_rate_custom"))
        item["email_notes_enabled"] = bool(item.get("email_notes_enabled"))
        item["email_review_before_send"] = bool(item.get("email_review_before_send"))
        item["hourly_rate"] = round(float(item.get("hourly_rate") or 0), 4) if item["hourly_rate_custom"] else 0
        clients.append(item)
    return clients


def validate_client(data: dict) -> tuple[dict, str | None]:
    name = str(data.get("name", "") or "").strip()
    if not name:
        return {}, "Nom du client obligatoire."
    email = str(data.get("email", "") or "").strip()
    if email and "@" not in email:
        return {}, "Mail invalide."
    if "hourly_rate_custom" in data:
        hourly_rate_custom = bool(data.get("hourly_rate_custom"))
    else:
        hourly_rate_custom = str(data.get("hourly_rate", "") or "").strip() not in ("", "0", "0.0", "0.00")
    try:
        hourly_rate = float(data.get("hourly_rate", 0) or 0) if hourly_rate_custom else 0.0
    except (TypeError, ValueError):
        return {}, "Prix horaire invalide."
    if hourly_rate < 0:
        return {}, "Le prix horaire ne peut pas être négatif."
    if hourly_rate <= 0:
        hourly_rate = 0.0
        hourly_rate_custom = False
    return {
        "name": name,
        "cesu": str(data.get("cesu", "") or "").strip(),
        "email": email,
        "email_notes_enabled": 1 if data.get("email_notes_enabled") else 0,
        "email_review_before_send": 1 if data.get("email_review_before_send") else 0,
        "hourly_rate": round(hourly_rate, 4),
        "hourly_rate_custom": 1 if hourly_rate_custom else 0,
        "address": str(data.get("address", "") or "").strip(),
        "phone": str(data.get("phone", "") or "").strip(),
        "activity": clean_text(data.get("activity")),
        "instructions": clean_text(data.get("instructions")),
        "access_info": clean_text(data.get("access_info")),
        "payment_preferences": clean_text(data.get("payment_preferences")),
        "usual_frequency": clean_text(data.get("usual_frequency")),
        "preferred_days": clean_text(data.get("preferred_days")),
        "is_archived": 1 if data.get("is_archived") else 0,
    }, None


def client_usual_duration(data: dict) -> tuple[float, str | None]:
    try:
        value = float(data.get("usual_duration_hours", 0) or 0)
    except (TypeError, ValueError):
        return 0.0, "Durée habituelle invalide."
    if value < 0:
        return 0.0, "La durée habituelle ne peut pas être négative."
    return round(value, 4), None


def create_client(data: dict) -> dict:
    item, error = validate_client(data)
    if error:
        raise ValueError(error)
    item["usual_duration_hours"], error = client_usual_duration(data)
    if error:
        raise ValueError(error)
    stamp = now_stamp()
    with db_connection() as db:
        try:
            db.execute(
                """
                INSERT INTO clients
                    (name, cesu, email, email_notes_enabled, email_review_before_send, hourly_rate, hourly_rate_custom, address, phone, activity, instructions,
                    access_info, payment_preferences, usual_duration_hours, usual_frequency, preferred_days,
                    is_archived, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    item["name"],
                    item["cesu"],
                    item["email"],
                    item["email_notes_enabled"],
                    item["email_review_before_send"],
                    item["hourly_rate"],
                    item["hourly_rate_custom"],
                    item["address"],
                    item["phone"],
                    item["activity"],
                    item["instructions"],
                    item["access_info"],
                    item["payment_preferences"],
                    item["usual_duration_hours"],
                    item["usual_frequency"],
                    item["preferred_days"],
                    item["is_archived"],
                    stamp,
                    stamp,
                ),
            )
        except sqlite3.IntegrityError as exc:
            raise ValueError("Ce client existe déjà.") from exc
        row = db.execute("SELECT * FROM clients WHERE name = ?", (item["name"],)).fetchone()
    return dict(row)


def update_client(original_name: str, data: dict) -> dict:
    item, error = validate_client(data)
    if error:
        raise ValueError(error)
    item["usual_duration_hours"], error = client_usual_duration(data)
    if error:
        raise ValueError(error)
    stamp = now_stamp()
    with db_connection() as db:
        existing = db.execute("SELECT * FROM clients WHERE name = ?", (original_name,)).fetchone()
        if not existing:
            raise KeyError("Client introuvable.")
        if item["name"] != original_name:
            duplicate = db.execute("SELECT 1 FROM clients WHERE name = ?", (item["name"],)).fetchone()
            if duplicate:
                raise ValueError("Un autre client porte déjà ce nom.")
        db.execute(
            """
            UPDATE clients
            SET name = ?, cesu = ?, email = ?, email_notes_enabled = ?, email_review_before_send = ?, hourly_rate = ?, hourly_rate_custom = ?, address = ?, phone = ?,
                activity = ?, instructions = ?, access_info = ?, payment_preferences = ?, usual_duration_hours = ?,
                usual_frequency = ?, preferred_days = ?, is_archived = ?, updated_at = ?
            WHERE name = ?
            """,
            (
                item["name"],
                item["cesu"],
                item["email"],
                item["email_notes_enabled"],
                item["email_review_before_send"],
                item["hourly_rate"],
                item["hourly_rate_custom"],
                item["address"],
                item["phone"],
                item["activity"],
                item["instructions"],
                item["access_info"],
                item["payment_preferences"],
                item["usual_duration_hours"],
                item["usual_frequency"],
                item["preferred_days"],
                item["is_archived"],
                stamp,
                original_name,
            ),
        )
        if item["name"] != original_name:
            db.execute("UPDATE interventions SET client = ?, updated_at = ? WHERE client = ?", (item["name"], stamp, original_name))
        row = db.execute("SELECT * FROM clients WHERE name = ?", (item["name"],)).fetchone()
    return dict(row)


def delete_client(name: str) -> None:
    with db_connection() as db:
        cursor = db.execute("DELETE FROM clients WHERE name = ?", (name,))
        if cursor.rowcount == 0:
            raise KeyError("Client introuvable.")


def client_custom_hourly_rate(client: str) -> float:
    with db_connection() as db:
        row = db.execute(
            "SELECT hourly_rate, hourly_rate_custom FROM clients WHERE name = ?",
            (client,),
        ).fetchone()
    if not row or not row["hourly_rate_custom"]:
        return 0.0
    rate = float(row["hourly_rate"] or 0)
    return rate if rate > 0 else 0.0


def validate_intervention(data: dict) -> tuple[dict, str | None]:
    try:
        when = datetime.fromisoformat(str(data.get("date", ""))).date()
    except ValueError:
        return {}, "Date invalide."
    client = str(data.get("client", "")).strip()
    if not client:
        return {}, "Client obligatoire."
    try:
        duration = float(data.get("duration_hours", 0))
    except (TypeError, ValueError):
        return {}, "Durée invalide."
    if duration <= 0:
        return {}, "La durée doit être supérieure à zéro."
    try:
        provided_rate = float(data.get("hourly_rate", DEFAULT_RATE) or DEFAULT_RATE)
    except (TypeError, ValueError):
        return {}, "Salaire horaire invalide."
    if provided_rate <= 0:
        return {}, "Salaire horaire invalide."
    rate = client_custom_hourly_rate(client) or provided_rate
    task = str(data.get("task", "") or "").strip()
    location = str(data.get("location", "") or "").strip()
    status = clean_text(data.get("status")) or "realisee"
    status = {"realized": "realisee", "planned": "prevue", "confirmed": "confirmee", "cancelled": "annulee", "rescheduled": "reportee"}.get(status, status)
    if status not in INTERVENTION_STATUSES:
        return {}, "Statut d'intervention invalide."
    payment_status = clean_text(data.get("payment_status"))
    payment_status = {"received": "recu", "to_receive": "a_recevoir", "partially_received": "partiellement_recu", "cancelled": "annule"}.get(payment_status, payment_status)
    if payment_status and payment_status not in PAYMENT_STATUSES:
        return {}, "Statut de paiement invalide."
    try:
        break_minutes = int(data.get("break_minutes", 0) or 0)
        travel_minutes = int(data.get("travel_minutes", 0) or 0)
        planned_amount = float(data.get("planned_amount", duration * rate) or 0)
        received_amount = float(data.get("received_amount", 0) or 0)
    except (TypeError, ValueError):
        return {}, "Les informations de temps ou de paiement sont invalides."
    if break_minutes < 0 or travel_minutes < 0 or planned_amount < 0 or received_amount < 0:
        return {}, "Les informations de temps ou de paiement ne peuvent pas être négatives."
    raw_category_id = data.get("category_id")
    try:
        category_id = int(raw_category_id) if raw_category_id not in (None, "") else None
    except (TypeError, ValueError):
        return {}, "Catégorie de prestation invalide."
    if category_id is not None:
        with db_connection() as db:
            if not db.execute("SELECT 1 FROM service_categories WHERE id = ?", (category_id,)).fetchone():
                return {}, "Catégorie de prestation introuvable."
    return {
        "date": when.isoformat(),
        "client": client,
        "duration_hours": round(duration, 4),
        "hourly_rate": round(rate, 4),
        "task": task,
        "location": location,
        "transmitted": 1 if data.get("transmitted") else 0,
        "declared": 1 if data.get("declared") else 0,
        "paid": 1 if data.get("paid") else 0,
        "planned_start": clean_text(data.get("planned_start")),
        "planned_end": clean_text(data.get("planned_end")),
        "actual_start": clean_text(data.get("actual_start")),
        "actual_end": clean_text(data.get("actual_end")),
        "break_minutes": break_minutes,
        "travel_minutes": travel_minutes,
        "status": status,
        "category_id": category_id,
        "planned_amount": round(planned_amount, 2),
        "received_amount": round(received_amount, 2),
        "payment_status": payment_status or ("recu" if data.get("paid") else "a_recevoir"),
    }, None


def add_or_update_client_from_intervention(item: dict) -> None:
    stamp = now_stamp()
    with db_connection() as db:
        db.execute(
            """
            INSERT INTO clients (name, cesu, email, hourly_rate, hourly_rate_custom, address, phone, updated_at)
            VALUES (?, '', '', 0, 0, ?, '', ?)
            ON CONFLICT(name) DO UPDATE SET
                address = CASE WHEN clients.address = '' THEN excluded.address ELSE clients.address END,
                updated_at = excluded.updated_at
            """,
            (item["client"], item.get("location", ""), stamp),
        )


def create_intervention(data: dict) -> dict:
    item, error = validate_intervention(data)
    if error:
        raise ValueError(error)
    add_or_update_client_from_intervention(item)
    stamp = now_stamp()
    with db_connection() as db:
        cursor = db.execute(
            """
            INSERT INTO interventions
                (date, client, duration_hours, hourly_rate, task, location, transmitted, declared, paid,
                planned_start, planned_end, actual_start, actual_end, break_minutes, travel_minutes, status,
                category_id, planned_amount, received_amount, payment_status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item["date"],
                item["client"],
                item["duration_hours"],
                item["hourly_rate"],
                item["task"],
                item["location"],
                item["transmitted"],
                item["declared"],
                item["paid"],
                item["planned_start"],
                item["planned_end"],
                item["actual_start"],
                item["actual_end"],
                item["break_minutes"],
                item["travel_minutes"],
                item["status"],
                item["category_id"],
                item["planned_amount"],
                item["received_amount"],
                item["payment_status"],
                stamp,
                stamp,
            ),
        )
        row = db.execute("SELECT * FROM interventions WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return row_to_dict(row)


def update_intervention(intervention_id: int, data: dict) -> dict:
    # Les anciennes versions de l'interface n'envoyaient pas les champs V2 :
    # on conserve donc leur valeur enregistrée lors d'une simple modification historique.
    with db_connection() as db:
        existing = db.execute("SELECT * FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
    if not existing:
        raise KeyError("Intervention introuvable.")
    merged_data = dict(existing)
    merged_data.update(data)
    item, error = validate_intervention(merged_data)
    if error:
        raise ValueError(error)
    add_or_update_client_from_intervention(item)
    stamp = now_stamp()
    with db_connection() as db:
        db.execute(
            """
            UPDATE interventions
            SET date = ?, client = ?, duration_hours = ?, hourly_rate = ?, task = ?, location = ?,
                transmitted = ?, declared = ?, paid = ?, planned_start = ?, planned_end = ?, actual_start = ?, actual_end = ?,
                break_minutes = ?, travel_minutes = ?, status = ?, category_id = ?, planned_amount = ?,
                received_amount = ?, payment_status = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                item["date"],
                item["client"],
                item["duration_hours"],
                item["hourly_rate"],
                item["task"],
                item["location"],
                item["transmitted"],
                item["declared"],
                item["paid"],
                item["planned_start"],
                item["planned_end"],
                item["actual_start"],
                item["actual_end"],
                item["break_minutes"],
                item["travel_minutes"],
                item["status"],
                item["category_id"],
                item["planned_amount"],
                item["received_amount"],
                item["payment_status"],
                stamp,
                intervention_id,
            ),
        )
        for reminder_type in ADMINISTRATIVE_STATUS_FIELDS:
            if item[reminder_type]:
                db.execute(
                    "DELETE FROM intervention_followup_ignores WHERE intervention_id = ? AND reminder_type = ?",
                    (intervention_id, reminder_type),
                )
        row = db.execute("SELECT * FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
    return row_to_dict(row)


def delete_intervention(intervention_id: int) -> None:
    with db_connection() as db:
        cursor = db.execute("DELETE FROM interventions WHERE id = ?", (intervention_id,))
        if cursor.rowcount == 0:
            raise KeyError("Intervention introuvable.")


def administrative_status_label(reminder_type: str) -> str:
    return {
        "transmitted": "À transmettre",
        "declared": "À déclarer",
        "paid": "À payer",
    }.get(reminder_type, reminder_type)


def update_intervention_administrative_status(intervention_id: int, reminder_type: str, checked: bool) -> dict:
    """Met à jour un seul état sans réécrire le reste de l'intervention."""

    reminder_type = clean_text(reminder_type)
    if reminder_type not in ADMINISTRATIVE_STATUS_FIELDS:
        raise ValueError("État administratif invalide.")
    checked_value = 1 if checked else 0
    with db_connection() as db:
        existing = db.execute("SELECT * FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
        if not existing:
            raise KeyError("Intervention introuvable.")
        if reminder_type == "transmitted":
            db.execute(
                "UPDATE interventions SET transmitted = ?, updated_at = ? WHERE id = ?",
                (checked_value, now_stamp(), intervention_id),
            )
        elif reminder_type == "declared":
            db.execute(
                "UPDATE interventions SET declared = ?, updated_at = ? WHERE id = ?",
                (checked_value, now_stamp(), intervention_id),
            )
        else:
            received_amount = max(0.0, float(existing["received_amount"] or 0))
            expected_amount = float(existing["planned_amount"] or 0)
            if expected_amount <= 0:
                expected_amount = float(existing["duration_hours"]) * float(existing["hourly_rate"])
            if checked and received_amount <= 0:
                received_amount = expected_amount
            payment_status = "recu" if checked else ("partiellement_recu" if received_amount > 0 else "a_recevoir")
            db.execute(
                """
                UPDATE interventions
                SET paid = ?, received_amount = ?, payment_status = ?, updated_at = ?
                WHERE id = ?
                """,
                (checked_value, round(received_amount, 2), payment_status, now_stamp(), intervention_id),
            )
            # Un paiement explicitement lié suit le même état, sans écraser un
            # montant reçu qui aurait été saisi manuellement.
            if checked:
                db.execute(
                    """
                    UPDATE pending_payments
                    SET status = 'recu',
                        received_amount = CASE WHEN received_amount > 0 THEN received_amount ELSE expected_amount END,
                        updated_at = ?
                    WHERE intervention_id = ?
                    """,
                    (now_stamp(), intervention_id),
                )
            else:
                db.execute(
                    """
                    UPDATE pending_payments
                    SET status = CASE WHEN received_amount > 0 THEN 'partiellement_recu' ELSE 'a_recevoir' END,
                        updated_at = ?
                    WHERE intervention_id = ?
                    """,
                    (now_stamp(), intervention_id),
                )
        if checked:
            db.execute(
                "DELETE FROM intervention_followup_ignores WHERE intervention_id = ? AND reminder_type = ?",
                (intervention_id, reminder_type),
            )
        row = db.execute("SELECT * FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
    return row_to_dict(row)


def list_intervention_followups(
    reminder_type: str = "",
    search: str = "",
    include_ignored: bool = False,
) -> list[dict]:
    reminder_type = clean_text(reminder_type)
    if reminder_type and reminder_type not in ADMINISTRATIVE_STATUS_FIELDS:
        raise ValueError("Filtre de suivi invalide.")
    search = clean_text(search)
    query = "SELECT * FROM interventions"
    params: list[object] = []
    if search:
        query += " WHERE client LIKE ? COLLATE NOCASE"
        params.append(f"%{search}%")
    query += " ORDER BY date ASC, client COLLATE NOCASE, id ASC"
    with db_connection() as db:
        rows = db.execute(query, params).fetchall()
        ignored_rows = db.execute(
            "SELECT intervention_id, reminder_type, created_at FROM intervention_followup_ignores"
        ).fetchall()
    ignored = {
        (int(row["intervention_id"]), str(row["reminder_type"])): str(row["created_at"])
        for row in ignored_rows
    }
    followups: list[dict] = []
    for row in rows:
        item = row_to_dict(row)
        all_missing = [status for status in ("transmitted", "declared", "paid") if not item[status]]
        active_missing = [status for status in all_missing if (item["id"], status) not in ignored]
        ignored_missing = [status for status in all_missing if (item["id"], status) in ignored]
        visible_statuses = active_missing + (ignored_missing if include_ignored else [])
        if reminder_type and reminder_type not in visible_statuses:
            continue
        if not visible_statuses:
            continue
        item["missing_reminders"] = active_missing
        item["missing_reminder_labels"] = [administrative_status_label(status) for status in active_missing]
        item["ignored_reminders"] = ignored_missing
        item["ignored_reminder_labels"] = [administrative_status_label(status) for status in ignored_missing]
        item["ignored_at"] = {status: ignored[(item["id"], status)] for status in ignored_missing}
        followups.append(item)
    return followups


def ignore_intervention_followup(intervention_id: int, reminder_type: str) -> dict:
    reminder_type = clean_text(reminder_type)
    if reminder_type not in ADMINISTRATIVE_STATUS_FIELDS:
        raise ValueError("Type de rappel invalide.")
    with db_connection() as db:
        intervention = db.execute("SELECT * FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
        if not intervention:
            raise KeyError("Intervention introuvable.")
        if bool(intervention[reminder_type]):
            raise ValueError("Cette action est déjà terminée.")
        stamp = now_stamp()
        db.execute(
            """
            INSERT INTO intervention_followup_ignores (intervention_id, reminder_type, created_at)
            VALUES (?, ?, ?)
            ON CONFLICT(intervention_id, reminder_type) DO UPDATE SET created_at = excluded.created_at
            """,
            (intervention_id, reminder_type, stamp),
        )
    return {"intervention_id": intervention_id, "reminder_type": reminder_type, "created_at": stamp}


def reactivate_intervention_followup(intervention_id: int, reminder_type: str) -> dict:
    reminder_type = clean_text(reminder_type)
    if reminder_type not in ADMINISTRATIVE_STATUS_FIELDS:
        raise ValueError("Type de rappel invalide.")
    with db_connection() as db:
        if not db.execute("SELECT 1 FROM interventions WHERE id = ?", (intervention_id,)).fetchone():
            raise KeyError("Intervention introuvable.")
        cursor = db.execute(
            "DELETE FROM intervention_followup_ignores WHERE intervention_id = ? AND reminder_type = ?",
            (intervention_id, reminder_type),
        )
    return {"intervention_id": intervention_id, "reminder_type": reminder_type, "reactivated": cursor.rowcount > 0}


def category_to_dict(row: sqlite3.Row) -> dict:
    item = dict(row)
    item["default_hourly_rate"] = round(float(item.get("default_hourly_rate") or 0), 2)
    item["default_duration_hours"] = round(float(item.get("default_duration_hours") or 0), 2)
    item["is_archived"] = bool(item.get("is_archived"))
    return item


def list_service_categories(include_archived: bool = False) -> list[dict]:
    query = "SELECT * FROM service_categories"
    if not include_archived:
        query += " WHERE is_archived = 0"
    query += " ORDER BY activity COLLATE NOCASE, name COLLATE NOCASE"
    with db_connection() as db:
        return [category_to_dict(row) for row in db.execute(query).fetchall()]


def validate_service_category(data: dict) -> tuple[dict, str | None]:
    name = clean_text(data.get("name"))
    if not name:
        return {}, "Nom de prestation obligatoire."
    activity = clean_text(data.get("activity"))
    if activity and activity not in ACTIVITIES:
        return {}, "Activité de prestation inconnue."
    try:
        rate = float(data.get("default_hourly_rate", 0) or 0)
        duration = float(data.get("default_duration_hours", 0) or 0)
    except (TypeError, ValueError):
        return {}, "Tarif ou durée de prestation invalide."
    if rate < 0 or duration < 0:
        return {}, "Tarif et durée ne peuvent pas être négatifs."
    color = clean_text(data.get("color"))
    if color and not re.fullmatch(r"#[0-9a-fA-F]{6}", color):
        return {}, "Couleur de prestation invalide."
    return {
        "name": name,
        "activity": activity,
        "icon_key": clean_text(data.get("icon_key")),
        "color": color,
        "default_hourly_rate": round(rate, 2),
        "default_duration_hours": round(duration, 2),
        "is_archived": 1 if data.get("is_archived") else 0,
    }, None


def create_service_category(data: dict) -> dict:
    item, error = validate_service_category(data)
    if error:
        raise ValueError(error)
    stamp = now_stamp()
    with db_connection() as db:
        try:
            cursor = db.execute(
                """
                INSERT INTO service_categories
                    (name, activity, icon_key, color, default_hourly_rate, default_duration_hours, is_archived, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (*item.values(), stamp, stamp),
            )
        except sqlite3.IntegrityError as exc:
            raise ValueError("Cette prestation existe déjà.") from exc
        row = db.execute("SELECT * FROM service_categories WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return category_to_dict(row)


def update_service_category(category_id: int, data: dict) -> dict:
    item, error = validate_service_category(data)
    if error:
        raise ValueError(error)
    with db_connection() as db:
        if not db.execute("SELECT 1 FROM service_categories WHERE id = ?", (category_id,)).fetchone():
            raise KeyError("Prestation introuvable.")
        db.execute(
            """
            UPDATE service_categories
            SET name = ?, activity = ?, icon_key = ?, color = ?, default_hourly_rate = ?,
                default_duration_hours = ?, is_archived = ?, updated_at = ?
            WHERE id = ?
            """,
            (*item.values(), now_stamp(), category_id),
        )
        row = db.execute("SELECT * FROM service_categories WHERE id = ?", (category_id,)).fetchone()
    return category_to_dict(row)


DOCUMENT_TEMPLATE_TYPE = "intervention_note"


def document_template_to_dict(row: sqlite3.Row) -> dict:
    item = dict(row)
    try:
        configuration = json.loads(str(item.pop("configuration_json")))
    except (TypeError, ValueError, json.JSONDecodeError):
        configuration = {}
    item["configuration"] = normalize_note_template_configuration(configuration)
    item["is_default"] = bool(item.get("is_default"))
    return item


def ensure_default_document_template() -> None:
    with db_connection() as db:
        existing = db.execute(
            "SELECT id FROM document_templates WHERE document_type = ? ORDER BY id LIMIT 1",
            (DOCUMENT_TEMPLATE_TYPE,),
        ).fetchone()
        if existing:
            has_default = db.execute(
                "SELECT 1 FROM document_templates WHERE document_type = ? AND is_default = 1 LIMIT 1",
                (DOCUMENT_TEMPLATE_TYPE,),
            ).fetchone()
            if not has_default:
                db.execute("UPDATE document_templates SET is_default = 1 WHERE id = ?", (existing["id"],))
            return
        stamp = now_stamp()
        db.execute(
            """
            INSERT INTO document_templates
                (name, document_type, is_default, configuration_json, created_at, updated_at)
            VALUES (?, ?, 1, ?, ?, ?)
            """,
            (
                "Note d’intervention classique",
                DOCUMENT_TEMPLATE_TYPE,
                json.dumps(default_note_template_configuration(), ensure_ascii=False),
                stamp,
                stamp,
            ),
        )


def list_document_templates() -> list[dict]:
    ensure_default_document_template()
    with db_connection() as db:
        rows = db.execute(
            """
            SELECT * FROM document_templates
            WHERE document_type = ?
            ORDER BY is_default DESC, name COLLATE NOCASE
            """,
            (DOCUMENT_TEMPLATE_TYPE,),
        ).fetchall()
    return [document_template_to_dict(row) for row in rows]


def default_document_template() -> dict:
    templates = list_document_templates()
    return next((item for item in templates if item["is_default"]), templates[0])


def validate_document_template(data: dict, existing: dict | None = None) -> dict:
    name = clean_text(data.get("name", existing.get("name") if existing else ""))
    if not name:
        raise ValueError("Le nom du modèle est obligatoire.")
    if len(name) > 80:
        raise ValueError("Le nom du modèle est trop long.")
    raw_configuration = data.get("configuration")
    configuration = (
        normalize_note_template_configuration(raw_configuration)
        if raw_configuration is not None
        else normalize_note_template_configuration((existing or {}).get("configuration"))
    )
    return {
        "name": name,
        "configuration": configuration,
        "is_default": bool(data.get("is_default", (existing or {}).get("is_default", False))),
    }


def create_document_template(data: dict) -> dict:
    item = validate_document_template(data)
    stamp = now_stamp()
    with db_connection() as db:
        count = int(
            db.execute(
                "SELECT COUNT(*) FROM document_templates WHERE document_type = ?",
                (DOCUMENT_TEMPLATE_TYPE,),
            ).fetchone()[0]
        )
        is_default = item["is_default"] or count == 0
        if is_default:
            db.execute(
                "UPDATE document_templates SET is_default = 0 WHERE document_type = ?",
                (DOCUMENT_TEMPLATE_TYPE,),
            )
        try:
            cursor = db.execute(
                """
                INSERT INTO document_templates
                    (name, document_type, is_default, configuration_json, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    item["name"],
                    DOCUMENT_TEMPLATE_TYPE,
                    1 if is_default else 0,
                    json.dumps(item["configuration"], ensure_ascii=False),
                    stamp,
                    stamp,
                ),
            )
        except sqlite3.IntegrityError as exc:
            raise ValueError("Un modèle porte déjà ce nom.") from exc
        row = db.execute("SELECT * FROM document_templates WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return document_template_to_dict(row)


def update_document_template(template_id: int, data: dict) -> dict:
    with db_connection() as db:
        row = db.execute("SELECT * FROM document_templates WHERE id = ?", (template_id,)).fetchone()
    if not row:
        raise KeyError("Modèle de note introuvable.")
    existing = document_template_to_dict(row)
    item = validate_document_template(data, existing)
    with db_connection() as db:
        if item["is_default"]:
            db.execute(
                "UPDATE document_templates SET is_default = 0 WHERE document_type = ?",
                (DOCUMENT_TEMPLATE_TYPE,),
            )
        elif existing["is_default"]:
            # Le modèle actif reste actif tant qu'un autre n'est pas choisi.
            item["is_default"] = True
        try:
            db.execute(
                """
                UPDATE document_templates
                SET name = ?, is_default = ?, configuration_json = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    item["name"],
                    1 if item["is_default"] else 0,
                    json.dumps(item["configuration"], ensure_ascii=False),
                    now_stamp(),
                    template_id,
                ),
            )
        except sqlite3.IntegrityError as exc:
            raise ValueError("Un modèle porte déjà ce nom.") from exc
        updated = db.execute("SELECT * FROM document_templates WHERE id = ?", (template_id,)).fetchone()
    return document_template_to_dict(updated)


def duplicate_document_template(template_id: int) -> dict:
    with db_connection() as db:
        row = db.execute("SELECT * FROM document_templates WHERE id = ?", (template_id,)).fetchone()
        names = {
            str(item["name"]).casefold()
            for item in db.execute(
                "SELECT name FROM document_templates WHERE document_type = ?",
                (DOCUMENT_TEMPLATE_TYPE,),
            ).fetchall()
        }
    if not row:
        raise KeyError("Modèle de note introuvable.")
    source = document_template_to_dict(row)
    base_name = f"{source['name']} - copie"
    name = base_name
    suffix = 2
    while name.casefold() in names:
        name = f"{base_name} {suffix}"
        suffix += 1
    return create_document_template({"name": name, "configuration": source["configuration"]})


def delete_document_template(template_id: int) -> None:
    with db_connection() as db:
        row = db.execute("SELECT * FROM document_templates WHERE id = ?", (template_id,)).fetchone()
        if not row:
            raise KeyError("Modèle de note introuvable.")
        count = int(
            db.execute(
                "SELECT COUNT(*) FROM document_templates WHERE document_type = ?",
                (DOCUMENT_TEMPLATE_TYPE,),
            ).fetchone()[0]
        )
        if count <= 1:
            raise ValueError("Le dernier modèle de note ne peut pas être supprimé.")
        was_default = bool(row["is_default"])
        db.execute("DELETE FROM document_templates WHERE id = ?", (template_id,))
        if was_default:
            replacement = db.execute(
                """
                SELECT id FROM document_templates
                WHERE document_type = ?
                ORDER BY id LIMIT 1
                """,
                (DOCUMENT_TEMPLATE_TYPE,),
            ).fetchone()
            db.execute("UPDATE document_templates SET is_default = 1 WHERE id = ?", (replacement["id"],))


def available_document_template_name(base_name: str) -> str:
    with db_connection() as db:
        names = {
            str(row["name"]).casefold()
            for row in db.execute(
                "SELECT name FROM document_templates WHERE document_type = ?",
                (DOCUMENT_TEMPLATE_TYPE,),
            ).fetchall()
        }
    candidate = clean_text(base_name) or "Modèle importé"
    suffix = 2
    while candidate.casefold() in names:
        candidate = f"{clean_text(base_name) or 'Modèle importé'} {suffix}"
        suffix += 1
    return candidate


def export_document_template(template_id: int, destination_dir: object) -> Path:
    template = next((item for item in list_document_templates() if item["id"] == template_id), None)
    if not template:
        raise KeyError("Modèle de note introuvable.")
    folder = ensure_folder(Path(str(destination_dir)).expanduser())
    filename = re.sub(r"[^A-Za-z0-9._-]+", "-", normalize_name(template["name"])).strip("-") or "modele-note"
    path = folder / f"EasyCESU-Modele-{filename}.json"
    payload = {
        "format": "easy-cesu-document-template",
        "format_version": 1,
        "document_type": DOCUMENT_TEMPLATE_TYPE,
        "name": template["name"],
        "configuration": template["configuration"],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def import_document_template(source_file: object) -> dict:
    path = Path(str(source_file)).expanduser()
    if not path.is_file():
        raise ValueError("Fichier de modèle introuvable.")
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError("Le fichier sélectionné n'est pas un modèle Easy CESU valide.") from exc
    if payload.get("format") != "easy-cesu-document-template" or payload.get("document_type") != DOCUMENT_TEMPLATE_TYPE:
        raise ValueError("Le fichier sélectionné n'est pas un modèle de note Easy CESU.")
    name = available_document_template_name(str(payload.get("name") or "Modèle importé"))
    return create_document_template({"name": name, "configuration": payload.get("configuration")})


def generate_document_template_test_pdf(configuration: object, destination_dir: object) -> Path:
    folder = ensure_folder(Path(str(destination_dir)).expanduser())
    output = folder / "Apercu modele note Easy CESU.pdf"
    sample_rows = [
        Intervention(
            annee=2026,
            mois=7,
            mois_libelle="Juillet",
            client="Client exemple",
            date="2026-07-03",
            duree_heures=2.0,
            salaire_net_horaire=22.0,
            montant_net=44.0,
            montant_brut=0.0,
            cesu="",
            adresse="",
            source_classeur="",
            source_onglet="",
        ),
        Intervention(
            annee=2026,
            mois=7,
            mois_libelle="Juillet",
            client="Client exemple",
            date="2026-07-10",
            duree_heures=1.5,
            salaire_net_horaire=22.0,
            montant_net=33.0,
            montant_brut=0.0,
            cesu="",
            adresse="",
            source_classeur="",
            source_onglet="",
        ),
    ]
    font, bold_font = register_fonts()
    generate_note_pdf(
        output,
        "Client exemple",
        "Juillet 2026",
        sample_rows,
        font,
        bold_font,
        active_employee_lines(),
        normalize_note_template_configuration(configuration),
    )
    return output


def note_to_dict(row: sqlite3.Row) -> dict:
    item = dict(row)
    item["carry_forward"] = bool(item.get("carry_forward"))
    item["is_private"] = bool(item.get("is_private"))
    return item


def validate_note(data: dict) -> tuple[dict, str | None]:
    body = clean_text(data.get("body"))
    if not body:
        return {}, "Le texte de la note est obligatoire."
    category = clean_text(data.get("category")) or "information"
    status = clean_text(data.get("status")) or "information"
    priority = clean_text(data.get("priority")) or "normale"
    if category not in NOTE_CATEGORIES or status not in NOTE_STATUSES or priority not in NOTE_PRIORITIES:
        return {}, "Catégorie, statut ou priorité de note invalide."
    intervention_id = data.get("intervention_id")
    try:
        intervention_id = int(intervention_id) if intervention_id not in (None, "") else None
    except (TypeError, ValueError):
        return {}, "Intervention liée invalide."
    client_name = clean_text(data.get("client_name"))
    if intervention_id is not None:
        with db_connection() as db:
            intervention = db.execute("SELECT client FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
        if not intervention:
            return {}, "Intervention liée introuvable."
        client_name = client_name or str(intervention["client"])
    if not client_name:
        return {}, "Client ou intervention obligatoire pour une note."
    reminder_date = clean_text(data.get("reminder_date"))
    if reminder_date:
        try:
            reminder_date = datetime.fromisoformat(reminder_date).date().isoformat()
        except ValueError:
            return {}, "Date de rappel invalide."
    return {
        "intervention_id": intervention_id,
        "client_name": client_name,
        "body": body,
        "category": category,
        "priority": priority,
        "status": status,
        "reminder_date": reminder_date,
        "carry_forward": 1 if data.get("carry_forward") else 0,
        "is_private": 1 if data.get("is_private") else 0,
    }, None


def list_notes(client_name: str = "", intervention_id: int | None = None, status: str = "") -> list[dict]:
    query = "SELECT * FROM intervention_notes WHERE 1=1"
    params: list[object] = []
    if client_name:
        query += " AND client_name = ?"
        params.append(client_name)
    if intervention_id is not None:
        query += " AND intervention_id = ?"
        params.append(intervention_id)
    if status:
        query += " AND status = ?"
        params.append(status)
    query += " ORDER BY CASE priority WHEN 'haute' THEN 0 WHEN 'normale' THEN 1 ELSE 2 END, reminder_date, updated_at DESC"
    with db_connection() as db:
        return [note_to_dict(row) for row in db.execute(query, params).fetchall()]


def create_note(data: dict) -> dict:
    item, error = validate_note(data)
    if error:
        raise ValueError(error)
    stamp = now_stamp()
    with db_connection() as db:
        cursor = db.execute(
            """
            INSERT INTO intervention_notes
                (intervention_id, client_name, body, category, priority, status, reminder_date, carry_forward, is_private, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (*item.values(), stamp, stamp),
        )
        row = db.execute("SELECT * FROM intervention_notes WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return note_to_dict(row)


def update_note(note_id: int, data: dict) -> dict:
    with db_connection() as db:
        existing = db.execute("SELECT * FROM intervention_notes WHERE id = ?", (note_id,)).fetchone()
    if not existing:
        raise KeyError("Note introuvable.")
    merged = dict(existing)
    merged.update(data)
    item, error = validate_note(merged)
    if error:
        raise ValueError(error)
    with db_connection() as db:
        db.execute(
            """
            UPDATE intervention_notes
            SET intervention_id = ?, client_name = ?, body = ?, category = ?, priority = ?, status = ?,
                reminder_date = ?, carry_forward = ?, is_private = ?, updated_at = ?
            WHERE id = ?
            """,
            (*item.values(), now_stamp(), note_id),
        )
        row = db.execute("SELECT * FROM intervention_notes WHERE id = ?", (note_id,)).fetchone()
    return note_to_dict(row)


def delete_note(note_id: int) -> None:
    with db_connection() as db:
        if db.execute("DELETE FROM intervention_notes WHERE id = ?", (note_id,)).rowcount == 0:
            raise KeyError("Note introuvable.")


def payment_to_dict(row: sqlite3.Row) -> dict:
    item = dict(row)
    item["expected_amount"] = round(float(item.get("expected_amount") or 0), 2)
    item["received_amount"] = round(float(item.get("received_amount") or 0), 2)
    return item


def validate_payment(data: dict) -> tuple[dict, str | None]:
    client_name = clean_text(data.get("client_name"))
    intervention_id = data.get("intervention_id")
    try:
        intervention_id = int(intervention_id) if intervention_id not in (None, "") else None
        expected_amount = float(data.get("expected_amount", 0) or 0)
        received_amount = float(data.get("received_amount", 0) or 0)
    except (TypeError, ValueError):
        return {}, "Montant ou intervention de paiement invalide."
    if expected_amount < 0 or received_amount < 0:
        return {}, "Les montants ne peuvent pas être négatifs."
    if intervention_id is not None:
        with db_connection() as db:
            intervention = db.execute("SELECT client, duration_hours * hourly_rate AS amount FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
        if not intervention:
            return {}, "Intervention liée introuvable."
        client_name = client_name or str(intervention["client"])
        if expected_amount == 0:
            expected_amount = float(intervention["amount"] or 0)
    if not client_name or expected_amount <= 0:
        return {}, "Client et montant attendu obligatoires."
    status = clean_text(data.get("status")) or "a_recevoir"
    if status not in PAYMENT_STATUSES:
        return {}, "Statut de paiement invalide."
    expected_date = clean_text(data.get("expected_date"))
    if expected_date:
        try:
            expected_date = datetime.fromisoformat(expected_date).date().isoformat()
        except ValueError:
            return {}, "Date de paiement attendue invalide."
    return {
        "intervention_id": intervention_id,
        "client_name": client_name,
        "expected_amount": round(expected_amount, 2),
        "received_amount": round(received_amount, 2),
        "expected_date": expected_date,
        "payment_method": clean_text(data.get("payment_method")),
        "status": status,
        "comment": clean_text(data.get("comment")),
    }, None


def list_pending_payments(status: str = "") -> list[dict]:
    query = "SELECT * FROM pending_payments"
    params: list[object] = []
    if status:
        query += " WHERE status = ?"
        params.append(status)
    query += " ORDER BY CASE status WHEN 'a_recevoir' THEN 0 WHEN 'partiellement_recu' THEN 1 ELSE 2 END, expected_date, updated_at DESC"
    with db_connection() as db:
        return [payment_to_dict(row) for row in db.execute(query, params).fetchall()]


def create_pending_payment(data: dict) -> dict:
    item, error = validate_payment(data)
    if error:
        raise ValueError(error)
    stamp = now_stamp()
    with db_connection() as db:
        cursor = db.execute(
            """
            INSERT INTO pending_payments
                (intervention_id, client_name, expected_amount, received_amount, expected_date, payment_method, status, comment, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (*item.values(), stamp, stamp),
        )
        row = db.execute("SELECT * FROM pending_payments WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return payment_to_dict(row)


def update_pending_payment(payment_id: int, data: dict) -> dict:
    with db_connection() as db:
        existing = db.execute("SELECT * FROM pending_payments WHERE id = ?", (payment_id,)).fetchone()
    if not existing:
        raise KeyError("Paiement en attente introuvable.")
    merged = dict(existing)
    merged.update(data)
    item, error = validate_payment(merged)
    if error:
        raise ValueError(error)
    with db_connection() as db:
        db.execute(
            """
            UPDATE pending_payments
            SET intervention_id = ?, client_name = ?, expected_amount = ?, received_amount = ?, expected_date = ?,
                payment_method = ?, status = ?, comment = ?, updated_at = ?
            WHERE id = ?
            """,
            (*item.values(), now_stamp(), payment_id),
        )
        row = db.execute("SELECT * FROM pending_payments WHERE id = ?", (payment_id,)).fetchone()
    return payment_to_dict(row)


def delete_pending_payment(payment_id: int) -> None:
    with db_connection() as db:
        if db.execute("DELETE FROM pending_payments WHERE id = ?", (payment_id,)).rowcount == 0:
            raise KeyError("Paiement en attente introuvable.")


def month_summary(year: int, month: int) -> dict:
    rows = list_interventions(year, month)
    clients = {row["client"] for row in rows}
    hours = sum(float(row["duration_hours"]) for row in rows)
    net = sum(float(row["amount_net"]) for row in rows)
    return {
        "year": year,
        "month": month,
        "month_label": f"{MONTH_LABELS.get(month, month)} {year}",
        "interventions": len(rows),
        "clients": len(clients),
        "hours": round(hours, 4),
        "hours_label": hours_label(hours),
        "amount_net": money(net),
        "rows": rows,
    }


def year_overview(year: int) -> dict:
    """Prépare les données légères affichées dans l'onglet Bilan."""
    monthly = {
        month: {
            "month": month,
            "label": MONTH_LABELS[month],
            "interventions": 0,
            "clients": 0,
            "hours": 0.0,
            "amount_net": 0.0,
        }
        for month in range(1, 13)
    }
    with db_connection() as db:
        rows = db.execute(
            """
            SELECT
                CAST(substr(date, 6, 2) AS INTEGER) AS month,
                COUNT(*) AS interventions,
                COUNT(DISTINCT client) AS clients,
                COALESCE(SUM(duration_hours), 0) AS hours,
                COALESCE(SUM(duration_hours * hourly_rate), 0) AS amount_net
            FROM interventions
            WHERE substr(date, 1, 4) = ?
            GROUP BY substr(date, 6, 2)
            ORDER BY month
            """,
            (f"{year:04d}",),
        ).fetchall()
    for row in rows:
        item = monthly[int(row["month"])]
        item.update(
            interventions=int(row["interventions"]),
            clients=int(row["clients"]),
            hours=round(float(row["hours"]), 4),
            amount_net=money(float(row["amount_net"])),
        )

    months = list(monthly.values())
    active_months = [item for item in months if item["interventions"]]
    total_hours = sum(item["hours"] for item in months)
    total_net = sum(item["amount_net"] for item in months)
    total_interventions = sum(item["interventions"] for item in months)
    total_clients = len({row["client"] for row in list_interventions(year)})
    busiest = max(active_months, key=lambda item: item["amount_net"], default=None)
    return {
        "year": year,
        "months": months,
        "totals": {
            "hours": round(total_hours, 4),
            "hours_label": hours_label(total_hours),
            "amount_net": money(total_net),
            "interventions": total_interventions,
            "clients": total_clients,
            "clients_month": sum(item["clients"] for item in months),
            "active_months": len(active_months),
            "average_net": money(total_net / len(active_months)) if active_months else 0.0,
        },
        "busiest_month": busiest,
    }


def comparison_reference_period(start: date, end: date, mode: str) -> tuple[date, date]:
    """Calcule une période de référence de même durée pour les comparaisons."""
    if mode == "last_year":
        def previous_year(value: date) -> date:
            try:
                return value.replace(year=value.year - 1)
            except ValueError:  # 29 février : comparaison avec le 28 février.
                return value.replace(year=value.year - 1, day=28)

        return previous_year(start), previous_year(end)
    duration = end - start
    reference_end = start - timedelta(days=1)
    return reference_end - duration, reference_end


def period_key(value: date, granularity: str) -> tuple:
    if granularity == "day":
        return (value,)
    if granularity == "week":
        return (value - timedelta(days=value.weekday()),)
    if granularity == "month":
        return (value.year, value.month)
    return (value.year,)


def period_label(key: tuple, granularity: str) -> str:
    if granularity == "day":
        return key[0].strftime("%d/%m")
    if granularity == "week":
        return f"Semaine du {key[0].strftime('%d/%m')}"
    if granularity == "month":
        return f"{MONTH_LABELS[key[1]]} {key[0]}"
    return str(key[0])


def period_id(key: tuple) -> str:
    return "|".join(value.isoformat() if isinstance(value, date) else str(value) for value in key)


def period_slots(start: date, end: date, granularity: str) -> list[tuple]:
    """Liste aussi les périodes vides pour que les graphiques restent comparables."""
    slots = []
    if granularity == "day":
        current = start
        while current <= end:
            slots.append((current,))
            current += timedelta(days=1)
    elif granularity == "week":
        current = start - timedelta(days=start.weekday())
        while current <= end:
            slots.append((current,))
            current += timedelta(days=7)
    elif granularity == "month":
        current = date(start.year, start.month, 1)
        while current <= end:
            slots.append((current.year, current.month))
            current = date(current.year + (current.month == 12), (current.month % 12) + 1, 1)
    else:
        for year in range(start.year, end.year + 1):
            slots.append((year,))
    return slots


def aggregate_period(start: date, end: date, granularity: str) -> list[dict]:
    """Regroupe les interventions par jour, semaine, mois ou année."""
    with db_connection() as db:
        rows = db.execute(
            """
            SELECT date, client, duration_hours, hourly_rate
            FROM interventions
            WHERE date >= ? AND date <= ?
            ORDER BY date, id
            """,
            (start.isoformat(), end.isoformat()),
        ).fetchall()
    grouped: dict[tuple, dict] = {}
    for row in rows:
        when = datetime.fromisoformat(row["date"]).date()
        key = period_key(when, granularity)
        item = grouped.setdefault(
            key,
            {
                "label": period_label(key, granularity),
                "hours": 0.0,
                "amount_net": 0.0,
                "interventions": 0,
                "clients": set(),
            },
        )
        item["hours"] += float(row["duration_hours"])
        item["amount_net"] += float(row["duration_hours"]) * float(row["hourly_rate"])
        item["interventions"] += 1
        item["clients"].add(row["client"])
    return [
        {
            "period_id": period_id(key),
            "label": item["label"],
            "hours": round(item["hours"], 4),
            "amount_net": money(item["amount_net"]),
            "interventions": item["interventions"],
            "clients": len(item["clients"]),
        }
        for key, item in sorted(grouped.items())
    ]


def comparison_totals(rows: list[dict]) -> dict:
    hours = sum(float(row["hours"]) for row in rows)
    amount_net = sum(float(row["amount_net"]) for row in rows)
    return {
        "hours": round(hours, 4),
        "hours_label": hours_label(hours),
        "amount_net": money(amount_net),
        "interventions": sum(int(row["interventions"]) for row in rows),
        "clients": sum(int(row["clients"]) for row in rows),
    }


def activity_overview(year: int, month: int, granularity: str) -> dict:
    """Prépare le graphique principal selon le niveau de détail demandé."""
    if granularity == "year":
        with db_connection() as db:
            bounds = db.execute("SELECT MIN(date), MAX(date) FROM interventions").fetchone()
        first = datetime.fromisoformat(bounds[0]).date() if bounds and bounds[0] else date(year, 1, 1)
        last = datetime.fromisoformat(bounds[1]).date() if bounds and bounds[1] else date(year, 12, 31)
        start, end = date(first.year, 1, 1), date(last.year, 12, 31)
        scope_label = f"{first.year} à {last.year}" if first.year != last.year else str(first.year)
    elif granularity == "day":
        start = date(year, month, 1)
        end = date(year + (month == 12), (month % 12) + 1, 1) - timedelta(days=1)
        scope_label = f"{MONTH_LABELS[month]} {year}"
    else:
        start, end = date(year, 1, 1), date(year, 12, 31)
        scope_label = str(year)

    values_by_id = {row["period_id"]: row for row in aggregate_period(start, end, granularity)}
    periods = [
        values_by_id.get(
            period_id(slot),
            {
                "label": period_label(slot, granularity),
                "hours": 0.0,
                "amount_net": 0.0,
                "interventions": 0,
                "clients": 0,
            },
        )
        for slot in period_slots(start, end, granularity)
    ]
    totals = comparison_totals(periods)
    totals["clients_periods"] = sum(int(item["clients"]) for item in periods)
    busiest = max(periods, key=lambda item: float(item["hours"]), default=None)
    return {
        "year": year,
        "month": month,
        "granularity": granularity,
        "scope_label": scope_label,
        "periods": periods,
        "totals": totals,
        "busiest": busiest if busiest and busiest["interventions"] else None,
    }


def variation(current: float, reference: float) -> dict:
    difference = current - reference
    return {
        "difference": money(difference),
        "percent": round((difference / reference) * 100, 1) if reference else None,
    }


def comparison_overview(start: date, end: date, granularity: str, reference_mode: str) -> dict:
    current_data = aggregate_period(start, end, granularity)
    reference_start, reference_end = comparison_reference_period(start, end, reference_mode)
    reference_data = aggregate_period(reference_start, reference_end, granularity)
    empty = {"hours": 0.0, "amount_net": 0.0, "interventions": 0, "clients": 0}
    current_by_id = {row["period_id"]: row for row in current_data}
    reference_by_id = {row["period_id"]: row for row in reference_data}
    current_slots = period_slots(start, end, granularity)
    reference_slots = period_slots(reference_start, reference_end, granularity)
    current_rows = [
        current_by_id.get(period_id(slot), {**empty, "label": period_label(slot, granularity)})
        for slot in current_slots
    ]
    reference_rows = [
        reference_by_id.get(period_id(slot), {**empty, "label": period_label(slot, granularity)})
        for slot in reference_slots
    ]
    points = []
    for index in range(max(len(current_rows), len(reference_rows))):
        current = current_rows[index] if index < len(current_rows) else empty
        reference = reference_rows[index] if index < len(reference_rows) else empty
        points.append({"label": current.get("label") or reference.get("label") or "Aucune donnée", "current": current, "reference": reference})
    current_totals = comparison_totals(current_rows)
    reference_totals = comparison_totals(reference_rows)
    return {
        "start": start.isoformat(),
        "end": end.isoformat(),
        "reference_start": reference_start.isoformat(),
        "reference_end": reference_end.isoformat(),
        "granularity": granularity,
        "reference_mode": reference_mode,
        "points": points,
        "current_totals": current_totals,
        "reference_totals": reference_totals,
        "variations": {
            "hours": variation(current_totals["hours"], reference_totals["hours"]),
            "amount_net": variation(current_totals["amount_net"], reference_totals["amount_net"]),
            "interventions": variation(float(current_totals["interventions"]), float(reference_totals["interventions"])),
        },
    }


def to_intervention_objects(rows: list[dict]) -> list[Intervention]:
    client_map = {row["name"]: row for row in clients_list()}
    objects: list[Intervention] = []
    for row in rows:
        when = datetime.fromisoformat(row["date"]).date()
        client = client_map.get(row["client"], {})
        amount_net = money(float(row["duration_hours"]) * float(row["hourly_rate"]))
        objects.append(
            Intervention(
                annee=when.year,
                mois=when.month,
                mois_libelle=f"{MONTH_LABELS[when.month]} {when.year}",
                client=row["client"],
                date=when.isoformat(),
                duree_heures=float(row["duration_hours"]),
                salaire_net_horaire=float(row["hourly_rate"]),
                montant_net=amount_net,
                montant_brut=money(amount_net * DEFAULT_BRUT_COEFF),
                cesu=client.get("cesu", ""),
                adresse=client.get("address", ""),
                source_classeur="Application de saisie",
                source_onglet="Saisie",
            )
        )
    return objects


def generate_month_notes(year: int, month: int, replace: bool = False, output_dir: object = None) -> dict:
    rows = list_interventions(year, month)
    if not rows:
        raise ValueError("Aucune intervention à générer pour ce mois.")
    objects = to_intervention_objects(rows)
    base_dir = set_notes_output_dir(output_dir, create=True) if output_dir else ensure_notes_output_dir(notes_output_dir())
    template = default_document_template()
    result = generate_notes(
        objects,
        base_dir,
        overwrite=replace,
        employee_lines=active_employee_lines(),
        template_configuration=template["configuration"],
    )
    return {
        "month_label": f"{MONTH_LABELS[month]} {year}",
        "output_base": str(base_dir),
        "output_dir": str(base_dir),
        "summary": month_summary(year, month),
        "template": {"id": template["id"], "name": template["name"]},
        "notes": result,
    }


def email_configuration_status(profile: dict | None = None) -> dict:
    selected_profile = profile or active_profile()
    try:
        settings = normalize_smtp_settings(selected_profile, require_complete=True)
    except ValueError as exc:
        return {"ready": False, "message": str(exc)}
    if settings["smtp_username"] and not password_saved(selected_profile["id"]):
        return {
            "ready": False,
            "message": "Renseigne le mot de passe SMTP dans les réglages.",
        }
    return {"ready": True, "message": "Configuration email prête."}


def test_email_configuration() -> dict:
    profile = active_profile()
    settings = normalize_smtp_settings(profile, require_complete=True)
    password = get_smtp_password(profile["id"])
    with smtp_connection(settings, password):
        pass
    return {"ok": True, "message": "Connexion au serveur de messagerie réussie."}


def month_email_preview(year: int, month: int) -> dict:
    rows = list_interventions(year, month)
    if not rows:
        raise ValueError("Aucune intervention pour ce mois.")
    clients = {item["name"]: item for item in clients_list()}
    profile = active_profile()
    settings = normalize_smtp_settings(profile)
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        grouped.setdefault(row["client"], []).append(row)
    recipients = []
    for client_name in sorted(grouped, key=sorted_name_key):
        client_rows = grouped[client_name]
        client = clients.get(client_name, {})
        email = clean_text(client.get("email"))
        total_hours = sum(float(row["duration_hours"]) for row in client_rows)
        total_amount = sum(
            float(row["duration_hours"]) * float(row["hourly_rate"])
            for row in client_rows
        )
        values = {
            "client": client_name,
            "mois": MONTH_LABELS[month],
            "annee": str(year),
            "heures": hours_label(total_hours),
            "montant": french_money(total_amount),
            "nom": clean_text(settings["smtp_sender_name"])
            or clean_text(profile.get("name"))
            or clean_text(profile.get("label")),
        }
        recipients.append(
            {
                "client": client_name,
                "email": email,
                "selectable": bool(email),
                "selected": bool(email and client.get("email_notes_enabled")),
                "review_before_send": bool(client.get("email_review_before_send")),
                "hours": values["heures"],
                "amount": values["montant"],
                "interventions": len(client_rows),
                "subject": render_email_template(settings["email_subject_template"], values),
                "body": render_email_template(settings["email_body_template"], values),
            }
        )
    return {
        "month_label": f"{MONTH_LABELS[month]} {year}",
        "configuration": email_configuration_status(),
        "recipients": recipients,
    }


def mark_month_interventions_transmitted(client_name: str, year: int, month: int) -> int:
    with db_connection() as db:
        rows = db.execute(
            """
            SELECT id FROM interventions
            WHERE client = ? AND substr(date, 1, 4) = ? AND substr(date, 6, 2) = ? AND transmitted = 0
            """,
            (client_name, f"{year:04d}", f"{month:02d}"),
        ).fetchall()
        intervention_ids = [int(row["id"]) for row in rows]
        if not intervention_ids:
            return 0
        placeholders = ", ".join("?" for _ in intervention_ids)
        db.execute(
            f"UPDATE interventions SET transmitted = 1, updated_at = ? WHERE id IN ({placeholders})",
            (now_stamp(), *intervention_ids),
        )
        db.execute(
            f"DELETE FROM intervention_followup_ignores WHERE reminder_type = 'transmitted' AND intervention_id IN ({placeholders})",
            intervention_ids,
        )
    return len(intervention_ids)


def send_month_emails(
    year: int,
    month: int,
    client_names: object,
    message_overrides: object = None,
    mark_transmitted: bool = False,
) -> dict:
    if not isinstance(client_names, list):
        raise ValueError("La liste des clients à contacter est invalide.")
    requested = {clean_text(name) for name in client_names if clean_text(name)}
    if not requested:
        raise ValueError("Coche au moins un client.")
    if message_overrides is None:
        message_overrides = {}
    if not isinstance(message_overrides, dict):
        raise ValueError("Les modifications des mails sont invalides.")

    preview = month_email_preview(year, month)
    candidates = {item["client"]: item for item in preview["recipients"]}
    unknown = requested - set(candidates)
    if unknown:
        raise ValueError("Un client sélectionné n'appartient pas au mois choisi.")
    missing_email = sorted(name for name in requested if not candidates[name]["selectable"])
    if missing_email:
        raise ValueError(f"Adresse email manquante pour : {', '.join(missing_email)}.")

    profile = active_profile()
    settings = normalize_smtp_settings(profile, require_complete=True)
    password = get_smtp_password(profile["id"])
    if settings["smtp_username"] and not password:
        raise ValueError("Renseigne le mot de passe SMTP dans les réglages.")

    generated = generate_month_notes(year, month, replace=False)
    notes_folder = Path(generated["output_dir"])
    sent: list[dict] = []
    errors: list[dict] = []
    with smtp_connection(settings, password) as server:
        for client_name in sorted(requested, key=sorted_name_key):
            recipient = candidates[client_name]
            try:
                note_path = matching_existing_note(notes_folder, year, month, client_name)
                if note_path is None:
                    raise ValueError("Le PDF de la note d'intervention est introuvable.")
                override = message_overrides.get(client_name, {})
                if not isinstance(override, dict):
                    override = {}
                subject = clean_text(override.get("subject")) or recipient["subject"]
                body = str(override.get("body") or recipient["body"]).strip()
                if not subject or not body:
                    raise ValueError("L'objet et le texte du mail ne peuvent pas être vides.")
                message = build_email_message(
                    settings,
                    recipient["email"],
                    subject,
                    body,
                    note_path,
                )
                server.send_message(message)
                transmitted_updated = (
                    mark_month_interventions_transmitted(client_name, year, month) if mark_transmitted else 0
                )
                sent.append(
                    {
                        "client": client_name,
                        "email": recipient["email"],
                        "note": str(note_path),
                        "transmitted_updated": transmitted_updated,
                    }
                )
            except Exception as exc:  # noqa: BLE001 - une erreur client ne bloque pas les suivants
                errors.append({"client": client_name, "error": str(exc)})
    return {
        "month_label": preview["month_label"],
        "sent": sent,
        "errors": errors,
        "output_dir": str(notes_folder),
    }


def payload_for_year(year: int) -> dict:
    rows = list_interventions(year)
    objects = to_intervention_objects(rows)
    summaries = build_summaries(objects)
    return {
        "generated_at": now_stamp(),
        "year": year,
        "month": None,
        "sources": {
            "suivi_paye": "Application de saisie",
            "fichier_clients": active_profile().get("fichier_clients", ""),
            "notes_intervention_dir": str(notes_output_dir()),
            "export_dir": str(export_output_dir()),
        },
        "parameters": {
            "salaire_net_horaire": DEFAULT_RATE,
            "coefficient_brut": DEFAULT_BRUT_COEFF,
            "replace_notes": False,
            "pdf_notes_enabled": True,
        },
        "interventions": [item.__dict__ for item in objects],
        "summaries": summaries,
        "anomalies": [],
        "notes": {"created": [], "skipped": [], "errors": []},
    }


def export_year(year: int, output_dir: object = None) -> dict:
    if output_dir:
        set_export_output_dir(output_dir, create=True)
    payload = payload_for_year(year)
    sorties = ensure_folder(export_output_dir())
    data_path = sorties / f"donnees_application_{year}.json"
    xlsx_path = sorties / f"Bilan activite application {year}.xlsx"
    data_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        export_bilan_excel(payload, xlsx_path, data_path)
    except PermissionError as exc:
        raise ValueError(f"Impossible d'ecrire le fichier Excel. Ferme le classeur s'il est ouvert : {xlsx_path}") from exc
    return {"data": str(data_path), "xlsx": str(xlsx_path), "totals": payload["summaries"]["totals"]}


def json_response(handler: SimpleHTTPRequestHandler, payload: object, status: int = 200) -> None:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def read_json_body(handler: SimpleHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0") or 0)
    if not length:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    return json.loads(raw)


def query_int(query: dict[str, list[str]], name: str, default: int | None = None) -> int | None:
    value = query.get(name, [None])[0]
    if value in (None, ""):
        return default
    return int(value)


def port_is_listening(port: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(0.2)
        return sock.connect_ex(("127.0.0.1", port)) == 0


def local_server_info(port: int) -> dict:
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:{port}/api/app-info", timeout=0.5) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception:
        return {}


def is_current_easy_cesu_server(port: int) -> bool:
    info = local_server_info(port)
    if info.get("app_name") != APP_NAME or info.get("app_version") != APP_VERSION:
        return False
    if not getattr(sys, "frozen", False):
        return True
    server_executable = Path(str(info.get("executable") or "")).resolve(strict=False)
    return server_executable == Path(sys.executable).resolve(strict=False)


def select_server_port(preferred_port: int, attempts: int = 50) -> tuple[int, bool]:
    # Chaque projet peut utiliser Chrome et son propre serveur sans partager le même port.
    for port in range(preferred_port, preferred_port + attempts):
        if not port_is_listening(port):
            return port, False
        if is_current_easy_cesu_server(port):
            return port, True
    raise RuntimeError("Aucun port local disponible pour Easy CESU.")


def request_server_stop(server: ThreadingHTTPServer) -> None:
    """Réveille et arrête la boucle HTTP sans dépendre d'un shutdown bloquant."""

    setattr(server, "_BaseServer__shutdown_request", True)
    try:
        with socket.create_connection(server.server_address, timeout=0.2):
            pass
    except OSError:
        pass
    try:
        server.socket.close()
    except OSError:
        pass


class LocalAppServer:
    """Pilote le serveur HTTP local indépendamment de la fenêtre d'affichage."""

    def __init__(self, preferred_port: int | None = None) -> None:
        self.preferred_port = (
            int(os.environ.get("NOTES_APP_PORT", "8765"))
            if preferred_port is None
            else preferred_port
        )
        self.port = 0
        self.url = ""
        self.browser_url = ""
        self.existing_server = False
        self.server: ThreadingHTTPServer | None = None
        self.monitor_stop: threading.Event | None = None
        self.server_thread: threading.Thread | None = None
        self.shutdown_timeout_seconds = 5.0

    def start(self, background: bool = True) -> str:
        if self.preferred_port == 0:
            # Le mode fenêtre native laisse Windows choisir un port libre afin de
            # ne jamais concurrencer les autres serveurs locaux de l'utilisateur.
            self.port, self.existing_server = 0, False
        else:
            self.port, self.existing_server = select_server_port(self.preferred_port)
        if self.existing_server:
            self.url = f"http://127.0.0.1:{self.port}"
            self.browser_url = f"{self.url}/?v=20260801-v314"
            return self.browser_url

        init_db()
        self.server = ThreadingHTTPServer(("127.0.0.1", self.port), AppHandler)
        self.port = int(self.server.server_address[1])
        self.url = f"http://127.0.0.1:{self.port}"
        self.browser_url = f"{self.url}/?v=20260801-v314"
        self.server.daemon_threads = True
        self.server.block_on_close = False
        self.monitor_stop = threading.Event()
        threading.Thread(
            target=monitor_browser_sessions,
            args=(self.server, self.monitor_stop),
            name="easy-cesu-session-monitor",
            daemon=True,
        ).start()

        if background:
            self.server_thread = threading.Thread(
                target=self.server.serve_forever,
                name="easy-cesu-local-server",
                daemon=True,
            )
            self.server_thread.start()
        return self.browser_url

    def serve_forever(self) -> None:
        if self.existing_server:
            return
        if self.server is None:
            self.start(background=False)
        if self.server is not None:
            self.server.serve_forever()

    def stop(self) -> None:
        # Une seconde fenêtre peut utiliser un serveur déjà lancé : elle ne doit pas l'arrêter.
        if self.existing_server or self.server is None:
            return
        if self.monitor_stop is not None:
            self.monitor_stop.set()
        server = self.server
        request_server_stop(server)
        if self.server_thread is not None and self.server_thread.is_alive():
            self.server_thread.join(timeout=self.shutdown_timeout_seconds)
        self.server = None


def update_browser_session(payload: dict) -> None:
    # Chaque onglet possède sa propre session afin qu'un onglet fermé n'arrête pas les autres.
    global BROWSER_CONNECTED_ONCE, BROWSER_EMPTY_SINCE
    session_id = str(payload.get("session_id") or "").strip()
    state = str(payload.get("state") or "active").strip().lower()
    if not session_id or len(session_id) > 100:
        raise ValueError("Session navigateur invalide.")
    now = time.monotonic()
    with BROWSER_SESSION_LOCK:
        if state == "active":
            BROWSER_SESSIONS[session_id] = now
            BROWSER_CONNECTED_ONCE = True
            BROWSER_EMPTY_SINCE = None
        elif state == "closed":
            BROWSER_SESSIONS.pop(session_id, None)
            BROWSER_STREAM_TOKENS.pop(session_id, None)
            if BROWSER_CONNECTED_ONCE and not BROWSER_SESSIONS and BROWSER_EMPTY_SINCE is None:
                BROWSER_EMPTY_SINCE = now
        else:
            raise ValueError("État de session navigateur invalide.")


def register_browser_stream(session_id: str) -> object:
    global BROWSER_CONNECTED_ONCE, BROWSER_EMPTY_SINCE
    if not session_id or len(session_id) > 100:
        raise ValueError("Session navigateur invalide.")
    token = object()
    with BROWSER_SESSION_LOCK:
        BROWSER_SESSIONS[session_id] = time.monotonic()
        BROWSER_STREAM_TOKENS[session_id] = token
        BROWSER_CONNECTED_ONCE = True
        BROWSER_EMPTY_SINCE = None
    return token


def close_browser_stream(session_id: str, token: object) -> None:
    global BROWSER_EMPTY_SINCE
    with BROWSER_SESSION_LOCK:
        # Un ancien flux ne doit jamais fermer la connexion qui l'a remplacé.
        if BROWSER_STREAM_TOKENS.get(session_id) is not token:
            return
        BROWSER_STREAM_TOKENS.pop(session_id, None)
        BROWSER_SESSIONS.pop(session_id, None)
        if BROWSER_CONNECTED_ONCE and not BROWSER_SESSIONS and BROWSER_EMPTY_SINCE is None:
            BROWSER_EMPTY_SINCE = time.monotonic()


def stream_browser_session(handler: SimpleHTTPRequestHandler, session_id: str) -> None:
    # Le flux continu permet de détecter une fermeture brutale, même sans signal JavaScript final.
    token = register_browser_stream(session_id)
    handler.send_response(HTTPStatus.OK)
    handler.send_header("Content-Type", "text/event-stream; charset=utf-8")
    handler.send_header("Connection", "keep-alive")
    handler.end_headers()
    try:
        while True:
            update_browser_session({"session_id": session_id, "state": "active"})
            handler.wfile.write(b": easy-cesu\n\n")
            handler.wfile.flush()
            time.sleep(2.0)
    except (BrokenPipeError, ConnectionResetError, OSError):
        close_browser_stream(session_id, token)


def monitor_browser_sessions(server: ThreadingHTTPServer, stop_event: threading.Event) -> None:
    # Le délai de grâce évite de couper le serveur pendant un simple rechargement de page.
    global BROWSER_EMPTY_SINCE
    while not stop_event.wait(1.0):
        should_stop = False
        now = time.monotonic()
        with BROWSER_SESSION_LOCK:
            stale_ids = [
                session_id
                for session_id, last_seen in BROWSER_SESSIONS.items()
                if now - last_seen > BROWSER_STALE_SECONDS
            ]
            for session_id in stale_ids:
                BROWSER_SESSIONS.pop(session_id, None)
                BROWSER_STREAM_TOKENS.pop(session_id, None)
            if BROWSER_CONNECTED_ONCE:
                if BROWSER_SESSIONS:
                    BROWSER_EMPTY_SINCE = None
                elif BROWSER_EMPTY_SINCE is None:
                    BROWSER_EMPTY_SINCE = now
                elif now - BROWSER_EMPTY_SINCE >= BROWSER_CLOSE_GRACE_SECONDS:
                    should_stop = True
        if should_stop:
            request_server_stop(server)
            return


def open_app_url(url: str) -> None:
    try:
        webbrowser.open(url)
    except Exception:
        pass


class AppHandler(SimpleHTTPRequestHandler):
    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def translate_path(self, path_value: str) -> str:
        parsed = urlparse(path_value)
        path_part = parsed.path
        if path_part == "/":
            return str(STATIC_DIR / "index.html")
        return str(STATIC_DIR / path_part.lstrip("/"))

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        if parsed.path == "/api/app-info":
            json_response(
                self,
                {
                    "app_name": APP_NAME,
                    "app_version": APP_VERSION,
                    "executable": str(Path(sys.executable).resolve()) if getattr(sys, "frozen", False) else "",
                    "runtime_dir": str(RUNTIME_DIR.resolve()),
                },
            )
            return
        if parsed.path == "/api/browser-events":
            stream_browser_session(self, query.get("session_id", [""])[0])
            return
        if parsed.path == "/api/bootstrap":
            json_response(self, bootstrap_payload())
            return
        if parsed.path == "/api/settings":
            json_response(self, {"settings": app_settings()})
            return
        if parsed.path == "/api/community":
            json_response(self, community_info())
            return
        if parsed.path == "/api/profiles":
            json_response(
                self,
                {
                    "profiles": [public_profile(profile) for profile in CONFIG.get("profiles", [])],
                    "active_profile_id": active_profile()["id"],
                },
            )
            return
        if parsed.path == "/api/clients":
            json_response(self, {"clients": clients_list()})
            return
        if parsed.path == "/api/service-categories":
            include_archived = query.get("include_archived", [""])[0].lower() in {"1", "true", "yes"}
            json_response(self, {"categories": list_service_categories(include_archived)})
            return
        if parsed.path == "/api/document-templates":
            json_response(self, {"templates": list_document_templates()})
            return
        if parsed.path == "/api/notes":
            raw_intervention_id = query.get("intervention_id", [""])[0]
            intervention_id = int(raw_intervention_id) if raw_intervention_id else None
            json_response(
                self,
                {"notes": list_notes(query.get("client", [""])[0], intervention_id, query.get("status", [""])[0])},
            )
            return
        if parsed.path == "/api/pending-payments":
            json_response(self, {"payments": list_pending_payments(query.get("status", [""])[0])})
            return
        if parsed.path == "/api/intervention-followups":
            json_response(
                self,
                {
                    "followups": list_intervention_followups(
                        query.get("type", [""])[0],
                        query.get("search", [""])[0],
                        query.get("include_ignored", ["0"])[0] in {"1", "true", "yes"},
                    )
                },
            )
            return
        if parsed.path == "/api/reminders":
            client_name = query.get("client", [""])[0]
            json_response(self, {"reminders": reminders_for_client(client_name)} if client_name else reminders_overview())
            return
        if parsed.path == "/api/reminders/overview":
            json_response(self, reminders_overview(query_int(query, "year"), query_int(query, "month")))
            return
        if parsed.path == "/api/interventions":
            year = query_int(query, "year")
            month = query_int(query, "month")
            client = query.get("client", [""])[0]
            json_response(self, {"interventions": list_interventions(year, month, client)})
            return
        if parsed.path == "/api/summary":
            year = query_int(query, "year", date.today().year)
            month = query_int(query, "month", date.today().month)
            json_response(self, month_summary(year, month))
            return
        if parsed.path == "/api/year-overview":
            year = query_int(query, "year", date.today().year)
            json_response(self, year_overview(year))
            return
        if parsed.path == "/api/activity-overview":
            year = query_int(query, "year", date.today().year)
            month = query_int(query, "month", date.today().month)
            granularity = query.get("granularity", ["month"])[0]
            if not 1 <= month <= 12 or granularity not in {"day", "week", "month", "year"}:
                json_response(self, {"error": "Paramètres de bilan invalides."}, HTTPStatus.BAD_REQUEST)
                return
            json_response(self, activity_overview(year, month, granularity))
            return
        if parsed.path == "/api/comparison-overview":
            try:
                start = datetime.fromisoformat(query.get("start", [""])[0]).date()
                end = datetime.fromisoformat(query.get("end", [""])[0]).date()
            except ValueError:
                json_response(self, {"error": "Dates de comparaison invalides."}, HTTPStatus.BAD_REQUEST)
                return
            if end < start:
                json_response(self, {"error": "La date de fin doit suivre la date de début."}, HTTPStatus.BAD_REQUEST)
                return
            granularity = query.get("granularity", ["month"])[0]
            if granularity not in {"day", "week", "month", "year"}:
                json_response(self, {"error": "Granularité de comparaison invalide."}, HTTPStatus.BAD_REQUEST)
                return
            reference_mode = query.get("reference", ["previous"])[0]
            if reference_mode not in {"previous", "last_year"}:
                json_response(self, {"error": "Période de référence invalide."}, HTTPStatus.BAD_REQUEST)
                return
            json_response(self, comparison_overview(start, end, granularity, reference_mode))
            return
        return super().do_GET()

    def do_POST(self) -> None:  # noqa: N802
        try:
            parsed = urlparse(self.path)
            body = read_json_body(self)
            if parsed.path == "/api/browser-session":
                update_browser_session(body)
                json_response(self, {"ok": True})
                return
            if parsed.path == "/api/shutdown":
                json_response(self, {"ok": True, "message": "Arrêt de Easy CESU."})
                self.wfile.flush()
                threading.Thread(target=request_server_stop, args=(self.server,), daemon=True).start()
                return
            if parsed.path == "/api/open-external":
                json_response(self, open_external_link(body.get("link_id")))
                return
            if parsed.path == "/api/support-reminder":
                json_response(self, {"support_reminder": update_support_reminder(body.get("action"))})
                return
            if parsed.path == "/api/clients":
                json_response(self, {"client": create_client(body)}, HTTPStatus.CREATED)
                return
            if parsed.path == "/api/service-categories":
                json_response(self, {"category": create_service_category(body)}, HTTPStatus.CREATED)
                return
            if parsed.path == "/api/document-templates":
                json_response(self, {"template": create_document_template(body)}, HTTPStatus.CREATED)
                return
            if parsed.path == "/api/document-templates/import":
                source_file = body.get("source_file")
                if not source_file:
                    selected, cancelled = choose_file(
                        "Importer un modèle de note Easy CESU",
                        export_output_dir(),
                        "Modèles Easy CESU (*.json)|*.json|Tous les fichiers (*.*)|*.*",
                    )
                    if cancelled or selected is None:
                        json_response(self, {"cancelled": True})
                        return
                    source_file = str(selected)
                json_response(self, {"cancelled": False, "template": import_document_template(source_file)})
                return
            if parsed.path == "/api/document-templates/preview-pdf":
                destination = body.get("destination_dir")
                if not destination:
                    selected, cancelled = choose_folder(
                        "Choisir le dossier du PDF d'essai",
                        export_output_dir(),
                    )
                    if cancelled or selected is None:
                        json_response(self, {"cancelled": True})
                        return
                    destination = str(selected)
                path = generate_document_template_test_pdf(body.get("configuration"), destination)
                json_response(self, {"cancelled": False, "path": str(path)})
                return
            if parsed.path.startswith("/api/document-templates/") and parsed.path.endswith("/duplicate"):
                template_id = int(parsed.path.split("/")[3])
                json_response(self, {"template": duplicate_document_template(template_id)}, HTTPStatus.CREATED)
                return
            if parsed.path.startswith("/api/document-templates/") and parsed.path.endswith("/reset"):
                template_id = int(parsed.path.split("/")[3])
                json_response(
                    self,
                    {
                        "template": update_document_template(
                            template_id,
                            {"configuration": default_note_template_configuration()},
                        )
                    },
                )
                return
            if parsed.path.startswith("/api/document-templates/") and parsed.path.endswith("/export"):
                template_id = int(parsed.path.split("/")[3])
                destination = body.get("destination_dir")
                if not destination:
                    selected, cancelled = choose_folder(
                        "Choisir le dossier d'export du modèle",
                        export_output_dir(),
                    )
                    if cancelled or selected is None:
                        json_response(self, {"cancelled": True})
                        return
                    destination = str(selected)
                path = export_document_template(template_id, destination)
                json_response(self, {"cancelled": False, "path": str(path)})
                return
            if parsed.path == "/api/notes":
                json_response(self, {"note": create_note(body)}, HTTPStatus.CREATED)
                return
            if parsed.path == "/api/pending-payments":
                json_response(self, {"payment": create_pending_payment(body)}, HTTPStatus.CREATED)
                return
            if parsed.path == "/api/intervention-followups/ignore":
                json_response(
                    self,
                    {
                        "ignore": ignore_intervention_followup(
                            int(body.get("intervention_id")),
                            str(body.get("reminder_type") or ""),
                        )
                    },
                    HTTPStatus.CREATED,
                )
                return
            if parsed.path == "/api/profiles":
                json_response(self, create_profile(body), HTTPStatus.CREATED)
                return
            if parsed.path == "/api/switch-profile":
                json_response(self, switch_profile(str(body.get("profile_id") or "")))
                return
            if parsed.path == "/api/select-notes-dir":
                selected, cancelled = choose_notes_output_dir()
                json_response(
                    self,
                    {
                        "cancelled": cancelled,
                        "settings": app_settings(),
                        "notes_intervention_dir": str(selected) if selected else str(notes_output_dir()),
                    },
                )
                return
            if parsed.path == "/api/select-export-dir":
                selected, cancelled = choose_export_output_dir()
                json_response(
                    self,
                    {
                        "cancelled": cancelled,
                        "settings": app_settings(),
                        "export_dir": str(selected) if selected else str(export_output_dir()),
                    },
                )
                return
            if parsed.path == "/api/select-data-dir":
                selected, cancelled = choose_data_output_dir()
                json_response(
                    self,
                    {
                        "cancelled": cancelled,
                        "settings": app_settings(),
                        "data_dir": str(selected) if selected else str(profile_data_dir(active_profile())),
                    },
                )
                return
            if parsed.path == "/api/select-workspace-root":
                configured, cancelled = choose_workspace_root()
                if cancelled or configured is None:
                    json_response(self, {"cancelled": True, "settings": app_settings()})
                else:
                    json_response(self, {"cancelled": False, **configured})
                return
            if parsed.path == "/api/select-database-file":
                selected, cancelled = choose_database_file()
                json_response(
                    self,
                    {
                        "cancelled": cancelled,
                        "settings": app_settings(),
                        "clients": clients_list(),
                        "database_path": str(selected) if selected else str(active_db_path()),
                    },
                )
                return
            if parsed.path == "/api/import-database":
                source_file = body.get("source_file")
                if not source_file:
                    selected, cancelled = choose_import_database_file()
                    if cancelled or selected is None:
                        json_response(self, {"cancelled": True, "settings": app_settings()})
                        return
                    source_file = str(selected)
                json_response(self, {"cancelled": False, **import_database_from(source_file)})
                return
            if parsed.path == "/api/import-backup":
                source_file = body.get("source_file")
                if not source_file:
                    selected, cancelled = choose_import_backup_file()
                    if cancelled or selected is None:
                        json_response(self, {"cancelled": True, "settings": app_settings()})
                        return
                    source_file = str(selected)
                json_response(self, {"cancelled": False, **restore_profile_from_backup(source_file)})
                return
            if parsed.path == "/api/select-source-dir":
                selected, cancelled = choose_source_data_dir()
                json_response(
                    self,
                    {
                        "cancelled": cancelled,
                        "settings": app_settings(),
                        "suivi_paye_dir": str(selected) if selected else active_profile().get("suivi_paye_dir", ""),
                    },
                )
                return
            if parsed.path == "/api/select-clients-file":
                selected, cancelled = choose_clients_file()
                json_response(
                    self,
                    {
                        "cancelled": cancelled,
                        "settings": app_settings(),
                        "fichier_clients": str(selected) if selected else active_profile().get("fichier_clients", ""),
                    },
                )
                return
            if parsed.path == "/api/backup-database":
                backup_dir = body.get("backup_dir")
                if not backup_dir:
                    selected, cancelled = choose_backup_database_dir()
                    if cancelled or selected is None:
                        json_response(self, {"cancelled": True, "settings": app_settings()})
                        return
                    backup_dir = str(selected)
                json_response(self, {"cancelled": False, **backup_database_to(backup_dir)})
                return
            if parsed.path == "/api/export-backup":
                backup_dir = body.get("backup_dir")
                if not backup_dir:
                    selected, cancelled = choose_backup_database_dir()
                    if cancelled or selected is None:
                        json_response(self, {"cancelled": True, "settings": app_settings()})
                        return
                    backup_dir = str(selected)
                json_response(self, {"cancelled": False, **backup_profile_to(backup_dir)})
                return
            if parsed.path == "/api/reminders":
                json_response(self, {"reminder": create_reminder(body)}, HTTPStatus.CREATED)
                return
            if parsed.path.startswith("/api/reminders/") and parsed.path.endswith("/occurrences"):
                parts = parsed.path.split("/")
                reminder_id = int(parts[3])
                json_response(
                    self,
                    set_reminder_occurrence_status(reminder_id, int(body.get("occurrence_id")), str(body.get("status") or "completed")),
                )
                return
            if parsed.path == "/api/interventions":
                json_response(self, {"intervention": create_intervention(body)}, HTTPStatus.CREATED)
                return
            if parsed.path == "/api/email-preview":
                year = int(body.get("year", date.today().year))
                month = int(body.get("month", date.today().month))
                json_response(self, month_email_preview(year, month))
                return
            if parsed.path == "/api/test-email":
                json_response(self, test_email_configuration())
                return
            if parsed.path == "/api/send-month-emails":
                year = int(body.get("year", date.today().year))
                month = int(body.get("month", date.today().month))
                json_response(
                    self,
                    send_month_emails(
                        year,
                        month,
                        body.get("clients"),
                        body.get("message_overrides"),
                        bool(body.get("mark_transmitted")),
                    ),
                )
                return
            if parsed.path == "/api/generate-month":
                year = int(body.get("year", date.today().year))
                month = int(body.get("month", date.today().month))
                json_response(
                    self,
                    generate_month_notes(
                        year,
                        month,
                        bool(body.get("replace")),
                        body.get("notes_intervention_dir") or body.get("output_dir"),
                    ),
                )
                return
            if parsed.path == "/api/export-year":
                year = int(body.get("year", date.today().year))
                json_response(self, export_year(year, body.get("export_dir") or body.get("output_dir")))
                return
            if parsed.path == "/api/refresh-clients":
                refresh_clients()
                json_response(self, {"clients": clients_list()})
                return
            json_response(self, {"error": "Route inconnue."}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        except KeyError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.NOT_FOUND)
        except Exception as exc:  # noqa: BLE001
            json_response(self, {"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_PUT(self) -> None:  # noqa: N802
        try:
            parsed = urlparse(self.path)
            if parsed.path.startswith("/api/clients/"):
                original_name = unquote(parsed.path.rsplit("/", 1)[1])
                body = read_json_body(self)
                json_response(self, {"client": update_client(original_name, body)})
                return
            if parsed.path.startswith("/api/service-categories/"):
                category_id = int(parsed.path.rsplit("/", 1)[1])
                body = read_json_body(self)
                json_response(self, {"category": update_service_category(category_id, body)})
                return
            if parsed.path.startswith("/api/document-templates/"):
                template_id = int(parsed.path.rsplit("/", 1)[1])
                body = read_json_body(self)
                json_response(self, {"template": update_document_template(template_id, body)})
                return
            if parsed.path.startswith("/api/notes/"):
                note_id = int(parsed.path.rsplit("/", 1)[1])
                body = read_json_body(self)
                json_response(self, {"note": update_note(note_id, body)})
                return
            if parsed.path.startswith("/api/pending-payments/"):
                payment_id = int(parsed.path.rsplit("/", 1)[1])
                body = read_json_body(self)
                json_response(self, {"payment": update_pending_payment(payment_id, body)})
                return
            if parsed.path.startswith("/api/reminders/"):
                reminder_id = int(parsed.path.rsplit("/", 1)[1])
                body = read_json_body(self)
                json_response(self, {"reminder": update_reminder(reminder_id, body)})
                return
            if parsed.path == "/api/settings":
                body = read_json_body(self)
                json_response(self, {"settings": update_settings(body)})
                return
            if parsed.path.startswith("/api/profiles/"):
                profile_id = unquote(parsed.path.rsplit("/", 1)[1])
                body = read_json_body(self)
                json_response(self, update_profile(profile_id, body))
                return
            if parsed.path.startswith("/api/interventions/") and parsed.path.endswith("/administrative-status"):
                intervention_id = int(parsed.path.split("/")[3])
                body = read_json_body(self)
                if not isinstance(body.get("checked"), bool):
                    raise ValueError("La valeur de l'état administratif doit être vraie ou fausse.")
                json_response(
                    self,
                    {
                        "intervention": update_intervention_administrative_status(
                            intervention_id,
                            str(body.get("reminder_type") or ""),
                            body["checked"],
                        )
                    },
                )
                return
            if parsed.path.startswith("/api/interventions/"):
                intervention_id = int(parsed.path.rsplit("/", 1)[1])
                body = read_json_body(self)
                json_response(self, {"intervention": update_intervention(intervention_id, body)})
                return
            json_response(self, {"error": "Route inconnue."}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        except KeyError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.NOT_FOUND)

    def do_DELETE(self) -> None:  # noqa: N802
        try:
            parsed = urlparse(self.path)
            if parsed.path.startswith("/api/service-categories/"):
                category_id = int(parsed.path.rsplit("/", 1)[1])
                category = next((item for item in list_service_categories(True) if item["id"] == category_id), None)
                if category is None:
                    raise KeyError("Prestation introuvable.")
                update_service_category(category_id, {**category, "is_archived": True})
                json_response(self, {"ok": True})
                return
            if parsed.path.startswith("/api/document-templates/"):
                delete_document_template(int(parsed.path.rsplit("/", 1)[1]))
                json_response(self, {"ok": True})
                return
            if parsed.path.startswith("/api/notes/"):
                delete_note(int(parsed.path.rsplit("/", 1)[1]))
                json_response(self, {"ok": True})
                return
            if parsed.path.startswith("/api/pending-payments/"):
                delete_pending_payment(int(parsed.path.rsplit("/", 1)[1]))
                json_response(self, {"ok": True})
                return
            if parsed.path.startswith("/api/profiles/"):
                profile_id = unquote(parsed.path.rsplit("/", 1)[1])
                json_response(self, delete_profile(profile_id))
                return
            if parsed.path.startswith("/api/reminders/"):
                reminder_id = int(parsed.path.rsplit("/", 1)[1])
                delete_reminder(reminder_id)
                json_response(self, {"ok": True})
                return
            if parsed.path.startswith("/api/intervention-followups/") and "/ignores/" in parsed.path:
                parts = parsed.path.strip("/").split("/")
                if len(parts) != 5 or parts[1] != "intervention-followups" or parts[3] != "ignores":
                    raise ValueError("Route de réactivation invalide.")
                json_response(self, reactivate_intervention_followup(int(parts[2]), parts[4]))
                return
            if parsed.path.startswith("/api/interventions/"):
                intervention_id = int(parsed.path.rsplit("/", 1)[1])
                delete_intervention(intervention_id)
                json_response(self, {"ok": True})
                return
            if parsed.path.startswith("/api/clients/"):
                client_name = unquote(parsed.path.rsplit("/", 1)[1])
                delete_client(client_name)
                json_response(self, {"ok": True})
                return
            json_response(self, {"error": "Route inconnue."}, HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        except KeyError as exc:
            json_response(self, {"error": str(exc)}, HTTPStatus.NOT_FOUND)

    def log_message(self, format_value: str, *args: object) -> None:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            print(f"[{stamp}] {self.address_string()} {format_value % args}", flush=True)
        except OSError:
            pass


def main(open_browser: bool | None = None) -> int:
    runtime = LocalAppServer()
    if open_browser is None:
        open_browser = bool(getattr(sys, "frozen", False)) or os.environ.get("NOTES_OPEN_BROWSER") == "1"
    browser_url = runtime.start(background=False)
    if runtime.existing_server:
        print(f"Application deja disponible : {runtime.url}", flush=True)
        if open_browser:
            open_app_url(browser_url)
        return 0
    print(f"Application demarree : {runtime.url}", flush=True)
    if open_browser:
        threading.Timer(0.8, open_app_url, args=[browser_url]).start()
    try:
        runtime.serve_forever()
    finally:
        runtime.stop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
