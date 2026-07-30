from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass, asdict
from datetime import date, datetime, time, timedelta
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


MONTHS = {
    "janvier": 1,
    "fevrier": 2,
    "février": 2,
    "mars": 3,
    "avril": 4,
    "mai": 5,
    "juin": 6,
    "juillet": 7,
    "aout": 8,
    "août": 8,
    "septembre": 9,
    "octobre": 10,
    "novembre": 11,
    "decembre": 12,
    "décembre": 12,
}

MONTH_LABELS = {
    1: "Janvier",
    2: "Février",
    3: "Mars",
    4: "Avril",
    5: "Mai",
    6: "Juin",
    7: "Juillet",
    8: "Août",
    9: "Septembre",
    10: "Octobre",
    11: "Novembre",
    12: "Décembre",
}

EMPLOYEE_LINES = [
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
]

NOTE_TEMPLATE_BLOCKS = ("identity", "title", "table")
DEFAULT_NOTE_TEMPLATE_CONFIGURATION = {
    "version": 1,
    "blocks": list(NOTE_TEMPLATE_BLOCKS),
    "visible": {"identity": True, "title": True, "table": True},
    "labels": {
        "title": "NOTE D’INTERVENTION",
        "employer": "Nom de l’employeur :",
        "month": "Mois concerné :",
        "date": "Dates d’intervention",
        "hourly_rate": "Salaire net",
        "hours": "Nombre d’heure",
        "amount": "Montant total",
        "total": "Total du mois :",
    },
    "page": {
        "left_margin_cm": 2.5,
        "right_margin_cm": 2.5,
        "top_margin_cm": 2.3,
        "bottom_margin_cm": 1.4,
    },
    "typography": {
        "body_size": 11.0,
        "title_size": 11.0,
        "table_size": 10.5,
        "text_color": "#000000",
        "title_color": "#000000",
    },
    "spacing": {
        "identity_after_cm": 1.35,
        "title_after_cm": 2.65,
        "table_after_cm": 0.0,
    },
    "table": {
        "header_background": "#E6E6E6",
        "total_background": "#E6E6E6",
        "border_color": "#000000",
        "row_height_cm": 0.70,
        "minimum_rows": 15,
    },
}


def default_note_template_configuration() -> dict[str, Any]:
    return deepcopy(DEFAULT_NOTE_TEMPLATE_CONFIGURATION)


def normalize_note_template_configuration(value: object) -> dict[str, Any]:
    """Normalise uniquement les réglages de mise en page pris en charge."""

    source = value if isinstance(value, dict) else {}
    result = default_note_template_configuration()

    blocks = source.get("blocks")
    if isinstance(blocks, list):
        selected = [str(item) for item in blocks if str(item) in NOTE_TEMPLATE_BLOCKS]
        result["blocks"] = list(dict.fromkeys(selected + list(NOTE_TEMPLATE_BLOCKS)))

    visible = source.get("visible")
    if isinstance(visible, dict):
        result["visible"]["identity"] = bool(visible.get("identity", True))
        result["visible"]["title"] = bool(visible.get("title", True))
    result["visible"]["table"] = True

    labels = source.get("labels")
    if isinstance(labels, dict):
        for key, fallback in result["labels"].items():
            text = str(labels.get(key, fallback) or "").strip()
            result["labels"][key] = text[:100] or fallback

    numeric_ranges = {
        ("page", "left_margin_cm"): (0.8, 4.0),
        ("page", "right_margin_cm"): (0.8, 4.0),
        ("page", "top_margin_cm"): (0.8, 4.0),
        ("page", "bottom_margin_cm"): (0.8, 4.0),
        ("typography", "body_size"): (8.0, 16.0),
        ("typography", "title_size"): (9.0, 24.0),
        ("typography", "table_size"): (7.0, 14.0),
        ("spacing", "identity_after_cm"): (0.0, 4.0),
        ("spacing", "title_after_cm"): (0.0, 5.0),
        ("spacing", "table_after_cm"): (0.0, 2.0),
        ("table", "row_height_cm"): (0.45, 1.2),
        ("table", "minimum_rows"): (1, 24),
    }
    for (section, key), (minimum, maximum) in numeric_ranges.items():
        section_value = source.get(section)
        if not isinstance(section_value, dict) or key not in section_value:
            continue
        try:
            number = float(section_value[key])
        except (TypeError, ValueError):
            continue
        number = max(minimum, min(maximum, number))
        result[section][key] = int(number) if key == "minimum_rows" else round(number, 2)

    color_pattern = re.compile(r"^#[0-9A-Fa-f]{6}$")
    for section, keys in {
        "typography": ("text_color", "title_color"),
        "table": ("header_background", "total_background", "border_color"),
    }.items():
        section_value = source.get(section)
        if not isinstance(section_value, dict):
            continue
        for key in keys:
            candidate = str(section_value.get(key, "")).strip()
            if color_pattern.fullmatch(candidate):
                result[section][key] = candidate.upper()
    return result


@dataclass
class Intervention:
    annee: int
    mois: int
    mois_libelle: str
    client: str
    date: str
    duree_heures: float
    salaire_net_horaire: float
    montant_net: float
    montant_brut: float
    cesu: str
    adresse: str
    source_classeur: str
    source_onglet: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_name(value: Any) -> str:
    text = normalize_text(value)
    stop = {"m", "mr", "mme", "mlle", "monsieur", "madame", "et"}
    tokens = [token for token in text.split() if token not in stop]
    return " ".join(tokens)


def sorted_name_key(value: Any) -> str:
    return " ".join(sorted(normalize_name(value).split()))


def parse_month_sheet_name(name: str) -> tuple[int, int] | None:
    match = re.match(r"^\s*([A-Za-zÀ-ÿ]+)\s+(\d{4})\s*$", name)
    if not match:
        return None
    month_name = normalize_text(match.group(1))
    month = MONTHS.get(month_name)
    if not month:
        return None
    return int(match.group(2)), month


def as_hours(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, time):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, timedelta):
        return value.total_seconds() / 3600
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if math.isnan(number):
            return None
        return number * 24 if 0 < number < 1 else number
    text = str(value).strip().lower().replace(",", ".")
    match = re.match(r"^(\d+(?:\.\d+)?)\s*h(?:\s*(\d+))?$", text)
    if match:
        return float(match.group(1)) + (float(match.group(2) or 0) / 60)
    match = re.match(r"^(\d+):(\d+)$", text)
    if match:
        return int(match.group(1)) + int(match.group(2)) / 60
    try:
        return float(text)
    except ValueError:
        return None


def as_date(value: Any, fallback: date) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return fallback


def money(value: float) -> float:
    return round(float(value) + 0.000000001, 2)


def french_money(value: float) -> str:
    return f"{money(value):,.2f} €".replace(",", "X").replace(".", ",").replace("X", " ")


def hours_label(hours: float) -> str:
    total_minutes = int(round(hours * 60))
    return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"


def safe_filename(value: str) -> str:
    text = re.sub(r'[<>:"/\\|?*]+', " ", value)
    text = re.sub(r"\s+", " ", text).strip()
    return text.rstrip(".")


def matching_existing_note(folder: Path, year: int, month: int, client: str) -> Path | None:
    if not folder.exists():
        return None
    prefix = f"Note d'intervention - {MONTH_LABELS[month]} {year} - "
    wanted = {normalize_name(client), sorted_name_key(client)}
    for pdf in folder.glob("*.pdf"):
        stem = pdf.stem
        if not stem.startswith(prefix):
            continue
        existing_client = stem[len(prefix) :]
        existing_keys = {normalize_name(existing_client), sorted_name_key(existing_client)}
        if wanted & existing_keys:
            return pdf
    return None


def month_folder(year: int, month: int) -> str:
    return f"{month:02d}. {MONTH_LABELS[month]} {year}"


def load_config(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def read_variables(workbook_path: Path, config: dict[str, Any]) -> tuple[float, float, dict[str, str]]:
    salaire = float(config.get("salaire_net_horaire_defaut", 22.0))
    coeff = float(config.get("coefficient_brut_defaut", 1.2873125))
    cesu_by_key: dict[str, str] = {}

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if "Variables" not in wb.sheetnames:
        return salaire, coeff, cesu_by_key
    ws = wb["Variables"]
    if isinstance(ws["B1"].value, (int, float)):
        salaire = float(ws["B1"].value)
    if isinstance(ws["B2"].value, (int, float)):
        coeff = float(ws["B2"].value)
    for row in ws.iter_rows(min_row=1, values_only=True):
        client = row[2] if len(row) >= 3 else None
        cesu = row[3] if len(row) >= 4 else None
        if client and cesu:
            cesu_by_key[normalize_name(client)] = str(cesu).strip()
            cesu_by_key[sorted_name_key(client)] = str(cesu).strip()
    wb.close()
    return salaire, coeff, cesu_by_key


def read_client_file(path: Path) -> dict[str, dict[str, str]]:
    clients: dict[str, dict[str, str]] = {}
    if not path.exists():
        return clients
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Liste clients" not in wb.sheetnames:
        wb.close()
        return clients
    ws = wb["Liste clients"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cesu, qualite, nom, prenom, adresse, telephone = row[:6]
        if not nom and not prenom:
            continue
        parts = [str(x).strip() for x in [qualite, nom, prenom] if x]
        display = " ".join(parts)
        reverse = " ".join(str(x).strip() for x in [prenom, nom] if x)
        normal = " ".join(str(x).strip() for x in [nom, prenom] if x)
        info = {
            "client_fichier": display,
            "adresse": str(adresse).strip() if adresse else "",
            "telephone": str(telephone).strip() if telephone else "",
            "cesu": str(cesu).strip() if cesu else "",
        }
        for key_source in [display, reverse, normal, f"{nom} {prenom}", f"{prenom} {nom}"]:
            key = normalize_name(key_source)
            if key:
                clients.setdefault(key, info)
            sorted_key = sorted_name_key(key_source)
            if sorted_key:
                clients.setdefault(sorted_key, info)
    wb.close()
    return clients


def lookup_client_info(client: str, client_info: dict[str, dict[str, str]]) -> dict[str, str]:
    return client_info.get(normalize_name(client)) or client_info.get(sorted_name_key(client)) or {}


def parse_interventions(
    workbook_path: Path,
    year_filter: int,
    month_filter: int | None,
    salaire_net: float,
    coeff_brut: float,
    cesu_by_key: dict[str, str],
    client_info: dict[str, dict[str, str]],
) -> tuple[list[Intervention], list[str]]:
    interventions: list[Intervention] = []
    anomalies: list[str] = []
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        parsed = parse_month_sheet_name(ws.title)
        if not parsed:
            continue
        year, month = parsed
        if year != year_filter:
            continue
        if month_filter and month != month_filter:
            continue
        month_start = date(year, month, 1)
        for row_index in range(2, ws.max_row + 1):
            client_cell = ws.cell(row=row_index, column=1).value
            client = str(client_cell).strip() if client_cell else ""
            if not client or normalize_text(client).startswith("totaux"):
                continue
            for col_index in range(2, 33):
                fallback = month_start + timedelta(days=col_index - 2)
                day = as_date(ws.cell(row=1, column=col_index).value, fallback)
                if day.year != year or day.month != month:
                    continue
                hours = as_hours(ws.cell(row=row_index, column=col_index).value)
                if not hours or hours <= 0:
                    continue
                info = lookup_client_info(client, client_info)
                cesu = (
                    cesu_by_key.get(normalize_name(client))
                    or cesu_by_key.get(sorted_name_key(client))
                    or info.get("cesu", "")
                )
                if not cesu:
                    anomalies.append(f"CESU manquant : {client} ({MONTH_LABELS[month]} {year})")
                montant_net = money(hours * salaire_net)
                interventions.append(
                    Intervention(
                        annee=year,
                        mois=month,
                        mois_libelle=f"{MONTH_LABELS[month]} {year}",
                        client=client,
                        date=day.isoformat(),
                        duree_heures=round(hours, 4),
                        salaire_net_horaire=salaire_net,
                        montant_net=montant_net,
                        montant_brut=money(montant_net * coeff_brut),
                        cesu=cesu,
                        adresse=info.get("adresse", ""),
                        source_classeur=str(workbook_path),
                        source_onglet=ws.title,
                    )
                )
    wb.close()
    interventions.sort(key=lambda item: (item.annee, item.mois, item.client, item.date))
    return interventions, sorted(set(anomalies))


def build_summaries(interventions: list[Intervention]) -> dict[str, Any]:
    monthly: dict[tuple[int, int], dict[str, Any]] = {}
    clients: dict[str, dict[str, Any]] = {}
    client_month: dict[tuple[str, int], dict[str, Any]] = {}
    for item in interventions:
        month_key = (item.annee, item.mois)
        monthly.setdefault(
            month_key,
            {
                "annee": item.annee,
                "mois": item.mois,
                "mois_libelle": item.mois_libelle,
                "clients": set(),
                "nb_interventions": 0,
                "heures": 0.0,
                "montant_net": 0.0,
                "montant_brut": 0.0,
            },
        )
        monthly[month_key]["clients"].add(item.client)
        monthly[month_key]["nb_interventions"] += 1
        monthly[month_key]["heures"] += item.duree_heures
        monthly[month_key]["montant_net"] += item.montant_net
        monthly[month_key]["montant_brut"] += item.montant_brut

        clients.setdefault(
            item.client,
            {
                "client": item.client,
                "cesu": item.cesu,
                "adresse": item.adresse,
                "mois_travailles": set(),
                "nb_interventions": 0,
                "heures": 0.0,
                "montant_net": 0.0,
                "montant_brut": 0.0,
                "derniere_intervention": item.date,
            },
        )
        clients[item.client]["mois_travailles"].add(item.mois_libelle)
        clients[item.client]["nb_interventions"] += 1
        clients[item.client]["heures"] += item.duree_heures
        clients[item.client]["montant_net"] += item.montant_net
        clients[item.client]["montant_brut"] += item.montant_brut
        clients[item.client]["derniere_intervention"] = max(clients[item.client]["derniere_intervention"], item.date)
        if item.cesu and not clients[item.client]["cesu"]:
            clients[item.client]["cesu"] = item.cesu
        if item.adresse and not clients[item.client]["adresse"]:
            clients[item.client]["adresse"] = item.adresse

        cm_key = (item.client, item.mois)
        client_month.setdefault(
            cm_key,
            {
                "client": item.client,
                "mois": item.mois,
                "mois_libelle": item.mois_libelle,
                "heures": 0.0,
                "montant_net": 0.0,
                "montant_brut": 0.0,
                "nb_interventions": 0,
            },
        )
        client_month[cm_key]["heures"] += item.duree_heures
        client_month[cm_key]["montant_net"] += item.montant_net
        client_month[cm_key]["montant_brut"] += item.montant_brut
        client_month[cm_key]["nb_interventions"] += 1

    monthly_rows = []
    for _, row in sorted(monthly.items()):
        monthly_rows.append(
            {
                "annee": row["annee"],
                "mois": row["mois"],
                "mois_libelle": row["mois_libelle"],
                "nb_clients": len(row["clients"]),
                "nb_interventions": row["nb_interventions"],
                "heures": round(row["heures"], 4),
                "montant_net": money(row["montant_net"]),
                "montant_brut": money(row["montant_brut"]),
            }
        )
    client_rows = []
    for _, row in sorted(clients.items(), key=lambda x: x[0].lower()):
        client_rows.append(
            {
                "client": row["client"],
                "cesu": row["cesu"],
                "adresse": row["adresse"],
                "nb_mois_travailles": len(row["mois_travailles"]),
                "nb_interventions": row["nb_interventions"],
                "heures": round(row["heures"], 4),
                "montant_net": money(row["montant_net"]),
                "montant_brut": money(row["montant_brut"]),
                "derniere_intervention": row["derniere_intervention"],
            }
        )
    client_month_rows = [
        {
            **row,
            "heures": round(row["heures"], 4),
            "montant_net": money(row["montant_net"]),
            "montant_brut": money(row["montant_brut"]),
        }
        for _, row in sorted(client_month.items(), key=lambda x: (x[0][0].lower(), x[0][1]))
    ]
    return {
        "monthly": monthly_rows,
        "clients": client_rows,
        "client_month": client_month_rows,
        "totals": {
            "nb_clients": len(clients),
            "nb_interventions": len(interventions),
            "heures": round(sum(item.duree_heures for item in interventions), 4),
            "montant_net": money(sum(item.montant_net for item in interventions)),
            "montant_brut": money(sum(item.montant_brut for item in interventions)),
        },
    }


def register_fonts() -> tuple[str, str]:
    for normal_path, bold_path in [
        (Path(r"C:\Windows\Fonts\arial.ttf"), Path(r"C:\Windows\Fonts\arialbd.ttf")),
        (Path(r"C:\Windows\Fonts\calibri.ttf"), Path(r"C:\Windows\Fonts\calibrib.ttf")),
        (Path(r"C:\Windows\Fonts\segoeui.ttf"), Path(r"C:\Windows\Fonts\segoeuib.ttf")),
    ]:
        if normal_path.exists():
            pdfmetrics.registerFont(TTFont("LocalSans", str(normal_path)))
            if bold_path.exists():
                pdfmetrics.registerFont(TTFont("LocalSans-Bold", str(bold_path)))
                pdfmetrics.registerFontFamily("LocalSans", normal="LocalSans", bold="LocalSans-Bold")
                return "LocalSans", "LocalSans-Bold"
            return "LocalSans", "LocalSans"
    return "Helvetica", "Helvetica-Bold"


def employee_block(employee_lines: list[str] | None = None) -> str:
    parts = []
    for index, line in enumerate(employee_lines or EMPLOYEE_LINES):
        if not line:
            parts.append("")
            continue
        text = escape(line)
        parts.append(f"<b>{text}</b>" if index == 0 else text)
    return "<br/>".join(parts)


def generate_note_pdf(
    path: Path,
    client: str,
    month_label: str,
    rows: list[Intervention],
    font_name: str,
    bold_font_name: str,
    employee_lines: list[str] | None = None,
    template_configuration: dict[str, Any] | None = None,
) -> None:
    configuration = normalize_note_template_configuration(template_configuration)
    page = configuration["page"]
    typography = configuration["typography"]
    spacing = configuration["spacing"]
    table_configuration = configuration["table"]
    labels = configuration["labels"]

    path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=page["right_margin_cm"] * cm,
        leftMargin=page["left_margin_cm"] * cm,
        topMargin=page["top_margin_cm"] * cm,
        bottomMargin=page["bottom_margin_cm"] * cm,
    )
    normal = ParagraphStyle(
        "NoteNormal",
        fontName=font_name,
        fontSize=typography["body_size"],
        leading=typography["body_size"] + 3,
        textColor=colors.HexColor(typography["text_color"]),
    )
    label = ParagraphStyle("NoteLabel", parent=normal, fontName=bold_font_name)
    title = ParagraphStyle(
        "NoteTitle",
        parent=label,
        alignment=1,
        fontSize=typography["title_size"],
        leading=typography["title_size"] + 3,
        textColor=colors.HexColor(typography["title_color"]),
    )

    usable_width = A4[0] - doc.leftMargin - doc.rightMargin
    employer = Table(
        [
            [Paragraph(escape(labels["employer"]), label), Paragraph(escape(client), normal)],
            [Paragraph(escape(labels["month"]), label), Paragraph(escape(month_label), normal)],
        ],
        colWidths=[usable_width * 0.2875, usable_width * 0.29375],
    )
    employer.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    header = Table(
        [[Paragraph(employee_block(employee_lines), normal), employer]],
        colWidths=[usable_width * 0.41875, usable_width * 0.58125],
    )
    header.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("LEFTPADDING", (0, 0), (0, 0), 0.15 * cm),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (1, 0), (1, 0), 1.1 * cm),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    data = [[labels["date"], labels["hourly_rate"], labels["hours"], labels["amount"]]]
    total_hours = 0.0
    total_net = 0.0
    for item in sorted(rows, key=lambda r: r.date):
        day = datetime.fromisoformat(item.date).strftime("%d/%m/%y")
        total_hours += item.duree_heures
        total_net += item.montant_net
        data.append(
            [
                day,
                french_money(item.salaire_net_horaire),
                hours_label(item.duree_heures),
                french_money(item.montant_net),
            ]
        )
    minimum_rows = int(table_configuration["minimum_rows"])
    while len(data) < minimum_rows + 1:
        data.append(["", "", "", ""])
    data.append([labels["total"], "", hours_label(total_hours), french_money(total_net)])

    table = Table(
        data,
        colWidths=[usable_width / 4] * 4,
        rowHeights=[0.72 * cm]
        + [table_configuration["row_height_cm"] * cm] * (len(data) - 2)
        + [0.72 * cm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), typography["table_size"]),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(table_configuration["header_background"])),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(table_configuration["total_background"])),
                ("FONTNAME", (0, 0), (-1, 0), bold_font_name),
                ("FONTNAME", (0, -1), (-1, -1), bold_font_name),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (0, -1), (1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(table_configuration["border_color"])),
                ("SPAN", (0, -1), (1, -1)),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    blocks = {
        "identity": header,
        "title": Paragraph(escape(labels["title"]), title),
        "table": table,
    }
    story: list[Any] = []
    for block_name in configuration["blocks"]:
        if not configuration["visible"].get(block_name, True):
            continue
        story.append(blocks[block_name])
        gap = float(spacing.get(f"{block_name}_after_cm", 0))
        if gap > 0:
            story.append(Spacer(1, gap * cm))
    doc.build(story)


def generate_notes(
    interventions: list[Intervention],
    notes_dir: Path,
    overwrite: bool,
    employee_lines: list[str] | None = None,
    template_configuration: dict[str, Any] | None = None,
) -> dict[str, Any]:
    grouped: dict[tuple[int, int, str], list[Intervention]] = defaultdict(list)
    for item in interventions:
        grouped[(item.annee, item.mois, item.client)].append(item)
    font, bold_font = register_fonts()
    result = {"created": [], "skipped": [], "errors": []}
    for (year, month, client), rows in sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1], x[0][2].lower())):
        folder = notes_dir
        filename = f"Note d'intervention - {MONTH_LABELS[month]} {year} - {safe_filename(client)}.pdf"
        output = folder / filename
        existing = matching_existing_note(folder, year, month, client)
        if existing and not overwrite:
            result["skipped"].append(str(existing))
            continue
        if existing and overwrite:
            output = existing
        elif output.exists() and not overwrite:
            result["skipped"].append(str(output))
            continue
        try:
            generate_note_pdf(
                output,
                client,
                f"{MONTH_LABELS[month]} {year}",
                rows,
                font,
                bold_font,
                employee_lines,
                template_configuration,
            )
            result["created"].append(str(output))
        except Exception as exc:  # noqa: BLE001
            result["errors"].append(f"{output}: {exc}")
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Genere les notes d'intervention et les donnees de bilan.")
    parser.add_argument("--config", default="config.json", help="Chemin du fichier config.json")
    parser.add_argument("--year", type=int, required=True, help="Annee a traiter")
    parser.add_argument("--month", type=int, default=0, help="Mois a traiter, optionnel")
    parser.add_argument("--replace", action="store_true", help="Remplace les notes PDF existantes")
    parser.add_argument("--no-pdf", action="store_true", help="Ne genere pas les notes PDF")
    parser.add_argument("--notes-output", default="", help="Dossier de sortie des notes, optionnel")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = root / config_path
    config = load_config(config_path)
    suivi_dir = Path(config["suivi_paye_dir"])
    suivi_name = config.get("suivi_paye_pattern", "Suivi de paye {year}.xlsx").format(year=args.year)
    suivi_path = suivi_dir / suivi_name
    if not suivi_path.exists():
        raise FileNotFoundError(f"Classeur introuvable : {suivi_path}")

    fichier_clients = Path(config.get("fichier_clients", suivi_dir / "Fichier client.xlsx"))
    sorties = root / "sorties"
    sorties.mkdir(parents=True, exist_ok=True)

    salaire_net, coeff_brut, cesu_by_key = read_variables(suivi_path, config)
    client_info = read_client_file(fichier_clients)
    interventions, anomalies = parse_interventions(
        suivi_path,
        args.year,
        args.month or None,
        salaire_net,
        coeff_brut,
        cesu_by_key,
        client_info,
    )
    summaries = build_summaries(interventions)

    notes_result = {"created": [], "skipped": [], "errors": []}
    if not args.no_pdf:
        notes_dir = Path(args.notes_output) if args.notes_output else Path(config["notes_intervention_dir"])
        overwrite = bool(args.replace or config.get("ecraser_notes_existantes", False))
        notes_result = generate_notes(interventions, notes_dir, overwrite)

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "year": args.year,
        "month": args.month or None,
        "sources": {
            "suivi_paye": str(suivi_path),
            "fichier_clients": str(fichier_clients),
            "notes_intervention_dir": str(Path(args.notes_output) if args.notes_output else Path(config["notes_intervention_dir"])),
        },
        "parameters": {
            "salaire_net_horaire": salaire_net,
            "coefficient_brut": coeff_brut,
            "replace_notes": bool(args.replace),
            "pdf_notes_enabled": not args.no_pdf,
        },
        "interventions": [asdict(item) for item in interventions],
        "summaries": summaries,
        "anomalies": anomalies,
        "notes": notes_result,
    }

    json_path = sorties / f"donnees_interventions_{args.year}.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    report_path = sorties / f"rapport_generation_{args.year}.txt"
    report_path.write_text(
        "\n".join(
            [
                f"Generation : {payload['generated_at']}",
                f"Annee : {args.year}",
                f"Interventions : {summaries['totals']['nb_interventions']}",
                f"Clients : {summaries['totals']['nb_clients']}",
                f"Heures : {hours_label(summaries['totals']['heures'])}",
                f"Net : {french_money(summaries['totals']['montant_net'])}",
                f"Notes creees : {len(notes_result['created'])}",
                f"Notes ignorees car deja existantes : {len(notes_result['skipped'])}",
                f"Erreurs notes : {len(notes_result['errors'])}",
                f"Anomalies : {len(anomalies)}",
            ]
        ),
        encoding="utf-8",
    )
    print(f"Donnees ecrites : {json_path}")
    print(f"Rapport ecrit : {report_path}")
    print(f"Interventions : {summaries['totals']['nb_interventions']}")
    print(f"Clients : {summaries['totals']['nb_clients']}")
    print(f"Notes creees : {len(notes_result['created'])}, ignorees : {len(notes_result['skipped'])}")
    if notes_result["errors"]:
        print("Erreurs notes :")
        for error in notes_result["errors"]:
            print(f"- {error}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
