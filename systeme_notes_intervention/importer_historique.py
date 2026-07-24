"""Importe les anciens suivis de paie dans la base Easy CESU, avec contrôle préalable."""

from __future__ import annotations

import argparse
import json
import os
import shutil
import sqlite3
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from generer_notes_et_donnees import as_hours, normalize_name, normalize_text, read_variables


ROOT = Path(__file__).resolve().parent
APP_DATA_ROOT = Path(os.environ.get("LOCALAPPDATA") or Path.home() / "AppData" / "Local") / "EasyCESU"
HISTORY_FILES = (
    ("Paye 2020-2021.xlsx", None),
    # Le suivi d'octobre à décembre 2022 est plus détaillé dans le classeur suivant.
    ("Paye 2022.xlsx", lambda value: value < date(2022, 10, 1)),
    ("Suivi de paye 2022.xlsx", None),
    ("Suivi de paye 2023 Janv Mai.xlsx", None),
    ("Suivi de paye 2023 Juin Dec.xlsx", None),
    ("Suivi de paye 2024.xlsx", None),
    ("Suivi de paye 2025 Janv Mai.xlsx", None),
    ("Suivi de paye 2025 juin Dec.xlsx", None),
)
MONTHS = {
    "janv": 1,
    "janvier": 1,
    "fev": 2,
    "fevrier": 2,
    "mars": 3,
    "avril": 4,
    "mai": 5,
    "juin": 6,
    "juillet": 7,
    "aout": 8,
    "sept": 9,
    "septembre": 9,
    "oct": 10,
    "octobre": 10,
    "nov": 11,
    "novembre": 11,
    "dec": 12,
    "decembre": 12,
}


def active_database() -> Path:
    """Retrouve la base réellement utilisée par le compte actif."""
    config_path = APP_DATA_ROOT / "config" / "config.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    active_id = config.get("active_profile_id")
    profile = next(item for item in config.get("profiles", []) if item.get("id") == active_id)
    raw_file = Path(str(profile.get("data_file") or "interventions.sqlite"))
    if raw_file.is_absolute():
        return raw_file
    data_dir = Path(str(profile.get("data_dir") or APP_DATA_ROOT / "data"))
    return data_dir / raw_file


def sheet_period(title: str) -> tuple[int, int] | None:
    """Accepte les noms d'onglets récents comme les anciens noms abrégés."""
    tokens = normalize_text(title).split()
    if len(tokens) < 2 or tokens[0] not in MONTHS or not tokens[1].isdigit():
        return None
    year = int(tokens[1])
    return (2000 + year if year < 100 else year, MONTHS[tokens[0]])


def history_rows(path: Path) -> Iterable[dict]:
    """Lit les deux modèles de fichiers historiques pris en charge."""
    default_rate, _, _ = read_variables(path, {"salaire_net_horaire_defaut": 22.0, "coefficient_brut_defaut": 1.2873125})
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            expected_period = sheet_period(worksheet.title)
            if not expected_period:
                continue  # Ignore les onglets Base et Variables.
            iterator = worksheet.iter_rows(values_only=True)
            header = next(iterator, ())
            header_names = [normalize_text(value) for value in header[:8]]
            if header_names[:3] == ["date", "client", "duree"]:
                # Ancien format : date, client, durée, CESU, payé, montant.
                for row in iterator:
                    when = row[0] if len(row) > 0 else None
                    client = row[1] if len(row) > 1 else None
                    hours = as_hours(row[2] if len(row) > 2 else None)
                    if not isinstance(when, (date, datetime)) or not client or not hours or hours <= 0:
                        continue
                    day = when.date() if isinstance(when, datetime) else when
                    amount_cell = row[5] if len(row) > 5 else None
                    amount = float(amount_cell) if isinstance(amount_cell, (int, float)) else hours * default_rate
                    yield {
                        "date": day,
                        "client": str(client).strip(),
                        "hours": round(hours, 4),
                        "amount": round(amount, 2),
                        "rate": round(amount / hours, 4),
                        "source": path.name,
                    }
                continue

            # Format récent : une ligne par client et une colonne par date.
            dated_columns = [
                (index, value.date() if isinstance(value, datetime) else value)
                for index, value in enumerate(header)
                if index and isinstance(value, (date, datetime))
            ]
            for row in iterator:
                client = row[0] if row else None
                if not client or normalize_text(client).startswith("totaux"):
                    continue
                for index, day in dated_columns:
                    hours = as_hours(row[index] if index < len(row) else None)
                    if not hours or hours <= 0 or (day.year, day.month) != expected_period:
                        continue
                    yield {
                        "date": day,
                        "client": str(client).strip(),
                        "hours": round(hours, 4),
                        "amount": round(hours * default_rate, 2),
                        "rate": round(default_rate, 4),
                        "source": path.name,
                    }
    finally:
        workbook.close()


def resolve_client_name(source_name: str, existing_names: set[str]) -> str:
    """Rapproche les anciennes formes courtes d'un client sans choix ambigu."""
    normalized = normalize_name(source_name)
    exact = [name for name in existing_names if normalize_name(name) == normalized]
    if len(exact) == 1:
        return exact[0]
    source_tokens = set(normalized.split())
    candidates = []
    for name in existing_names:
        name_tokens = set(normalize_name(name).split())
        if source_tokens and (source_tokens <= name_tokens or name_tokens <= source_tokens):
            candidates.append(name)
    return candidates[0] if len(candidates) == 1 else source_name


def backup_database(database: Path) -> Path:
    """Copie la base avec l'API SQLite afin de garantir une sauvegarde cohérente."""
    target_dir = ROOT / "sauvegardes"
    target_dir.mkdir(exist_ok=True)
    target = target_dir / f"EasyCESU-avant-import-historique-{datetime.now():%Y%m%d_%H%M%S}.sqlite"
    with sqlite3.connect(database) as source, sqlite3.connect(target) as destination:
        source.backup(destination)
    with sqlite3.connect(target) as check:
        if check.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
            raise RuntimeError("La sauvegarde de précaution est invalide.")
    return target


def load_candidates(source_dir: Path) -> list[dict]:
    rows = []
    for filename, predicate in HISTORY_FILES:
        path = source_dir / filename
        if not path.exists():
            raise FileNotFoundError(f"Classeur historique introuvable : {path}")
        for row in history_rows(path):
            if predicate and not predicate(row["date"]):
                continue
            rows.append(row)
    return sorted(rows, key=lambda item: (item["date"], normalize_name(item["client"]), item["hours"]))


def existing_signatures(connection: sqlite3.Connection) -> set[tuple[str, str, float, float]]:
    return {
        (row["date"], normalize_name(row["client"]), round(float(row["duration_hours"]), 4), round(float(row["hourly_rate"]), 4))
        for row in connection.execute("SELECT date, client, duration_hours, hourly_rate FROM interventions")
    }


def import_history(source_dir: Path, apply: bool) -> dict:
    database = active_database()
    if not database.exists():
        raise FileNotFoundError(f"Base Easy CESU introuvable : {database}")
    candidates = load_candidates(source_dir)
    with sqlite3.connect(database) as connection:
        connection.row_factory = sqlite3.Row
        existing_clients = {row["name"] for row in connection.execute("SELECT name FROM clients")}
        signatures = existing_signatures(connection)
        to_insert = []
        for row in candidates:
            client = resolve_client_name(row["client"], existing_clients)
            signature = (row["date"].isoformat(), normalize_name(client), row["hours"], row["rate"])
            if signature in signatures:
                continue
            to_insert.append({**row, "client": client})
            existing_clients.add(client)
            signatures.add(signature)

        summary = {
            "database": str(database),
            "candidates": len(candidates),
            "new_interventions": len(to_insert),
            "years": dict(sorted(Counter(item["date"].year for item in to_insert).items())),
            "hours": round(sum(item["hours"] for item in to_insert), 2),
            "amount_net": round(sum(item["amount"] for item in to_insert), 2),
        }
        if not apply:
            return summary

        backup = backup_database(database)
        stamp = datetime.now().isoformat(timespec="seconds")
        for row in to_insert:
            connection.execute(
                """
                INSERT INTO clients (name, cesu, email, hourly_rate, hourly_rate_custom, address, phone, updated_at)
                VALUES (?, '', '', 0, 0, '', '', ?)
                ON CONFLICT(name) DO NOTHING
                """,
                (row["client"], stamp),
            )
            connection.execute(
                """
                INSERT INTO interventions
                    (date, client, duration_hours, hourly_rate, task, location, transmitted, paid, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, '', 0, 0, ?, ?)
                """,
                (row["date"].isoformat(), row["client"], row["hours"], row["rate"], f"Import historique - {row['source']}", stamp, stamp),
            )
        connection.commit()
        if connection.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
            raise RuntimeError("La base est invalide après l'import historique.")
        return {**summary, "backup": str(backup)}


def main() -> int:
    parser = argparse.ArgumentParser(description="Import historique Easy CESU")
    parser.add_argument("source_dir", type=Path, help="Dossier contenant les classeurs de suivi de paye")
    parser.add_argument("--apply", action="store_true", help="Écrit les interventions après le contrôle")
    args = parser.parse_args()
    print(json.dumps(import_history(args.source_dir, args.apply), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
